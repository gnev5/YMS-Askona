from datetime import timedelta
from io import BytesIO
from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
from fastapi.security import OAuth2PasswordRequestForm
from sqlalchemy.orm import Session
from fastapi.responses import StreamingResponse

from ..db import get_db
from .. import models, schemas
from ..security import get_password_hash, verify_password, create_access_token, ACCESS_TOKEN_EXPIRE_MINUTES
from ..deps import get_current_user, get_current_admin
from typing import List
from openpyxl import Workbook, load_workbook

router = APIRouter()


@router.post("/register", response_model=schemas.User, status_code=status.HTTP_201_CREATED)
def register(user_in: schemas.UserCreate, db: Session = Depends(get_db)):
    exists = db.query(models.User).filter(models.User.email == user_in.email).first()
    if exists:
        raise HTTPException(status_code=400, detail="User already exists")

    try:
        role_value = user_in.role or models.UserRole.carrier.value
        role = models.UserRole(role_value)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid role value")

    new_user = models.User(
        email=user_in.email,
        full_name=user_in.full_name,
        password_hash=get_password_hash(user_in.password),
        role=role,
    )
    db.add(new_user)
    db.commit()
    db.refresh(new_user)
    return new_user


@router.post("/login", response_model=schemas.Token)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    # debug trace
    print("/auth/login: start for", form_data.username)
    user = db.query(models.User).filter(models.User.email == form_data.username).first()
    print("/auth/login: queried user ->", bool(user))
    if not user or not verify_password(form_data.password, user.password_hash):
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Incorrect email or password")
    print("/auth/login: password verified")
    access_token = create_access_token(subject=user.email, expires_delta=timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES))
    print("/auth/login: token issued")
    return {"access_token": access_token, "token_type": "bearer"}


@router.get("/me", response_model=schemas.User)
def me(current_user: models.User = Depends(get_current_user)):
    return current_user

@router.post("/change-password")
def change_password(
    payload: schemas.PasswordChangeRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    if not verify_password(payload.current_password, current_user.password_hash):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid current password")
    if len(payload.new_password) < 8:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="New password must be at least 8 characters")
    if verify_password(payload.new_password, current_user.password_hash):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="New password must differ from current password")

    current_user.password_hash = get_password_hash(payload.new_password)
    db.add(current_user)
    db.commit()
    return {"message": "Password updated. Please log in again."}

# Admin: users management
@router.get("/users", response_model=List[schemas.User])
def list_users(db: Session = Depends(get_db), _: models.User = Depends(get_current_admin)):
    return db.query(models.User).all()

@router.post("/users", response_model=schemas.User, status_code=status.HTTP_201_CREATED)
def create_user_admin(user_in: schemas.UserCreate, db: Session = Depends(get_db), _: models.User = Depends(get_current_admin)):
    exists = db.query(models.User).filter(models.User.email == user_in.email).first()
    if exists:
        raise HTTPException(status_code=400, detail="User already exists")
    try:
        role = models.UserRole(user_in.role) if user_in.role else models.UserRole.carrier
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid role value")
    user = models.User(
        email=user_in.email,
        full_name=user_in.full_name,
        password_hash=get_password_hash(user_in.password),
        role=role,
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    return user

@router.put("/users/{user_id}", response_model=schemas.User)
def update_user_admin(user_id: int, payload: schemas.UserUpdate, db: Session = Depends(get_db), _: models.User = Depends(get_current_admin)):
    user = db.query(models.User).get(user_id)
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    if payload.full_name is not None:
        user.full_name = payload.full_name
    if payload.role is not None:
        try:
            user.role = models.UserRole(payload.role)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid role value")
    if payload.is_active is not None:
        user.is_active = payload.is_active
    if payload.password:
        user.password_hash = get_password_hash(payload.password)
    db.commit()
    db.refresh(user)
    return user

@router.delete("/users/{user_id}")
def delete_user_admin(user_id: int, db: Session = Depends(get_db), _: models.User = Depends(get_current_admin)):
    user = db.query(models.User).get(user_id)
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    db.delete(user)
    db.commit()
    return {"message": "User deleted"}


@router.get("/users/template")
def download_user_template(_: models.User = Depends(get_current_admin)):
    """Provide Excel template for bulk user creation."""
    wb = Workbook()
    ws = wb.active
    ws.title = "users"
    headers = ["email", "password", "full_name", "role"]
    ws.append(headers)
    # sample rows for clarity
    ws.append(["user@example.com", "Password123!", "Иван Иванов", "carrier"])
    ws.append(["admin@yms.local", "Admin1234!", "Администратор", "admin"])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = "users_template.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.post("/users/import")
def import_users_from_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: models.User = Depends(get_current_admin),
):
    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Invalid file format. Please upload an .xlsx file")

    try:
        wb = load_workbook(filename=BytesIO(file.file.read()))
    except Exception:
        raise HTTPException(status_code=400, detail="Failed to read Excel file")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Excel file is empty")

    headers = [str(h).strip().lower() if h else "" for h in rows[0]]
    expected = ["email", "password", "full_name", "role"]
    if headers[: len(expected)] != expected:
        raise HTTPException(status_code=400, detail=f"Invalid headers. Expected {expected}")

    created = 0
    skipped = 0
    errors: List[str] = []

    for idx, row in enumerate(rows[1:], start=2):
        email, password, full_name, role_raw = row[:4]
        if not email or not password or not full_name:
            skipped += 1
            errors.append(f"Row {idx}: missing required fields")
            continue
        try:
            role = models.UserRole(str(role_raw).lower()) if role_raw else models.UserRole.carrier
        except ValueError:
            skipped += 1
            errors.append(f"Row {idx}: invalid role '{role_raw}'")
            continue

        existing = db.query(models.User).filter(models.User.email == email).first()
        if existing:
            skipped += 1
            errors.append(f"Row {idx}: user {email} already exists")
            continue

        user = models.User(
            email=email,
            full_name=full_name,
            password_hash=get_password_hash(str(password)),
            role=role,
        )
        db.add(user)
        created += 1

    db.commit()
    return {"created": created, "skipped": skipped, "errors": errors}
