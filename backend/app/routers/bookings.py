ÔĽŅfrom fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Body
from fastapi.responses import StreamingResponse
from dataclasses import dataclass, field
from fastapi import Query
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func, and_
from typing import List, Optional
from collections import defaultdict
from datetime import date, datetime, timedelta, time, timezone
import uuid
import logging
from .. import models, schemas
from ..db import get_db
from ..deps import get_current_user
from .prr_limits import get_duration
from ..quota_utils import calculate_used_volume, get_quota_for_date
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from io import BytesIO

router = APIRouter()
MSK_TZ = timezone(timedelta(hours=3))
NOON_MSK = time(15, 0)
EXPORT_RED_FILL = PatternFill(fill_type="solid", fgColor="FFFEE2E2")
EXPORT_ORANGE_FILL = PatternFill(fill_type="solid", fgColor="FFFED7AA")
EXPORT_YELLOW_FILL = PatternFill(fill_type="solid", fgColor="FFFEF9C3")
EXPORT_SUMMARY_HEADER_FILL = PatternFill(fill_type="solid", fgColor="FFE0F2FE")
OWN_PRODUCTION_REPORT_ROWS = [
    "–õ–ĺ–Ņ–į—ā–ł–Ĺ–į",
    "–ü–ĺ—á–į–Ķ–≤—Ā–ļ–ł–Ļ",
    "–°–ĺ–Ľ–≤–ł—Ā",
    "–°–ĺ—Ą—ā –°–Ľ–ł–Ņ",
    "–°–ĺ—Ü–ł–į–Ľ–ł—Ā—ā–ł—á–Ķ—Ā–ļ–į—Ź",
    "–¶–†–°–ď–ü",
]
OWN_PRODUCTION_FIXED_CUBES = {
    "–õ–ĺ–Ņ–į—ā–ł–Ĺ–į": 38,
    "–ü–ĺ—á–į–Ķ–≤—Ā–ļ–ł–Ļ": 64,
    "–°–ĺ–Ľ–≤–ł—Ā": 31,
    "–¶–†–°–ď–ü": 64,
}


def _normalize_report_text(value: str | None) -> str:
    return (value or "").strip().lower().replace("—Ď", "–Ķ")


def _resolve_own_production_report_direction(supplier_name: str | None) -> str | None:
    normalized = _normalize_report_text(supplier_name)
    if "–Ľ–ĺ–Ņ–į—ā–ł–Ĺ–į" in normalized and "–Ľ–ł—ā–≤—É–ī" in normalized:
        return "–õ–ĺ–Ņ–į—ā–ł–Ĺ–į"
    if "–Ņ–ĺ—á–į–Ļ–ļ" in normalized or "–Ņ–ĺ—á–į–Ķ–≤" in normalized:
        return "–ü–ĺ—á–į–Ķ–≤—Ā–ļ–ł–Ļ"
    if "—Ā–ĺ–Ľ–≤–ł—Ā" in normalized:
        return "–°–ĺ–Ľ–≤–ł—Ā"
    if "—Ā–ĺ—Ą—ā —Ā–Ľ–ł–Ņ" in normalized and ("–ļ–ĺ–≤—Ä–ĺ–≤" in normalized or "–≤–Ľ–į–ī–ł–ľ–ł—Ä" in normalized):
        return "–°–ĺ—Ą—ā –°–Ľ–ł–Ņ"
    if "—Ā–ĺ—Ü–ł–į–Ľ—Ć—Ā—ā" in normalized or "—Ā–ĺ—Ü–ł–į–Ľ–ł—Ā—ā" in normalized:
        return "–°–ĺ—Ü–ł–į–Ľ–ł—Ā—ā–ł—á–Ķ—Ā–ļ–į—Ź"
    if "—Ü—Ä—Ā–≥–Ņ" in normalized:
        return "–¶–†–°–ď–ü"
    return None


def _is_purchased_report_row(transport_type_name: str | None) -> bool:
    return _normalize_report_text(transport_type_name) == "–∑–į–ļ—É–Ņ–Ĺ–į—Ź"


def _is_own_production_report_row(transport_type_name: str | None) -> bool:
    normalized = _normalize_report_text(transport_type_name)
    return normalized in {"—Ā–ĺ–Ī—Ā—ā–≤–Ķ–Ĺ–Ĺ–ĺ–Ķ –Ņ—Ä–ĺ–ł–∑–≤–ĺ–ī—Ā—ā–≤–ĺ", "—Ā–ĺ–Ī—Ā—ā–≤–Ķ–Ĺ–Ĺ–ĺ–Ķ"}


def _parse_report_booking_date(serialized: dict) -> date | None:
    raw_date = serialized.get("booking_date")
    if not raw_date:
        return None
    try:
        booking_date = datetime.strptime(str(raw_date)[:10], "%Y-%m-%d").date()
    except ValueError:
        return None

    raw_start = (serialized.get("start_time") or "")[:5]
    try:
        start_time = datetime.strptime(raw_start, "%H:%M").time()
    except ValueError:
        start_time = time(0, 0)

    if start_time >= time(22, 0):
        return booking_date + timedelta(days=1)
    return booking_date


def _report_number(value) -> float:
    if value is None or value == "":
        return 0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0


def _report_cell_number(value: float):
    return int(value) if float(value).is_integer() else value


def _append_report_matrix_table(ws, column_labels: list[str], values_by_date_and_column: dict[date, dict[str, float]]) -> None:
    ordered_dates = sorted(values_by_date_and_column)
    header = ["–Ē–į—ā–į", *column_labels, "–ě–Ī—Č–ł–Ļ –ł—ā–ĺ–≥"]
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = EXPORT_SUMMARY_HEADER_FILL

    grand_total_by_column = defaultdict(float)
    for report_date in ordered_dates:
        values_by_column = values_by_date_and_column[report_date]
        row_values = [_report_cell_number(values_by_column.get(label, 0)) for label in column_labels]
        row_total = sum(values_by_column.values())
        for label, value in values_by_column.items():
            grand_total_by_column[label] += value
        ws.append([
            report_date.isoformat(),
            *row_values,
            _report_cell_number(row_total),
        ])

    ws.append([
        "–ě–Ī—Č–ł–Ļ –ł—ā–ĺ–≥",
        *[_report_cell_number(grand_total_by_column.get(label, 0)) for label in column_labels],
        _report_cell_number(sum(grand_total_by_column.values())),
    ])

    ws.freeze_panes = "B2"
    ws.column_dimensions["A"].width = 14
    for idx in range(2, len(header) + 1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = 18


def _append_bookings_report_sheets(wb: Workbook, serialized_rows: list[dict]) -> None:
    report_columns = [*OWN_PRODUCTION_REPORT_ROWS, "–ó–į–ļ—É–Ņ–Ĺ–į—Ź"]
    values_by_date_and_column: dict[date, dict[str, float]] = defaultdict(lambda: defaultdict(float))

    for serialized in serialized_rows:
        report_date = _parse_report_booking_date(serialized)
        if report_date is None:
            continue

        transport_type_name = serialized.get("transport_type_name")
        supplier_name = serialized.get("supplier_name")
        direction = _resolve_own_production_report_direction(supplier_name)

        if direction is not None and _is_own_production_report_row(transport_type_name):
            cubes = OWN_PRODUCTION_FIXED_CUBES.get(direction)
            if cubes is None:
                cubes = _report_number(serialized.get("cubes"))
            values_by_date_and_column[report_date][direction] += cubes
            continue

        if _is_purchased_report_row(transport_type_name):
            values_by_date_and_column[report_date]["–ó–į–ļ—É–Ņ–Ĺ–į—Ź"] += _report_number(serialized.get("cubes"))

    own_ws = wb.create_sheet(title="–°–ĺ–Ī—Ā—ā–≤–Ķ–Ĺ–Ĺ–ĺ–Ķ –Ņ—Ä–ĺ–ł–∑–≤–ĺ–ī—Ā—ā–≤–ĺ")
    _append_report_matrix_table(own_ws, report_columns, values_by_date_and_column)


@dataclass
class BookingListParams:
    page: int = 1
    page_size: int = 50
    supplier: Optional[str] = None
    zone: Optional[str] = None
    transport_type: Optional[str] = None
    vehicle_plate: Optional[str] = None
    driver_name: Optional[str] = None
    transport_sheet: Optional[str] = None
    booking_type: Optional[str] = None
    user_email: Optional[str] = None
    object_ids: list[int] = field(default_factory=list)
    date_from: Optional[date] = None
    date_to: Optional[date] = None
    only_owner: bool = False


def get_booking_list_params(
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    supplier: Optional[str] = Query(None),
    zone: Optional[str] = Query(None),
    transport_type: Optional[str] = Query(None),
    vehicle_plate: Optional[str] = Query(None),
    driver_name: Optional[str] = Query(None),
    transport_sheet: Optional[str] = Query(None),
    booking_type: Optional[str] = Query(None, description="in|out"),
    user_email: Optional[str] = Query(None),
    object_id: Optional[List[int]] = Query(None),
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    only_owner: bool = Query(False),
) -> BookingListParams:
    return BookingListParams(
        page=page,
        page_size=page_size,
        supplier=supplier,
        zone=zone,
        transport_type=transport_type,
        vehicle_plate=vehicle_plate,
        driver_name=driver_name,
        transport_sheet=transport_sheet,
        booking_type=booking_type,
        user_email=user_email,
        object_ids=object_id or [],
        date_from=date_from,
        date_to=date_to,
        only_owner=only_owner,
    )

def _dock_matches_supplier_zone(dock: models.Dock, supplier_zone_id: int | None) -> bool:
    if supplier_zone_id is None:
        return True
    if not dock.available_zones:
        return True
    return any(zone.id == supplier_zone_id for zone in dock.available_zones)


def _to_msk(created_at: datetime) -> datetime:
    if created_at.tzinfo is None:
        created_utc = created_at.replace(tzinfo=timezone.utc)
    else:
        created_utc = created_at.astimezone(timezone.utc)
    return created_utc.astimezone(MSK_TZ)


def _is_after_noon_for_next_day_msk(created_at: datetime, booking_date) -> bool:
    created_msk = _to_msk(created_at)
    return created_msk.time() > NOON_MSK and booking_date == (created_msk.date() + timedelta(days=1))


def _is_post_factum_by_start_msk(created_at: datetime, booking_date, slot_start_time: time) -> bool:
    created_msk = _to_msk(created_at)
    slot_start_msk = datetime.combine(booking_date, slot_start_time).replace(tzinfo=MSK_TZ)
    return created_msk > slot_start_msk


def _is_created_today_for_today_msk(created_at: datetime, booking_date) -> bool:
    created_msk = _to_msk(created_at)
    return created_msk.date() == booking_date


def _normalize_search(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    cleaned = value.strip()
    return cleaned or None


def _build_first_slot_subquery(db: Session):
    ranked_slots = (
        db.query(
            models.BookingTimeSlot.booking_id.label("booking_id"),
            models.TimeSlot.slot_date.label("slot_date"),
            models.TimeSlot.start_time.label("start_time"),
            models.TimeSlot.dock_id.label("dock_id"),
            func.row_number().over(
                partition_by=models.BookingTimeSlot.booking_id,
                order_by=(
                    models.TimeSlot.slot_date.asc(),
                    models.TimeSlot.start_time.asc(),
                    models.TimeSlot.id.asc(),
                ),
            ).label("rn"),
        )
        .join(models.TimeSlot, models.TimeSlot.id == models.BookingTimeSlot.time_slot_id)
        .subquery()
    )

    return (
        db.query(
            ranked_slots.c.booking_id.label("booking_id"),
            ranked_slots.c.slot_date.label("slot_date"),
            ranked_slots.c.start_time.label("start_time"),
            ranked_slots.c.dock_id.label("dock_id"),
        )
        .filter(ranked_slots.c.rn == 1)
        .subquery()
    )


def _build_booking_listing_query(db: Session, params: BookingListParams, current_user: models.User):
    if params.date_from and params.date_to and params.date_from > params.date_to:
        raise HTTPException(status_code=400, detail="date_from must be before or equal to date_to")

    first_slot_subquery = _build_first_slot_subquery(db)

    query = (
        db.query(
            models.Booking.id.label("booking_id"),
            first_slot_subquery.c.slot_date.label("booking_date"),
            first_slot_subquery.c.start_time.label("booking_start_time"),
        )
        .join(first_slot_subquery, first_slot_subquery.c.booking_id == models.Booking.id)
        .join(models.Dock, models.Dock.id == first_slot_subquery.c.dock_id)
        .outerjoin(models.Supplier, models.Supplier.id == models.Booking.supplier_id)
        .outerjoin(models.Zone, models.Zone.id == models.Booking.zone_id)
        .outerjoin(models.TransportTypeRef, models.TransportTypeRef.id == models.Booking.transport_type_id)
        .outerjoin(models.User, models.User.id == models.Booking.user_id)
        .filter(models.Booking.status == "confirmed")
    )

    if params.only_owner:
        query = query.filter(models.Booking.user_id == current_user.id)
    if params.date_from:
        query = query.filter(first_slot_subquery.c.slot_date >= params.date_from)
    if params.date_to:
        query = query.filter(first_slot_subquery.c.slot_date <= params.date_to)
    if params.object_ids:
        query = query.filter(models.Dock.object_id.in_(params.object_ids))

    supplier = _normalize_search(params.supplier)
    if supplier:
        query = query.filter(func.lower(func.coalesce(models.Supplier.name, "")).like(f"%{supplier.lower()}%"))

    zone = _normalize_search(params.zone)
    if zone:
        query = query.filter(func.lower(func.coalesce(models.Zone.name, "")).like(f"%{zone.lower()}%"))

    transport_type = _normalize_search(params.transport_type)
    if transport_type:
        query = query.filter(
            func.lower(func.coalesce(models.TransportTypeRef.name, "")).like(f"%{transport_type.lower()}%")
        )

    vehicle_plate = _normalize_search(params.vehicle_plate)
    if vehicle_plate:
        query = query.filter(
            func.lower(func.coalesce(models.Booking.vehicle_plate, "")).like(f"%{vehicle_plate.lower()}%")
        )

    driver_name = _normalize_search(params.driver_name)
    if driver_name:
        query = query.filter(
            func.lower(func.coalesce(models.Booking.driver_full_name, "")).like(f"%{driver_name.lower()}%")
        )

    transport_sheet = _normalize_search(params.transport_sheet)
    if transport_sheet:
        query = query.filter(
            func.lower(func.coalesce(models.Booking.transport_sheet, "")).like(f"%{transport_sheet.lower()}%")
        )

    if params.booking_type:
        try:
            booking_direction = models.BookingDirection(params.booking_type)
        except ValueError:
            raise HTTPException(status_code=400, detail="booking_type must be 'in' or 'out'")
        query = query.filter(models.Booking.booking_type == booking_direction)

    user_email = _normalize_search(params.user_email)
    if user_email and current_user.role == models.UserRole.admin:
        query = query.filter(
            (func.lower(func.coalesce(models.User.email, "")).like(f"%{user_email.lower()}%")) |
            (func.lower(func.coalesce(models.User.full_name, "")).like(f"%{user_email.lower()}%"))
        )

    return query.order_by(
        first_slot_subquery.c.slot_date.desc(),
        first_slot_subquery.c.start_time.desc(),
        models.Booking.id.desc(),
    )


def _serialize_bookings_bulk(
    db: Session,
    bookings: List[models.Booking],
    include_user: bool = False,
) -> dict[int, dict]:
    if not bookings:
        return {}

    booking_ids = [booking.id for booking in bookings]
    slot_rows = (
        db.query(
            models.BookingTimeSlot.booking_id.label("booking_id"),
            models.TimeSlot.slot_date.label("slot_date"),
            models.TimeSlot.start_time.label("start_time"),
            models.TimeSlot.end_time.label("end_time"),
            models.Dock.name.label("dock_name"),
            models.Dock.object_id.label("object_id"),
            models.Object.name.label("object_name"),
        )
        .join(models.TimeSlot, models.TimeSlot.id == models.BookingTimeSlot.time_slot_id)
        .join(models.Dock, models.Dock.id == models.TimeSlot.dock_id)
        .outerjoin(models.Object, models.Object.id == models.Dock.object_id)
        .filter(models.BookingTimeSlot.booking_id.in_(booking_ids))
        .order_by(
            models.BookingTimeSlot.booking_id,
            models.TimeSlot.slot_date,
            models.TimeSlot.start_time,
            models.TimeSlot.id,
        )
        .all()
    )

    rows_by_booking: dict[int, list] = {}
    for row in slot_rows:
        rows_by_booking.setdefault(row.booking_id, []).append(row)

    serialized: dict[int, dict] = {}
    for booking in bookings:
        rows = rows_by_booking.get(booking.id)
        if not rows:
            logging.warning(f"No slots found for booking {booking.id}")
            continue

        first_row = rows[0]
        last_row = rows[-1]
        end_date = last_row.slot_date + timedelta(days=1) if last_row.end_time <= last_row.start_time else last_row.slot_date

        data = {
            "id": booking.id,
            "booking_date": first_row.slot_date.isoformat(),
            "start_time": first_row.start_time.strftime("%H:%M:%S"),
            "end_date": end_date.isoformat(),
            "end_time": last_row.end_time.strftime("%H:%M:%S"),
            "user_id": booking.user_id,
            "vehicle_plate": booking.vehicle_plate or "",
            "driver_name": booking.driver_full_name or "",
            "driver_full_name": booking.driver_full_name or "",
            "driver_phone": booking.driver_phone or "",
            "vehicle_type_name": booking.vehicle_type.name if booking.vehicle_type else "",
            "dock_name": first_row.dock_name or "",
            "status": booking.status,
            "slots_count": len(rows),
            "created_at": booking.created_at.isoformat(),
            "supplier_name": booking.supplier.name if booking.supplier else None,
            "zone_name": booking.zone.name if booking.zone else None,
            "transport_type_name": booking.transport_type.name if booking.transport_type else None,
            "cubes": booking.cubes,
            "transport_sheet": booking.transport_sheet,
            "object_id": first_row.object_id,
            "object_name": first_row.object_name,
            "booking_type": booking.booking_type.value if getattr(booking, "booking_type", None) else None,
            "is_after_noon_for_next_day_msk": _is_after_noon_for_next_day_msk(booking.created_at, first_row.slot_date),
            "is_post_factum_msk": _is_post_factum_by_start_msk(
                booking.created_at,
                first_row.slot_date,
                first_row.start_time,
            ),
            "is_created_today_for_today_msk": _is_created_today_for_today_msk(
                booking.created_at,
                first_row.slot_date,
            ),
        }

        if include_user and booking.user:
            data["user_email"] = booking.user.email
            data["user_login"] = booking.user.email
            data["user_full_name"] = booking.user.full_name

        serialized[booking.id] = data

    return serialized


def _serialize_booking(db: Session, booking: models.Booking, include_user: bool = False):
    return _serialize_bookings_bulk(db, [booking], include_user=include_user).get(booking.id)


def _build_paginated_bookings_response(
    db: Session,
    current_user: models.User,
    params: BookingListParams,
):
    base_query = _build_booking_listing_query(db, params, current_user)
    total = base_query.order_by(None).count()
    total_pages = (total + params.page_size - 1) // params.page_size if total > 0 else 0
    current_page = min(params.page, total_pages) if total_pages > 0 else 1
    offset = (current_page - 1) * params.page_size

    page_rows = base_query.offset(offset).limit(params.page_size).all()
    booking_ids = [row.booking_id for row in page_rows]

    bookings = (
        db.query(models.Booking)
        .options(
            joinedload(models.Booking.user),
            joinedload(models.Booking.vehicle_type),
            joinedload(models.Booking.supplier),
            joinedload(models.Booking.zone),
            joinedload(models.Booking.transport_type),
        )
        .filter(models.Booking.id.in_(booking_ids))
        .all()
    ) if booking_ids else []

    booking_by_id = {booking.id: booking for booking in bookings}
    serialized_by_id = _serialize_bookings_bulk(db, bookings, include_user=True)

    items = []
    for booking_id in booking_ids:
        booking = booking_by_id.get(booking_id)
        serialized = serialized_by_id.get(booking_id)
        if not booking or not serialized:
            continue
        is_owner = booking.user_id == current_user.id
        serialized["is_owner"] = is_owner
        serialized["can_modify"] = is_owner or current_user.role == models.UserRole.admin
        items.append(serialized)

    return {
        "items": items,
        "total": total,
        "page": current_page,
        "page_size": params.page_size,
        "total_pages": total_pages,
    }


def _resolve_export_row_fill(serialized: dict) -> PatternFill | None:
    # Keep same priority as UI: red > orange > yellow.
    if bool(serialized.get("is_post_factum_msk")):
        return EXPORT_RED_FILL
    if bool(serialized.get("is_created_today_for_today_msk")):
        return EXPORT_ORANGE_FILL
    if bool(serialized.get("is_after_noon_for_next_day_msk")):
        return EXPORT_YELLOW_FILL
    return None


def _format_export_date(value: date | datetime | str | None) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, date):
        return value.strftime("%d.%m.%Y")
    try:
        return datetime.fromisoformat(value).strftime("%d.%m.%Y")
    except ValueError:
        try:
            return date.fromisoformat(value).strftime("%d.%m.%Y")
        except ValueError:
            return value

@router.post("/", response_model=schemas.Booking)
def create_booking(booking: schemas.BookingCreateUpdated, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    """–†–é–†—ē–†¬∑–†“Ď–†¬į–†–Ö–†—Ď–†¬Ķ –†–Ö–†—ē–†–Ü–†—ē–†‚ĄĖ –†¬∑–†¬į–†—ó–†—Ď–°–É–†—Ď –†–Ö–†¬į –†—ü–†¬†–†¬† (–†—ē–†¬Ī–†–Ö–†—ē–†–Ü–†¬Ľ–†¬Ķ–†–Ö–†–Ö–†¬į–°–Ź –†–Ü–†¬Ķ–°–ā–°–É–†—Ď–°–Ź)"""
    logging.info(f"--- create_booking START for user {current_user.id} ---")
    logging.info(f"Received booking data: {booking.dict()}")

    # –†‚Äô–†¬į–†¬Ľ–†—Ď–†“Ď–†¬į–°‚Ä†–†—Ď–°–Ź –°‚Äö–†—Ď–†—ó–†¬į –°‚Äö–°–ā–†¬į–†–Ö–°–É–†—ó–†—ē–°–ā–°‚Äö–†¬į
    vehicle_type = db.query(models.VehicleType).filter(models.VehicleType.id == booking.vehicle_type_id).first()
    if not vehicle_type:
        raise HTTPException(status_code=404, detail="Vehicle type not found")
    
    obj = db.query(models.Object).filter(models.Object.id == booking.object_id).first()
    if not obj:
        raise HTTPException(status_code=404, detail="Object not found")

    supplier_zone_id: int | None = None
    if booking.supplier_id:
        supplier = db.query(models.Supplier).options(
            joinedload(models.Supplier.vehicle_types),
            joinedload(models.Supplier.zone),
        ).filter(models.Supplier.id == booking.supplier_id).first()
        if not supplier:
            raise HTTPException(status_code=404, detail="Supplier not found")
        supplier_zone_id = supplier.zone_id

        if booking.zone_id is not None and booking.zone_id != supplier_zone_id:
            raise HTTPException(status_code=400, detail="Booking zone must match supplier zone")

        if supplier.vehicle_types:
            allowed_ids = {vt.id for vt in supplier.vehicle_types}
            if booking.vehicle_type_id not in allowed_ids:
                raise HTTPException(status_code=400, detail="Selected vehicle type is not allowed for this supplier")
    
    try:
        duration = get_duration(
            object_id=booking.object_id,
            supplier_id=booking.supplier_id,
            transport_type_id=booking.transport_type_id,
            vehicle_type_id=booking.vehicle_type_id,
            db=db
        ).duration_minutes
    except HTTPException as e:
        if e.status_code == 404:
            duration = vehicle_type.duration_minutes
        else:
            raise e

    if duration <= 0:
        raise HTTPException(status_code=400, detail="Invalid duration")
    
    # –†‚Äô–°‚ÄĻ–°‚Ä°–†—Ď–°–É–†¬Ľ–°–Ź–†¬Ķ–†—ė –°‚Äö–°–ā–†¬Ķ–†¬Ī–°—ď–†¬Ķ–†—ė–†—ē–†¬Ķ –†—Ē–†—ē–†¬Ľ–†—Ď–°‚Ä°–†¬Ķ–°–É–°‚Äö–†–Ü–†—ē –°–É–†¬Ľ–†—ē–°‚Äö–†—ē–†–Ü
    required_slots = duration // 30 + (1 if duration % 30 != 0 else 0)
    logging.info(f"Calculated duration: {duration} mins, required_slots: {required_slots}")
    
    # –†—ü–†¬į–°–ā–°–É–†—Ď–†—ė –†“Ď–†¬į–°‚Äö–°—ď –†—Ď –†–Ü–°–ā–†¬Ķ–†—ė–°–Ź –†–Ö–†¬į–°‚Ä°–†¬į–†¬Ľ–†¬į
    booking_date = datetime.strptime(booking.booking_date, "%Y-%m-%d").date()
    start_time = datetime.strptime(booking.start_time, "%H:%M").time()
    next_date = booking_date + timedelta(days=1)  # –†“Ď–†—ē–†—ó–°—ď–°–É–†—Ē–†¬į–†¬Ķ–†—ė –†—ó–†¬Ķ–°–ā–†¬Ķ–°‚Äö–†¬Ķ–†—Ē–†¬į–†–Ö–†—Ď–†¬Ķ –†–Ö–†¬į –°–É–†¬Ľ–†¬Ķ–†“Ď–°—ď–°–č–°‚Äį–†—Ď–†‚ĄĖ –†“Ď–†¬Ķ–†–Ö–°–ä
    logging.info(f"Parsed booking_date: {booking_date}, start_time: {start_time}")

    try:
        booking_direction = models.BookingDirection(booking.booking_type or "in")
    except Exception:
        raise HTTPException(status_code=400, detail="booking_type must be 'in' or 'out'")

    # –†–é–†–Ö–†¬į–°‚Ä°–†¬į–†¬Ľ–†¬į –†—ó–°–ā–†—ē–†–Ü–†¬Ķ–°–ā–†—Ď–†—ė –†–Ü–°‚ÄĻ–†¬Ī–°–ā–†¬į–†–Ö–†–Ö–°‚ÄĻ–†‚ĄĖ –°–É–†¬Ľ–†—ē–°‚Äö, –†¬Ķ–°–É–†¬Ľ–†—Ď –†—ó–†¬Ķ–°–ā–†¬Ķ–†“Ď–†¬į–†–Ö
    chosen_slots = None
    if booking.time_slot_id:
        logging.info(f"Attempting to book with specific time_slot_id: {booking.time_slot_id}")
        initial_slot = db.query(models.TimeSlot).filter(models.TimeSlot.id == booking.time_slot_id).first()
        
        if initial_slot:
            logging.info(f"Initial slot found: id={initial_slot.id}, available={initial_slot.is_available}, date={initial_slot.slot_date}, time={initial_slot.start_time}")
        else:
            logging.info("Initial slot not found.")

        if initial_slot and initial_slot.is_available and initial_slot.slot_date == booking_date and initial_slot.start_time == start_time:
            logging.info("Initial slot checks passed.")
            # –†—ü–†—ē–†¬Ľ–°—ď–°‚Ä°–†¬į–†¬Ķ–†—ė –°–É–†¬Ľ–†—ē–°‚Äö–°‚ÄĻ –†“Ď–†¬Ľ–°–Ź –°–Ć–°‚Äö–†—ē–†—Ė–†—ē –†“Ď–†—ē–†—Ē–†¬į –†–Ö–†¬į–°‚Ä°–†—Ď–†–Ö–†¬į–°–Ź –°–É –†–Ü–°‚ÄĻ–†¬Ī–°–ā–†¬į–†–Ö–†–Ö–†—ē–†—Ė–†—ē (–°–ā–†¬į–†¬∑–°–ā–†¬Ķ–°‚ā¨–†¬į–†¬Ķ–†—ė –°–É–†¬Ľ–†¬Ķ–†“Ď–°—ď–°–č–°‚Äį–†—Ď–†‚ĄĖ –†“Ď–†¬Ķ–†–Ö–°–ä)
            dock_slots = db.query(models.TimeSlot).join(models.Dock).filter(
                models.TimeSlot.dock_id == initial_slot.dock_id,
                models.TimeSlot.slot_date.in_([booking_date, next_date]),
                models.TimeSlot.is_available == True,
                models.Dock.object_id == booking.object_id
            ).order_by(models.TimeSlot.slot_date, models.TimeSlot.start_time).all()
            logging.info(f"Found {len(dock_slots)} subsequent slots for the same dock (can span to next day).")

            # –†—ú–†¬į–†‚ĄĖ–†“Ď–†¬Ķ–†—ė –†—Ď–†–Ö–†“Ď–†¬Ķ–†—Ē–°–É –†–Ö–†¬į–°‚Ä°–†¬į–†¬Ľ–°–ä–†–Ö–†—ē–†—Ė–†—ē –°–É–†¬Ľ–†—ē–°‚Äö–†¬į
            start_index = None
            for idx, slot in enumerate(dock_slots):
                if slot.id == initial_slot.id:
                    start_index = idx
                    break
            logging.info(f"Found start_index: {start_index}")

            if start_index is not None and start_index + required_slots <= len(dock_slots):
                logging.info("Sufficient subsequent slots found.")
                candidate_chain = dock_slots[start_index:start_index + required_slots]
                logging.info(f"Candidate chain of {len(candidate_chain)} slots: {[s.id for s in candidate_chain]}")

                # –†—ü–°–ā–†—ē–†–Ü–†¬Ķ–°–ā–†—Ď–†—ė –†–Ö–†¬Ķ–†—ó–°–ā–†¬Ķ–°–ā–°‚ÄĻ–†–Ü–†–Ö–†—ē–°–É–°‚Äö–°–ä –†—ó–†—ē datetime (–°—ď–°‚Ä°–†—Ď–°‚Äö–°‚ÄĻ–†–Ü–†¬į–°–Ź –†—ó–†¬Ķ–°–ā–†¬Ķ–°‚Ä¶–†—ē–†“Ď –†–Ö–†¬į –°–É–†¬Ľ–†¬Ķ–†“Ď–°—ď–°–č–°‚Äį–†—Ď–†‚ĄĖ –†“Ď–†¬Ķ–†–Ö–°–ä)
                is_continuous = True
                for j in range(len(candidate_chain) - 1):
                    current_end = datetime.combine(candidate_chain[j].slot_date, candidate_chain[j].end_time)
                    next_start = datetime.combine(candidate_chain[j + 1].slot_date, candidate_chain[j + 1].start_time)
                    if current_end != next_start:
                        is_continuous = False
                        logging.info(
                            f"Chain is not continuous at index {j}. Slot {candidate_chain[j].id} "
                            f"ends at {current_end}, next slot {candidate_chain[j+1].id} starts at {next_start}"
                        )
                        break
                
                if is_continuous:
                    logging.info("Chain is continuous.")
                    # –†—ü–°–ā–†—ē–†–Ü–†¬Ķ–°–ā–†—Ď–†—ė –†¬Ľ–†—Ď–†—ė–†—Ď–°‚Äö–°‚ÄĻ –†—ó–°–ā–†—ē–†—ó–°—ď–°–É–†—Ē–†–Ö–†—ē–†‚ĄĖ –°–É–†—ó–†—ē–°–É–†—ē–†¬Ī–†–Ö–†—ē–°–É–°‚Äö–†—Ď –†—ē–†¬Ī–°–Č–†¬Ķ–†—Ē–°‚Äö–†¬į
                    dock = db.query(models.Dock).options(
                        joinedload(models.Dock.available_zones)
                    ).filter(models.Dock.id == initial_slot.dock_id).first()
                    obj_for_check = dock.object if dock else None

                    capacity_block = False
                    if dock and obj_for_check:
                        if not _dock_matches_supplier_zone(dock, supplier_zone_id):
                            logging.info(
                                f"Initial slot dock {dock.id} is not allowed for supplier zone {supplier_zone_id}."
                            )
                            capacity_block = True
                        logging.info(f"Checking object capacity for object_id: {obj_for_check.id}")
                        limits_to_check = []
                        if dock.dock_type == models.DockType.entrance:
                            limits_to_check.append(("in", obj_for_check.capacity_in, [models.DockType.entrance, models.DockType.universal]))
                        elif dock.dock_type == models.DockType.exit:
                            limits_to_check.append(("out", obj_for_check.capacity_out, [models.DockType.exit, models.DockType.universal]))
                        else:  # universal -> –†—ó–°–ā–†—ē–†–Ü–†¬Ķ–°–ā–°–Ź–†¬Ķ–†—ė –†¬Ľ–†—Ď–†—ė–†—Ď–°‚Äö –†–Ü –†¬∑–†¬į–†–Ü–†—Ď–°–É–†—Ď–†—ė–†—ē–°–É–°‚Äö–†—Ď –†—ē–°‚Äö –°‚Äö–†—Ď–†—ó–†¬į –†¬Ī–°–ā–†—ē–†–Ö–†—Ď–°–ā–†—ē–†–Ü–†¬į–†–Ö–†—Ď–°–Ź
                            if booking_direction == models.BookingDirection.inbound:
                                limits_to_check.append(("in", obj_for_check.capacity_in, [models.DockType.entrance, models.DockType.universal]))
                            elif booking_direction == models.BookingDirection.outbound:
                                limits_to_check.append(("out", obj_for_check.capacity_out, [models.DockType.exit, models.DockType.universal]))
                        
                        logging.info(f"Limits to check: {limits_to_check}")

                        for direction, cap_limit, types_to_use in limits_to_check:
                            if not cap_limit or cap_limit <= 0:
                                logging.info("No capacity limit set or limit is zero. Skipping check.")
                                continue
                            logging.info(f"Checking {direction} capacity. Limit: {cap_limit}")
                            for slot in candidate_chain:
                                occupancy_obj = db.query(func.count(models.BookingTimeSlot.id)).join(models.Booking, models.BookingTimeSlot.booking_id == models.Booking.id).join(
                                    models.TimeSlot, models.BookingTimeSlot.time_slot_id == models.TimeSlot.id
                                ).join(
                                    models.Dock, models.TimeSlot.dock_id == models.Dock.id
                                ).filter(
                                    models.Dock.object_id == obj_for_check.id,
                                    models.Dock.dock_type.in_(types_to_use),
                                    models.TimeSlot.slot_date == slot.slot_date,
                                    models.TimeSlot.start_time == slot.start_time,
                                    models.TimeSlot.end_time == slot.end_time,
                                    models.Booking.status == "confirmed"
                                ).scalar() or 0
                                logging.info(f"Slot {slot.id} ({slot.start_time}): Object occupancy is {occupancy_obj}")
                                if occupancy_obj >= cap_limit:
                                    capacity_block = True
                                    logging.warning(f"Object capacity limit reached for slot {slot.id}. Occupancy ({occupancy_obj}) >= Limit ({cap_limit})")
                                    break
                            if capacity_block:
                                break

                    if not capacity_block:
                        logging.info("Object capacity check passed.")
                        # –†—ü–°–ā–†—ē–†–Ü–†¬Ķ–°–ā–†—Ď–†—ė –†“Ď–†—ē–°–É–°‚Äö–°—ď–†—ó–†–Ö–†—ē–°–É–°‚Äö–°–ä
                        all_available = True
                        for slot in candidate_chain:
                            current_occupancy = db.query(func.count(models.BookingTimeSlot.id)).filter(
                                models.BookingTimeSlot.time_slot_id == slot.id
                            ).scalar() or 0
                            logging.info(f"Slot {slot.id}: current_occupancy={current_occupancy}, capacity={slot.capacity}")
                            if current_occupancy >= slot.capacity:
                                all_available = False
                                logging.warning(f"Slot capacity limit reached for slot {slot.id}. Occupancy ({current_occupancy}) >= Capacity ({slot.capacity})")
                                break

                        if all_available:
                            logging.info("Slot capacity check passed. Setting chosen_slots.")
                            chosen_slots = candidate_chain
                        else:
                            logging.warning("Slot capacity check failed.")
                    else:
                        logging.warning("Object capacity check failed.")
                else:
                    logging.warning("Continuity check failed.")
            else:
                logging.warning("Not enough subsequent slots available in dock_slots.")
        else:
            logging.warning("Initial slot check failed (is_available, date, or time mismatch).")

    logging.info(f"Value of chosen_slots after specific slot check: {[s.id for s in chosen_slots] if chosen_slots else 'None'}")

    # –†‚ÄĘ–°–É–†¬Ľ–†—Ď –†–Ü–°‚ÄĻ–†¬Ī–°–ā–†¬į–†–Ö–†–Ö–°‚ÄĻ–†‚ĄĖ –°–É–†¬Ľ–†—ē–°‚Äö –†–Ö–†¬Ķ –†—ó–†—ē–†“Ď–†—ē–°‚ā¨–†¬Ķ–†¬Ľ, –†—Ď–°‚Äį–†¬Ķ–†—ė –†¬Ľ–°–č–†¬Ī–†—ē–†‚ĄĖ –†“Ď–†—ē–°–É–°‚Äö–°—ď–†—ó–†–Ö–°‚ÄĻ–†‚ĄĖ
    if not chosen_slots:
        logging.info("Specific slot not booked. Searching for any available slot.")
        # –†—ú–†¬į–°‚Ä¶–†—ē–†“Ď–†—Ď–†—ė –†“Ď–†—ē–°–É–°‚Äö–°—ď–†—ó–†–Ö–°‚ÄĻ–†¬Ķ –°–É–†¬Ľ–†—ē–°‚Äö–°‚ÄĻ
        available_slots = db.query(models.TimeSlot).join(models.Dock).filter(
            models.TimeSlot.slot_date.in_([booking_date, next_date]),
            models.TimeSlot.is_available == True,
            models.Dock.object_id == booking.object_id
        ).filter(
            (models.TimeSlot.slot_date > booking_date) | (models.TimeSlot.start_time >= start_time)
        ).order_by(models.TimeSlot.slot_date, models.TimeSlot.start_time).all()
        logging.info(f"Found {len(available_slots)} total available slots for the object (spanning start and next day).")
    else:
        available_slots = []

    # –†‚Äú–°–ā–°—ď–†—ó–†—ó–†—Ď–°–ā–°—ď–†¬Ķ–†—ė –°–É–†¬Ľ–†—ē–°‚Äö–°‚ÄĻ –†—ó–†—ē –†“Ď–†—ē–†—Ē–†¬į–†—ė
    slots_by_dock = {}
    for slot in available_slots:
        if slot.dock_id not in slots_by_dock:
            slots_by_dock[slot.dock_id] = []
        slots_by_dock[slot.dock_id].append(slot)

    dock_ids = list(slots_by_dock.keys())
    docks_from_db = (
        db.query(models.Dock).options(
            joinedload(models.Dock.available_transport_types),
            joinedload(models.Dock.available_zones),
        ).filter(models.Dock.id.in_(dock_ids)).all() if dock_ids else []
    )
    dock_map = {d.id: d for d in docks_from_db}

    # –†–é–†—ē–°–ā–°‚Äö–†—Ď–°–ā–°—ď–†¬Ķ–†—ė –†“Ď–†—ē–†—Ē–†—Ď –†—ó–†—ē –†—ó–°–ā–†—Ď–†—ē–°–ā–†—Ď–°‚Äö–†¬Ķ–°‚Äö–°—ď –†–Ü –†¬∑–†¬į–†–Ü–†—Ď–°–É–†—Ď–†—ė–†—ē–°–É–°‚Äö–†—Ď –†—ē–°‚Äö –°‚Äö–†—Ď–†—ó–†¬į –†¬Ī–°–ā–†—ē–†–Ö–†—Ď–°–ā–†—ē–†–Ü–†¬į–†–Ö–†—Ď–°–Ź
    def sort_key(dock_id):
        dock = dock_map.get(dock_id)
        if not dock:
            return 3 # Should not happen
        
        if booking_direction == models.BookingDirection.outbound:
            if dock.dock_type == models.DockType.exit:
                return 0
            if dock.dock_type == models.DockType.universal:
                return 1
        elif booking_direction == models.BookingDirection.inbound:
            if dock.dock_type == models.DockType.entrance:
                return 0
            if dock.dock_type == models.DockType.universal:
                return 1
        return 2 # Other types last

    sorted_dock_ids = sorted(dock_ids, key=sort_key)
    logging.info(f"Searching docks in order: {sorted_dock_ids}")
    
    # –†¬ė–°‚Äį–†¬Ķ–†—ė –†—ó–†—ē–†“Ď–°‚Ä¶–†—ē–†“Ď–°–Ź–°‚Äį–°—ď–°–č –°‚Ä†–†¬Ķ–†—ó–†—ē–°‚Ä°–†—Ē–°—ď –°–É–†¬Ľ–†—ē–°‚Äö–†—ē–†–Ü
    if not chosen_slots: # Ensure chosen_slots is explicitly reset if the first block failed
        chosen_slots = None 
    for dock_id in sorted_dock_ids:
        logging.info(f"--- Checking Dock ID: {dock_id} ---")
        dock_slots = slots_by_dock[dock_id]
        # –†–é–†—ē–°–ā–°‚Äö–†—Ď–°–ā–°—ď–†¬Ķ–†—ė –°–É–†¬Ľ–†—ē–°‚Äö–°‚ÄĻ –†—ó–†—ē –†“Ď–†¬į–°‚Äö–†¬Ķ –†—Ď –†–Ü–°–ā–†¬Ķ–†—ė–†¬Ķ–†–Ö–†—Ď, –°‚Ä°–°‚Äö–†—ē–†¬Ī–°‚ÄĻ –†—Ē–†—ē–°–ā–°–ā–†¬Ķ–†—Ē–°‚Äö–†–Ö–†—ē –†—ē–†¬Ī–°–ā–†¬į–†¬Ī–†¬į–°‚Äö–°‚ÄĻ–†–Ü–†¬į–°‚Äö–°–ä –†—ó–†¬Ķ–°–ā–†¬Ķ–°‚Ä¶–†—ē–†“Ď –°‚Ä°–†¬Ķ–°–ā–†¬Ķ–†¬∑ –°–É–°—ď–°‚Äö–†—Ē–†—Ď
        dock_slots.sort(key=lambda x: (x.slot_date, x.start_time))
        
        dock = dock_map.get(dock_id)
        obj = dock.object if dock else None

        # –†—ü–°–ā–†—ē–†—ó–°—ď–°–É–†—Ē–†¬į–†¬Ķ–†—ė –†“Ď–†—ē–†—Ē–†—Ď –†–Ö–†¬Ķ–†—ó–†—ē–†“Ď–°‚Ä¶–†—ē–†“Ď–°–Ź–°‚Äį–†¬Ķ–†—Ė–†—ē –°‚Äö–†—Ď–†—ó–†¬į
        if booking_direction == models.BookingDirection.inbound and dock and dock.dock_type == models.DockType.exit:
            logging.info(f"Skipping dock {dock_id}: type is 'exit' for an 'in' booking.")
            continue
        if booking_direction == models.BookingDirection.outbound and dock and dock.dock_type == models.DockType.entrance:
            logging.info(f"Skipping dock {dock_id}: type is 'entrance' for an 'out' booking.")
            continue
        if dock and not _dock_matches_supplier_zone(dock, supplier_zone_id):
            logging.info(f"Skipping dock {dock_id}: not allowed for supplier zone {supplier_zone_id}.")
            continue
        
        # –†—ü–°–ā–†—ē–†–Ü–†¬Ķ–°–ā–°–Ź–†¬Ķ–†—ė, –°–ā–†¬į–†¬∑–°–ā–†¬Ķ–°‚ā¨–†¬Ķ–†–Ö –†¬Ľ–†—Ď –°‚Äö–†—Ď–†—ó –†—ó–†¬Ķ–°–ā–†¬Ķ–†–Ü–†—ē–†¬∑–†—Ē–†—Ď –†“Ď–†¬Ľ–°–Ź –°–Ć–°‚Äö–†—ē–†—Ė–†—ē –†“Ď–†—ē–†—Ē–†¬į
        if booking.transport_type_id and dock and dock.available_transport_types:
            allowed_transport_ids = {t.id for t in dock.available_transport_types}
            if booking.transport_type_id not in allowed_transport_ids:
                logging.info(f"Skipping dock {dock_id}: transport_type_id {booking.transport_type_id} not allowed.")
                continue # –†—ü–†¬Ķ–°–ā–†¬Ķ–°‚Ä¶–†—ē–†“Ď–†—Ď–†—ė –†—Ē –°–É–†¬Ľ–†¬Ķ–†“Ď–°—ď–°–č–°‚Äį–†¬Ķ–†—ė–°—ď –†“Ď–†—ē–†—Ē–°—ď

        # –†¬ė–°‚Äį–†¬Ķ–†—ė –°–É–†¬Ľ–†—ē–°‚Äö–°‚ÄĻ –°–É–°‚Äö–°–ā–†—ē–†—Ė–†—ē –†–Ö–†¬į–°‚Ä°–†—Ď–†–Ö–†¬į–°–Ź –°–É –†¬∑–†¬į–†—ó–°–ā–†—ē–°‚ā¨–†¬Ķ–†–Ö–†–Ö–†—ē–†—Ė–†—ē –†–Ü–°–ā–†¬Ķ–†—ė–†¬Ķ–†–Ö–†—Ď; –†“Ď–†—ē–†—ó–°—ď–°–É–†—Ē–†¬į–†¬Ķ–†—ė –†—ó–†¬Ķ–°–ā–†¬Ķ–°–ā–°‚ÄĻ–†–Ü–°‚ÄĻ –†–Ü –°–ā–†¬į–°–É–†—ó–†—Ď–°–É–†¬į–†–Ö–†—Ď–†—Ď,
        # –†–Ö–†—ē –†–Ö–†¬Ķ –†“Ď–†—ē–†—ó–°—ď–°–É–†—Ē–†¬į–†¬Ķ–†—ė –†¬∑–†¬į–†–Ö–°–Ź–°‚Äö–°‚ÄĻ–†¬Ķ –°–É–†¬Ľ–†—ē–°‚Äö–°‚ÄĻ (capacity) –†–Ü –°–É–†¬Ķ–°–ā–†¬Ķ–†“Ď–†—Ď–†–Ö–†¬Ķ –°‚Ä†–†¬Ķ–†—ó–†—ē–°‚Ä°–†—Ē–†—Ď.
        start_idx = next((idx for idx, s in enumerate(dock_slots) if s.slot_date == booking_date and s.start_time == start_time), None)
        if start_idx is None:
            logging.info(f"No starting slot at {start_time} in dock {dock_id}. Skipping dock.")
            continue

        accumulated_minutes = 0
        candidate_chain = []
        valid_dock = True

        for slot in dock_slots[start_idx:]:
            # –†—ü–°–ā–†—ē–†–Ü–†¬Ķ–°–ā–°–Ź–†¬Ķ–†—ė –†¬Ľ–†—Ď–†—ė–†—Ď–°‚Äö–°‚ÄĻ –†—ó–°–ā–†—ē–†—ó–°—ď–°–É–†—Ē–†–Ö–†—ē–†‚ĄĖ –°–É–†—ó–†—ē–°–É–†—ē–†¬Ī–†–Ö–†—ē–°–É–°‚Äö–†—Ď –†—ē–†¬Ī–°–Č–†¬Ķ–†—Ē–°‚Äö–†¬į –†—ó–†—ē –†–Ö–†¬į–†—ó–°–ā–†¬į–†–Ü–†¬Ľ–†¬Ķ–†–Ö–†—Ď–°–č
            if dock and obj:
                limits_to_check = []
                if dock.dock_type == models.DockType.entrance:
                    limits_to_check.append(("in", obj.capacity_in, [models.DockType.entrance, models.DockType.universal]))
                elif dock.dock_type == models.DockType.exit:
                    limits_to_check.append(("out", obj.capacity_out, [models.DockType.exit, models.DockType.universal]))
                else:  # universal
                    if booking_direction == models.BookingDirection.inbound:
                        limits_to_check.append(("in", obj.capacity_in, [models.DockType.entrance, models.DockType.universal]))
                    elif booking_direction == models.BookingDirection.outbound:
                        limits_to_check.append(("out", obj.capacity_out, [models.DockType.exit, models.DockType.universal]))

                capacity_block = False
                for _, cap_limit, types_to_use in limits_to_check:
                    if not cap_limit or cap_limit <= 0:
                        continue
                    occupancy_obj = db.query(func.count(models.BookingTimeSlot.id)).join(models.Booking, models.BookingTimeSlot.booking_id == models.Booking.id).join(
                        models.TimeSlot, models.BookingTimeSlot.time_slot_id == models.TimeSlot.id
                    ).join(
                        models.Dock, models.TimeSlot.dock_id == models.Dock.id
                    ).filter(
                        models.Dock.object_id == obj.id,
                        models.Dock.dock_type.in_(types_to_use),
                        models.TimeSlot.slot_date == slot.slot_date,
                        models.TimeSlot.start_time == slot.start_time,
                        models.TimeSlot.end_time == slot.end_time,
                        models.Booking.status == "confirmed"
                    ).scalar() or 0
                    if occupancy_obj >= cap_limit:
                        capacity_block = True
                        break
                if capacity_block:
                    logging.info(f"Object capacity block for slot {slot.id} in dock {dock_id}.")
                    valid_dock = False
                    break

            # –†—ü–°–ā–†—ē–†–Ü–†¬Ķ–°–ā–°–Ź–†¬Ķ–†—ė –†¬∑–†¬į–†–Ö–°–Ź–°‚Äö–†—ē–°–É–°‚Äö–°–ä –°–É–†¬Ľ–†—ē–°‚Äö–†¬į
            current_occupancy = db.query(func.count(models.BookingTimeSlot.id)).filter(
                models.BookingTimeSlot.time_slot_id == slot.id
            ).scalar() or 0
            if current_occupancy >= slot.capacity:
                logging.info(f"Slot {slot.id} in dock {dock_id} is fully occupied. Dock rejected.")
                valid_dock = False
                break

            candidate_chain.append(slot)
            slot_minutes = int((datetime.combine(slot.slot_date, slot.end_time) - datetime.combine(slot.slot_date, slot.start_time)).total_seconds() // 60)
            accumulated_minutes += slot_minutes

            if accumulated_minutes >= duration:
                chosen_slots = candidate_chain
                logging.info(f"Accumulated required duration in dock {dock_id}. Slots: {[s.id for s in candidate_chain]}")
                break

        if chosen_slots:
            break
        if not valid_dock:
            logging.info(f"Dock {dock_id} rejected due to occupied slot or capacity block.")
            continue
    
    if not chosen_slots:
        logging.error("--- No suitable slots found. Raising 409 Conflict. ---")
        raise HTTPException(
            status_code=409, 
            detail="–†—ú–†¬į –†¬∑–†¬į–†—ó–°–ā–†—ē–°‚ā¨–†¬Ķ–†–Ö–†–Ö–°‚ÄĻ–†‚ĄĖ –†—ó–†¬Ķ–°–ā–†—Ď–†—ē–†“Ď –†–Ö–†¬Ķ –†–Ö–†¬į–†‚ĄĖ–†“Ď–†¬Ķ–†–Ö–†—ē –†“Ď–†—ē–°–É–°‚Äö–°—ď–†—ó–†–Ö–°‚ÄĻ–°‚Ä¶ –†–Ü–°–ā–†¬Ķ–†—ė–†¬Ķ–†–Ö–†–Ö–°‚ÄĻ–°‚Ä¶ –°–É–†¬Ľ–†—ē–°‚Äö–†—ē–†–Ü"
        )

    quota, total_quota_volume = get_quota_for_date(
        db=db,
        object_id=booking.object_id,
        transport_type_id=booking.transport_type_id,
        target_date=booking_date,
        direction=booking_direction,
    )
    if quota and total_quota_volume is not None:
        if booking.cubes is None:
            raise HTTPException(status_code=400, detail="Volume (cubes) is required because a quota applies on this date")
        used_volume = calculate_used_volume(
            db=db,
            object_id=booking.object_id,
            transport_type_id=booking.transport_type_id,
            target_date=booking_date,
            direction=booking_direction,
        )
        remaining_volume = total_quota_volume - used_volume
        if not quota.allow_overbooking and booking.cubes > remaining_volume:
            raise HTTPException(
                status_code=400,
                detail=f"Quota exceeded for {booking_date}. Remaining: {remaining_volume}, requested: {booking.cubes}",
            )
    
    logging.info(f"--- Booking successful. Creating booking with slots: {[s.id for s in chosen_slots]} ---")
    # –†–é–†—ē–†¬∑–†“Ď–†¬į–†¬Ķ–†—ė –†¬∑–†¬į–†—ó–†—Ď–°–É–°–ä
    new_booking = models.Booking(
        user_id=current_user.id,
        vehicle_type_id=booking.vehicle_type_id,
        vehicle_plate=booking.vehicle_plate or "",
        driver_full_name=booking.driver_full_name or "",
        driver_phone=booking.driver_phone or "",
        status="confirmed",
        supplier_id=booking.supplier_id,
        zone_id=booking.zone_id,
        transport_type_id=booking.transport_type_id,
        cubes=booking.cubes,
        transport_sheet=booking.transport_sheet,
        booking_type=booking_direction,
    )
    db.add(new_booking)
    db.flush()  # –†—ü–†—ē–†¬Ľ–°—ď–°‚Ä°–†¬į–†¬Ķ–†—ė ID
    
    # –†–é–†—ē–†¬∑–†“Ď–†¬į–†¬Ķ–†—ė –°–É–†–Ü–°–Ź–†¬∑–†—Ď –°–É –†–Ü–°–ā–†¬Ķ–†—ė–†¬Ķ–†–Ö–†–Ö–°‚ÄĻ–†—ė–†—Ď –°–É–†¬Ľ–†—ē–°‚Äö–†¬į–†—ė–†—Ď
    for slot in chosen_slots:
        booking_slot = models.BookingTimeSlot(
            booking_id=new_booking.id,
            time_slot_id=slot.id
        )
        db.add(booking_slot)
    
    db.commit()
    db.refresh(new_booking)
    return new_booking
@router.put("/{booking_id}/cancel")
def cancel_booking(
    booking_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    """–†—õ–°‚Äö–†—ė–†¬Ķ–†–Ö–†—Ď–°‚Äö–°–ä –†¬∑–†¬į–†—ó–†—Ď–°–É–°–ä"""
    booking_query = db.query(models.Booking).filter(models.Booking.id == booking_id)
    if current_user.role != models.UserRole.admin:
        booking_query = booking_query.filter(models.Booking.user_id == current_user.id)
    booking = booking_query.first()
    
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")
    
    if booking.status != "confirmed":
        raise HTTPException(status_code=400, detail="Booking is not in confirmed status")
    
    booking.status = "cancelled"
    booking.updated_at = datetime.utcnow()
    
    # –†–ą–†“Ď–†¬į–†¬Ľ–°–Ź–†¬Ķ–†—ė –°–É–†–Ü–°–Ź–†¬∑–†—Ď –°–É –†–Ü–°–ā–†¬Ķ–†—ė–†¬Ķ–†–Ö–†–Ö–°‚ÄĻ–†—ė–†—Ď –°–É–†¬Ľ–†—ē–°‚Äö–†¬į–†—ė–†—Ď, –°‚Ä°–°‚Äö–†—ē–†¬Ī–°‚ÄĻ –†—ē–†–Ö–†—Ď –°–É–†–Ö–†—ē–†–Ü–†¬į –°–É–°‚Äö–†¬į–†¬Ľ–†—Ď –†“Ď–†—ē–°–É–°‚Äö–°—ď–†—ó–†–Ö–°‚ÄĻ
    db.query(models.BookingTimeSlot).filter(
        models.BookingTimeSlot.booking_id == booking_id
    ).delete()
    
    db.commit()
    
    return {"message": "Booking cancelled successfully"}

@router.get("/all", response_model=schemas.BookingListPage)
def get_all_bookings(
    params: BookingListParams = Depends(get_booking_list_params),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    """–†—ü–†—ē–†¬Ľ–°—ď–°‚Ä°–†—Ď–°‚Äö–°–ä –†–Ü–°–É–†¬Ķ –†¬∑–†¬į–†—ó–†—Ď–°–É–†—Ď (–°‚Äö–†—ē–†¬Ľ–°–ä–†—Ē–†—ē –†“Ď–†¬Ľ–°–Ź –†¬į–†“Ď–†—ė–†—Ď–†–Ö–†—Ď–°–É–°‚Äö–°–ā–†¬į–°‚Äö–†—ē–°–ā–†—ē–†–Ü)"""
    # –†—ü–°–ā–†—ē–†–Ü–†¬Ķ–°–ā–°–Ź–†¬Ķ–†—ė –†—ó–°–ā–†¬į–†–Ü–†¬į –†¬į–†“Ď–†—ė–†—Ď–†–Ö–†—Ď–°–É–°‚Äö–°–ā–†¬į–°‚Äö–†—ē–°–ā–†¬į
    if current_user.role != models.UserRole.admin:
        raise HTTPException(status_code=403, detail="–†—ú–†¬Ķ–†“Ď–†—ē–°–É–°‚Äö–†¬į–°‚Äö–†—ē–°‚Ä°–†–Ö–†—ē –†—ó–°–ā–†¬į–†–Ü")

    return _build_paginated_bookings_response(db, current_user, params)


@router.get("/my", response_model=schemas.BookingListPage)
def get_my_bookings(
    params: BookingListParams = Depends(get_booking_list_params),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    """–†—ü–†—ē–†¬Ľ–°—ď–°‚Ä°–†—Ď–°‚Äö–°–ä –†–Ü–°–É–†¬Ķ –†¬∑–†¬į–†—ó–†—Ď–°–É–†—Ď (–†¬Ī–°‚ÄĻ–†–Ü–°‚ā¨–†—Ď–†¬Ķ \"–†—ė–†—ē–†—Ď\"), –†–Ü–†—Ď–†“Ď–†–Ö–°‚ÄĻ –†–Ü–°–É–†¬Ķ–†—ė –†—ó–†—ē–†¬Ľ–°–ä–†¬∑–†—ē–†–Ü–†¬į–°‚Äö–†¬Ķ–†¬Ľ–°–Ź–†—ė"""
    return _build_paginated_bookings_response(db, current_user, params)


@router.post("/export/xlsx")
def export_bookings_xlsx(
    booking_ids: Optional[List[int]] = Body(None),
    variant: str = "default",
    params: BookingListParams = Depends(get_booking_list_params),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    """–†¬≠–†—Ē–°–É–†—ó–†—ē–°–ā–°‚Äö –†–Ü–°‚ÄĻ–†¬Ī–°–ā–†¬į–†–Ö–†–Ö–°‚ÄĻ–°‚Ä¶ –†¬Ī–°–ā–†—ē–†–Ö–†—Ď–°–ā–†—ē–†–Ü–†¬į–†–Ö–†—Ď–†‚ĄĖ –†–Ü XLSX."""
    if variant not in {"default", "start-end"}:
        raise HTTPException(status_code=400, detail="Unsupported export variant")

    if booking_ids:
        unique_ids = list(dict.fromkeys(booking_ids))
    else:
        filtered_rows = _build_booking_listing_query(db, params, current_user).all()
        unique_ids = [row.booking_id for row in filtered_rows]

    if not unique_ids:
        raise HTTPException(status_code=400, detail="No booking IDs provided")

    bookings = (
        db.query(models.Booking)
        .options(
            joinedload(models.Booking.user),
            joinedload(models.Booking.vehicle_type),
            joinedload(models.Booking.supplier),
            joinedload(models.Booking.zone),
            joinedload(models.Booking.transport_type),
        )
        .filter(
            models.Booking.id.in_(unique_ids),
            models.Booking.status == "confirmed",
        )
        .all()
    )
    booking_by_id = {b.id: b for b in bookings}
    serialized_by_id = _serialize_bookings_bulk(db, bookings, include_user=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "bookings"
    variant_headers = [
        "\u0414\u0430\u0442\u0430 \u043d\u0430\u0447\u0430\u043b\u0430",
        "\u0412\u0440\u0435\u043c\u044f \u043d\u0430\u0447\u0430\u043b\u0430",
        "\u0414\u0430\u0442\u0430 \u043e\u043a\u043e\u043d\u0447\u0430\u043d\u0438\u044f",
        "\u0412\u0440\u0435\u043c\u044f \u043e\u043a\u043e\u043d\u0447\u0430\u043d\u0438\u044f",
        "\u0422\u0440\u0430\u043d\u0441\u043f\u043e\u0440\u0442\u043d\u044b\u0439 \u043b\u0438\u0441\u0442",
        "\u041f\u043e\u0441\u0442\u0430\u0432\u0449\u0438\u043a",
        "\u041a\u0443\u0431\u044b",
        "\u0422\u0438\u043f \u0422\u0421",
        "\u041e\u0431\u044a\u0435\u043a\u0442",
        "\u0417\u043e\u043d\u0430",
        "\u0422\u0438\u043f \u043f\u0435\u0440\u0435\u0432\u043e\u0437\u043a\u0438",
    ] if variant == "start-end" else None
    default_headers = [
        "\u0414\u0430\u0442\u0430",
        "\u0412\u0440\u0435\u043c\u044f",
        "\u0414\u043e\u043a",
        "\u041e\u0431\u044a\u0435\u043a\u0442",
        "\u0422\u0438\u043f \u0422\u0421",
        "\u041f\u043e\u0441\u0442\u0430\u0432\u0449\u0438\u043a",
        "\u0417\u043e\u043d\u0430",
        "\u0422\u0438\u043f \u043f\u0435\u0440\u0435\u0432\u043e\u0437\u043a\u0438",
        "\u041a\u0443\u0431\u044b",
        "\u0422\u0440\u0430\u043d\u0441\u043f\u043e\u0440\u0442\u043d\u044b\u0439 \u043b\u0438\u0441\u0442",
        "\u0421\u0442\u0430\u0442\u0443\u0441",
        "\u041f\u043e\u043b\u044c\u0437\u043e\u0432\u0430\u0442\u0435\u043b\u044c",
        "ID \u0431\u0440\u043e\u043d\u0438\u0440\u043e\u0432\u0430\u043d\u0438\u044f",
    ]
    ws.append([
        "–†‚ÄĚ–†¬į–°‚Äö–†¬į",
        "–†‚Äô–°–ā–†¬Ķ–†—ė–°–Ź",
        "–†‚ÄĚ–†—ē–†—Ē",
        "–†—õ–†¬Ī–°–Č–†¬Ķ–†—Ē–°‚Äö",
        "–†—ě–†—Ď–†—ó –†—ě–†–é",
        "–†—ü–†—ē–°–É–°‚Äö–†¬į–†–Ü–°‚Äį–†—Ď–†—Ē",
        "–†‚ÄĒ–†—ē–†–Ö–†¬į",
        "–†—ě–†—Ď–†—ó –†—ó–†¬Ķ–°–ā–†¬Ķ–†–Ü–†—ē–†¬∑–†—Ē–†—Ď",
        "–†—ô–°—ď–†¬Ī–°‚ÄĻ",
        "–†—ě–°–ā–†¬į–†–Ö–°–É–†—ó–†—ē–°–ā–°‚Äö–†–Ö–°‚ÄĻ–†‚ĄĖ –†¬Ľ–†—Ď–°–É–°‚Äö",
        "–†–é–°‚Äö–†¬į–°‚Äö–°—ď–°–É",
        "–†—ü–†—ē–†¬Ľ–°–ä–†¬∑–†—ē–†–Ü–†¬į–°‚Äö–†¬Ķ–†¬Ľ–°–ä",
        "ID –†¬Ī–°–ā–†—ē–†–Ö–†—Ď–°–ā–†—ē–†–Ü–†¬į–†–Ö–†—Ď–°–Ź",
    ])
    headers_to_apply = variant_headers or default_headers
    for col_idx, header in enumerate(headers_to_apply, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    for booking_id in unique_ids:
        booking = booking_by_id.get(booking_id)
        if not booking:
            continue

        serialized = serialized_by_id.get(booking_id)
        if not serialized:
            continue

        start_time = (serialized.get("start_time") or "")[:5]
        end_time = (serialized.get("end_time") or "")[:5]
        time_range = f"{start_time} - {end_time}" if start_time or end_time else ""
        user_label = serialized.get("user_full_name") or serialized.get("user_email") or serialized.get("user_login") or ""

        if variant == "start-end":
            ws.append([
                _format_export_date(serialized.get("booking_date")),
                start_time,
                _format_export_date(serialized.get("end_date")),
                end_time,
                serialized.get("transport_sheet") or "",
                serialized.get("supplier_name") or "",
                serialized.get("cubes") if serialized.get("cubes") is not None else "",
                serialized.get("vehicle_type_name") or "",
                serialized.get("object_name") or "",
                serialized.get("zone_name") or "",
                serialized.get("transport_type_name") or "",
            ])
        else:
            ws.append([
                serialized.get("booking_date") or "",
                time_range,
                serialized.get("dock_name") or "",
                serialized.get("object_name") or "",
                serialized.get("vehicle_type_name") or "",
                serialized.get("supplier_name") or "",
                serialized.get("zone_name") or "",
                serialized.get("transport_type_name") or "",
                serialized.get("cubes") if serialized.get("cubes") is not None else "",
                serialized.get("transport_sheet") or "",
                serialized.get("status") or "",
                user_label,
                serialized.get("id") or "",
            ])

        row_fill = _resolve_export_row_fill(serialized)
        if row_fill is not None:
            current_row = ws.max_row
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=current_row, column=col_idx).fill = row_fill

    _append_bookings_report_sheets(wb, [serialized_by_id[booking_id] for booking_id in unique_ids if booking_id in serialized_by_id])

    legend_ws = wb.create_sheet(title="legend")
    legend_ws.append(["–†¬¶–†–Ü–†¬Ķ–°‚Äö", "HEX", "–†‚ÄĒ–†–Ö–†¬į–°‚Ä°–†¬Ķ–†–Ö–†—Ď–†¬Ķ"])
    legend_ws.append(["", "#fee2e2", "–†—ó–†—ē–°–É–°‚Äö–°‚Äě–†¬į–†—Ē–°‚Äö–°—ď–†—ė"])
    legend_ws.append(["", "#fed7aa", "–°–É–†¬Ķ–†—Ė–†—ē–†“Ď–†–Ö–°–Ź –†–Ö–†¬į –°–É–†¬Ķ–†—Ė–†—ē–†“Ď–†–Ö–°–Ź"])
    legend_ws.append(["", "#fef9c3", "–°–É–†¬Ķ–†—Ė–†—ē–†“Ď–†–Ö–°–Ź –†—ó–†—ē–°–É–†¬Ľ–†¬Ķ 15:00 –†–Ö–†¬į –†¬∑–†¬į–†–Ü–°‚Äö–°–ā–†¬į"])
    legend_rows = [
        ("\u0426\u0432\u0435\u0442", "HEX", "\u0417\u043d\u0430\u0447\u0435\u043d\u0438\u0435"),
        ("", "#fee2e2", "\u043f\u043e\u0441\u0442\u0444\u0430\u043a\u0442\u0443\u043c"),
        ("", "#fed7aa", "\u0441\u0435\u0433\u043e\u0434\u043d\u044f \u043d\u0430 \u0441\u0435\u0433\u043e\u0434\u043d\u044f"),
        ("", "#fef9c3", "\u0441\u0435\u0433\u043e\u0434\u043d\u044f \u043f\u043e\u0441\u043b\u0435 15:00 \u043d\u0430 \u0437\u0430\u0432\u0442\u0440\u0430"),
    ]
    for row_idx, row in enumerate(legend_rows, start=1):
        for col_idx, value in enumerate(row, start=1):
            legend_ws.cell(row=row_idx, column=col_idx, value=value)

    legend_ws.cell(row=2, column=1).fill = EXPORT_RED_FILL
    legend_ws.cell(row=3, column=1).fill = EXPORT_ORANGE_FILL
    legend_ws.cell(row=4, column=1).fill = EXPORT_YELLOW_FILL

    if variant_headers:
        ws.delete_cols(12, 3)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename_prefix = "my_bookings_start_end_export" if variant == "start-end" else "my_bookings_export"
    filename = f"{filename_prefix}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

@router.put("/{booking_id}/transport-sheet", response_model=schemas.BookingWithDetails)
def update_transport_sheet(
    booking_id: int,
    payload: schemas.BookingTransportSheetUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    """–†—õ–†¬Ī–†–Ö–†—ē–†–Ü–†—Ď–°‚Äö–°–ä –°‚Äö–°–ā–†¬į–†–Ö–°–É–†—ó–†—ē–°–ā–°‚Äö–†–Ö–°‚ÄĻ–†‚ĄĖ –†¬Ľ–†—Ď–°–É–°‚Äö –†“Ď–†¬Ľ–°–Ź –†¬Ī–°–ā–†—ē–†–Ö–†—Ď"""
    query = db.query(models.Booking).filter(models.Booking.id == booking_id)
    if current_user.role != models.UserRole.admin:
        query = query.filter(models.Booking.user_id == current_user.id)
    booking = query.first()

    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")

    if payload.transport_sheet and len(payload.transport_sheet) > 20:
        raise HTTPException(status_code=400, detail="Transport sheet must be at most 20 characters")

    booking.transport_sheet = payload.transport_sheet or None
    booking.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(booking)

    # –†‚Äô–°–É–†¬Ķ–†—Ė–†“Ď–†¬į –†–Ü–†—Ē–†¬Ľ–°–č–°‚Ä°–†¬į–†¬Ķ–†—ė –†“Ď–†¬į–†–Ö–†–Ö–°‚ÄĻ–†¬Ķ –†¬į–†–Ü–°‚Äö–†—ē–°–ā–†¬į, –°‚Ä°–°‚Äö–†—ē–†¬Ī–°‚ÄĻ –†–Ö–†¬į –°‚Äě–°–ā–†—ē–†–Ö–°‚Äö–†¬Ķ –†–Ü–†—Ď–†“Ď–†–Ö–†—ē, –†—Ē–°‚Äö–†—ē –°–É–†—ē–†¬∑–†“Ď–†¬į–†¬Ľ –†¬Ī–°–ā–†—ē–†–Ö–°–ä
    serialized = _serialize_booking(db, booking, include_user=True)
    if not serialized:
        raise HTTPException(status_code=500, detail="Booking slots not found")

    is_owner = booking.user_id == current_user.id
    serialized["is_owner"] = is_owner
    serialized["can_modify"] = is_owner or current_user.role == models.UserRole.admin
    
    return serialized

@router.get("/{booking_id}/slots")
def get_booking_slots(
    booking_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    """–†—ü–†—ē–†¬Ľ–°—ď–°‚Ä°–†—Ď–°‚Äö–°–ä –°–É–†¬Ľ–†—ē–°‚Äö–°‚ÄĻ –†—Ē–†—ē–†–Ö–†—Ē–°–ā–†¬Ķ–°‚Äö–†–Ö–†—ē–†‚ĄĖ –†¬∑–†¬į–†—ó–†—Ď–°–É–†—Ď"""
    booking = db.query(models.Booking).filter(
        models.Booking.id == booking_id
    ).first()
    
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")
    
    slots = db.query(models.TimeSlot, models.Dock).join(
        models.BookingTimeSlot, models.TimeSlot.id == models.BookingTimeSlot.time_slot_id
    ).join(
        models.Dock, models.TimeSlot.dock_id == models.Dock.id
    ).filter(models.BookingTimeSlot.booking_id == booking_id).all()
    
    result = []
    for slot, dock in slots:
        result.append({
            "id": slot.id,
            "dock_name": dock.name,
            "slot_date": slot.slot_date.isoformat(),
            "start_time": slot.start_time.strftime("%H:%M"),
            "end_time": slot.end_time.strftime("%H:%M"),
            "capacity": slot.capacity
        })
    
    return result

@router.delete("/{booking_id}")
def delete_booking(
    booking_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    """–†–ą–†“Ď–†¬į–†¬Ľ–†—Ď–°‚Äö–°–ä –†¬∑–†¬į–†—ó–†—Ď–°–É–°–ä (–°‚Äö–†—ē–†¬Ľ–°–ä–†—Ē–†—ē –†¬Ķ–°–É–†¬Ľ–†—Ď –†—ē–†–Ö–†¬į –†—ē–°‚Äö–†—ė–†¬Ķ–†–Ö–†¬Ķ–†–Ö–†¬į)"""
    booking = db.query(models.Booking).filter(
        models.Booking.id == booking_id,
        models.Booking.user_id == current_user.id
    ).first()
    
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")
    
    if booking.status == "confirmed":
        raise HTTPException(status_code=400, detail="Cannot delete confirmed booking. Cancel it first.")
    
    # –†–ą–†“Ď–†¬į–†¬Ľ–°–Ź–†¬Ķ–†—ė –°–É–†–Ü–°–Ź–†¬∑–†—Ď –°–É –°–É–†¬Ľ–†—ē–°‚Äö–†¬į–†—ė–†—Ď
    db.query(models.BookingTimeSlot).filter(models.BookingTimeSlot.booking_id == booking_id).delete()
    
    # –†–ą–†“Ď–†¬į–†¬Ľ–°–Ź–†¬Ķ–†—ė –†¬∑–†¬į–†—ó–†—Ď–°–É–°–ä
    db.delete(booking)
    db.commit()
    
    return {"message": "Booking deleted successfully"}


@router.get("/import/template")
def download_booking_import_template(
    direction: str,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    """
    XLSX-–°‚ā¨–†¬į–†¬Ī–†¬Ľ–†—ē–†–Ö –†“Ď–†¬Ľ–°–Ź –†—Ď–†—ė–†—ó–†—ē–°–ā–°‚Äö–†¬į –†¬Ī–°–ā–†—ē–†–Ö–†—Ď–°–ā–†—ē–†–Ü–†¬į–†–Ö–†—Ď–†‚ĄĖ. direction: in|out.
    """
    dir_normalized = direction.lower()
    if dir_normalized not in ("in", "out"):
        raise HTTPException(status_code=400, detail="direction must be 'in' or 'out'")
    try:
        direction_enum = models.BookingDirection(dir_normalized)
    except Exception:
        raise HTTPException(status_code=400, detail="direction must be 'in' or 'out'")

    wb = Workbook()
    ws = wb.active
    ws.title = "bookings"
    ws.append(["transport_sheet", "supplier_name", "cubes", "booking_date", "start_time", "transport_type", "vehicle_type", "object_name", "driver_full_name", "driver_phone"])
    ws.append(["TS-001", "–†—õ–†—õ–†—õ –†—ü–°–ā–†—Ď–†—ė–†¬Ķ–°–ā", "12.5", "2025-01-10", "09:00", "–†¬∑–†¬į–†—Ē–°—ď–†—ó–†—Ē–†¬į", "–†¬§–°—ď–°–ā–†¬į 20'", "–†—õ–†¬Ī–°—ď–°‚Ä¶–†—ē–†–Ü–†—ē", "–†¬ė–†–Ü–†¬į–†–Ö–†—ē–†–Ü –†¬ė.–†¬ė.", "+7 999 000-00-00"])

    ws_sup = wb.create_sheet("suppliers")
    ws_sup.append(["supplier_name", "zone_name"])
    suppliers = db.query(models.Supplier).options(joinedload(models.Supplier.zone)).all()
    for s in suppliers:
        ws_sup.append([s.name, s.zone.name if s.zone else ""])

    ws_obj = wb.create_sheet("objects")
    ws_obj.append(["object_name"])
    for obj in db.query(models.Object).all():
        ws_obj.append([obj.name])

    ws_tt = wb.create_sheet("transport_types")
    ws_tt.append(["transport_type"])
    for t in db.query(models.TransportTypeRef).all():
        ws_tt.append([t.name])

    ws_vt = wb.create_sheet("vehicle_types")
    ws_vt.append(["vehicle_type"])
    for vt in db.query(models.VehicleType).all():
        ws_vt.append([vt.name])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="booking_import_template_{dir_normalized}.xlsx"'},
    )


@router.post("/import", response_model=schemas.BookingImportResult)
def import_bookings_from_excel(
    direction: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    """
    –†¬ė–†—ė–†—ó–†—ē–°–ā–°‚Äö –†¬Ī–°–ā–†—ē–†–Ö–†—Ď–°–ā–†—ē–†–Ü–†¬į–†–Ö–†—Ď–†‚ĄĖ –†—Ď–†¬∑ Excel. direction: in|out. –†‚Äô–†¬į–†¬Ľ–†—Ď–†“Ď–†–Ö–°‚ÄĻ–†¬Ķ –°–É–°‚Äö–°–ā–†—ē–†—Ē–†—Ď –°–É–†—ē–†¬∑–†“Ď–†¬į–°–č–°‚Äö–°–É–°–Ź, –†—ē–°‚ā¨–†—Ď–†¬Ī–†—Ē–†—Ď –†–Ü–†—ē–†¬∑–†–Ü–°–ā–†¬į–°‚Äį–†¬į–°–č–°‚Äö–°–É–°–Ź.
    """
    dir_normalized = direction.lower()
    if dir_normalized not in ("in", "out"):
        raise HTTPException(status_code=400, detail="direction must be 'in' or 'out'")

    # Normalize and validate direction to enum once, reuse below
    try:
        direction_enum = models.BookingDirection(dir_normalized)
    except Exception:
        raise HTTPException(status_code=400, detail="direction must be 'in' or 'out'")

    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="–†—õ–†¬∂–†—Ď–†“Ď–†¬į–†¬Ķ–°‚Äö–°–É–°–Ź Excel –°‚Äě–†¬į–†‚ĄĖ–†¬Ľ (.xlsx)")

    try:
        wb = load_workbook(BytesIO(file.file.read()))
        ws = wb.active
    except Exception:
        raise HTTPException(status_code=400, detail="–†—ú–†¬Ķ –°—ď–†“Ď–†¬į–†¬Ľ–†—ē–°–É–°–ä –†—ó–°–ā–†—ē–°‚Ä°–†—Ď–°‚Äö–†¬į–°‚Äö–°–ä Excel –°‚Äě–†¬į–†‚ĄĖ–†¬Ľ")

    expected_headers = ["transport_sheet", "supplier_name", "cubes", "booking_date", "start_time", "transport_type", "vehicle_type", "object_name", "driver_full_name", "driver_phone"]
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    if [h.lower() for h in headers] != expected_headers:
        raise HTTPException(status_code=400, detail=f"–†—õ–†¬∂–†—Ď–†“Ď–†¬į–†¬Ķ–°‚Äö–°–É–°–Ź –†¬∑–†¬į–†—Ė–†—ē–†¬Ľ–†—ē–†–Ü–†—ē–†—Ē: {', '.join(expected_headers)}")

    suppliers = db.query(models.Supplier).options(joinedload(models.Supplier.zone)).all()
    supplier_map = {s.name.strip().lower(): s for s in suppliers}

    objects = db.query(models.Object).all()
    object_map = {o.name.strip().lower(): o for o in objects}

    transport_types = db.query(models.TransportTypeRef).all()
    transport_type_map = {t.name.strip().lower(): t for t in transport_types}

    vehicle_types = db.query(models.VehicleType).all()
    vehicle_type_map = {v.name.strip().lower(): v for v in vehicle_types}

    docks = db.query(models.Dock).options(joinedload(models.Dock.available_zones)).all()
    allowed_types = [models.DockType.universal]
    if dir_normalized == "in":
        allowed_types.append(models.DockType.entrance)
    else:
        allowed_types.append(models.DockType.exit)

    docks_by_object = {}
    for d in docks:
        if d.dock_type not in allowed_types:
            continue
        docks_by_object.setdefault(d.object_id, []).append(d)
    for lst in docks_by_object.values():
        lst.sort(key=lambda x: x.name or "")

    errors: list[schemas.BookingImportError] = []
    created = 0
    provisional_usage: dict[tuple[int, int, str, datetime.date], float] = {}

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        raw_transport_sheet, raw_supplier, raw_cubes, raw_date, raw_time, raw_transport_type, raw_vehicle_type, raw_object, raw_driver_name, raw_driver_phone = row
        transport_sheet = (raw_transport_sheet or "").strip()
        supplier_name = (raw_supplier or "").strip()
        cubes = None if raw_cubes in (None, "") else float(raw_cubes)
        booking_date_str = (raw_date or "").strip() if isinstance(raw_date, str) else (raw_date.strftime("%Y-%m-%d") if hasattr(raw_date, "strftime") else "")
        start_time_str = (raw_time or "").strip() if isinstance(raw_time, str) else (raw_time.strftime("%H:%M") if hasattr(raw_time, "strftime") else "")
        transport_type_name = (raw_transport_type or "").strip()
        vehicle_type_name = (raw_vehicle_type or "").strip()
        object_name = (raw_object or "").strip()
        driver_name = (raw_driver_name or "").strip()
        driver_phone = (raw_driver_phone or "").strip()

        row_errors = []
        if not transport_sheet:
            row_errors.append("transport_sheet –†—ē–†¬Ī–°–Ź–†¬∑–†¬į–°‚Äö–†¬Ķ–†¬Ľ–†¬Ķ–†–Ö")
        if not supplier_name:
            row_errors.append("supplier_name –†—ē–†¬Ī–°–Ź–†¬∑–†¬į–°‚Äö–†¬Ķ–†¬Ľ–†¬Ķ–†–Ö")
        if cubes is None:
            row_errors.append("cubes –†—ē–†¬Ī–°–Ź–†¬∑–†¬į–°‚Äö–†¬Ķ–†¬Ľ–†¬Ķ–†–Ö")
        if not booking_date_str:
            row_errors.append("booking_date –†—ē–†¬Ī–°–Ź–†¬∑–†¬į–°‚Äö–†¬Ķ–†¬Ľ–†¬Ķ–†–Ö")
        if not start_time_str:
            row_errors.append("start_time –†—ē–†¬Ī–°–Ź–†¬∑–†¬į–°‚Äö–†¬Ķ–†¬Ľ–†¬Ķ–†–Ö")
        if not transport_type_name:
            row_errors.append("transport_type –†—ē–†¬Ī–°–Ź–†¬∑–†¬į–°‚Äö–†¬Ķ–†¬Ľ–†¬Ķ–†–Ö")
        if not vehicle_type_name:
            row_errors.append("vehicle_type –†—ē–†¬Ī–°–Ź–†¬∑–†¬į–°‚Äö–†¬Ķ–†¬Ľ–†¬Ķ–†–Ö")
        if not object_name:
            row_errors.append("object_name –†—ē–†¬Ī–°–Ź–†¬∑–†¬į–°‚Äö–†¬Ķ–†¬Ľ–†¬Ķ–†–Ö")

        supplier = supplier_map.get(supplier_name.lower()) if supplier_name else None
        if supplier is None:
            row_errors.append(f"supplier '{supplier_name}' –†–Ö–†¬Ķ –†–Ö–†¬į–†‚ĄĖ–†“Ď–†¬Ķ–†–Ö")
        zone_id = supplier.zone_id if supplier else None

        obj = object_map.get(object_name.lower()) if object_name else None
        if obj is None:
            row_errors.append(f"object '{object_name}' –†–Ö–†¬Ķ –†–Ö–†¬į–†‚ĄĖ–†“Ď–†¬Ķ–†–Ö")

        transport_type = transport_type_map.get(transport_type_name.lower()) if transport_type_name else None
        if transport_type is None:
            row_errors.append(f"transport_type '{transport_type_name}' –†–Ö–†¬Ķ –†–Ö–†¬į–†‚ĄĖ–†“Ď–†¬Ķ–†–Ö")

        vehicle_type = vehicle_type_map.get(vehicle_type_name.lower()) if vehicle_type_name else None
        if vehicle_type is None:
            row_errors.append(f"vehicle_type '{vehicle_type_name}' –†–Ö–†¬Ķ –†–Ö–†¬į–†‚ĄĖ–†“Ď–†¬Ķ–†–Ö")

        try:
            booking_date = datetime.strptime(booking_date_str, "%Y-%m-%d").date()
        except Exception:
            row_errors.append(f"booking_date '{booking_date_str}' –†–Ö–†¬Ķ–†—Ē–†—ē–°–ā–°–ā–†¬Ķ–†—Ē–°‚Äö–†¬Ķ–†–Ö")
            booking_date = None

        try:
            start_time = datetime.strptime(start_time_str, "%H:%M").time()
        except Exception:
            row_errors.append(f"start_time '{start_time_str}' –†–Ö–†¬Ķ–†—Ē–†—ē–°–ā–°–ā–†¬Ķ–†—Ē–°‚Äö–†¬Ķ–†–Ö")
            start_time = None

        if row_errors:
            errors.append(schemas.BookingImportError(row_number=idx, message="; ".join(row_errors)))
            continue

        # duration
        try:
            duration = get_duration(
                object_id=obj.id,
                supplier_id=supplier.id if supplier else None,
                transport_type_id=transport_type.id if transport_type else None,
                vehicle_type_id=vehicle_type.id if vehicle_type else None,
                db=db
            ).duration_minutes
        except HTTPException as e:
            if e.status_code == 404:
                duration = vehicle_type.duration_minutes
            else:
                errors.append(schemas.BookingImportError(row_number=idx, message=e.detail or "–†—ú–†¬Ķ –°—ď–†“Ď–†¬į–†¬Ľ–†—ē–°–É–°–ä –†–Ü–°‚ÄĻ–°‚Ä°–†—Ď–°–É–†¬Ľ–†—Ď–°‚Äö–°–ä –†“Ď–†¬Ľ–†—Ď–°‚Äö–†¬Ķ–†¬Ľ–°–ä–†–Ö–†—ē–°–É–°‚Äö–°–ä"))
                continue

        if duration <= 0:
            errors.append(schemas.BookingImportError(row_number=idx, message="–†‚ÄĚ–†¬Ľ–†—Ď–°‚Äö–†¬Ķ–†¬Ľ–°–ä–†–Ö–†—ē–°–É–°‚Äö–°–ä –†“Ď–†—ē–†¬Ľ–†¬∂–†–Ö–†¬į –†¬Ī–°‚ÄĻ–°‚Äö–°–ä –†¬Ī–†—ē–†¬Ľ–°–ä–°‚ā¨–†¬Ķ 0"))
            continue

        required_slots = duration // 30 + (1 if duration % 30 != 0 else 0)
        candidate_docks = docks_by_object.get(obj.id, [])
        chosen_chain = None
        chosen_dock_id = None

        for dock in candidate_docks:
            # zone check
            if dock.available_zones:
                zone_ids = {z.id for z in dock.available_zones}
                if zone_id not in zone_ids:
                    continue

            slots = db.query(models.TimeSlot).filter(
                models.TimeSlot.dock_id == dock.id,
                models.TimeSlot.slot_date == booking_date
            ).order_by(models.TimeSlot.start_time).all()

            start_idx = next((i for i, s in enumerate(slots) if s.start_time == start_time), None)
            if start_idx is None:
                continue

            accumulated_minutes = 0
            chain: list[models.TimeSlot] = []
            dock_ok = True

            for s in slots[start_idx:]:
                occ = db.query(func.count(models.BookingTimeSlot.id)).join(
                    models.Booking, models.BookingTimeSlot.booking_id == models.Booking.id
                ).filter(
                    models.BookingTimeSlot.time_slot_id == s.id,
                    models.Booking.status == "confirmed"
                ).scalar() or 0
                if occ >= s.capacity:
                    dock_ok = False
                    break

                chain.append(s)
                slot_minutes = int((datetime.combine(s.slot_date, s.end_time) - datetime.combine(s.slot_date, s.start_time)).total_seconds() // 60)
                accumulated_minutes += slot_minutes

                if accumulated_minutes >= duration:
                    chosen_chain = chain
                    chosen_dock_id = dock.id
                    break

            if not dock_ok:
                continue
            if chosen_chain:
                break

        if not chosen_chain:
            errors.append(schemas.BookingImportError(row_number=idx, message="–†—ú–†¬Ķ–°‚Äö –°–É–†–Ü–†—ē–†¬Ī–†—ē–†“Ď–†–Ö–†—ē–†—Ė–†—ē –°–É–†¬Ľ–†—ē–°‚Äö–†¬į –†–Ö–†¬į –†—ē–†¬Ī–°–Č–†¬Ķ–†—Ē–°‚Äö–†¬Ķ –†“Ď–†¬Ľ–°–Ź –°–Ć–°‚Äö–†—ē–†‚ĄĖ –†¬∑–†—ē–†–Ö–°‚ÄĻ/–†–Ü–°–ā–†¬Ķ–†—ė–†¬Ķ–†–Ö–†—Ď"))
            continue


        quota = None
        total_quota_volume = None
        if transport_type:
            quota, total_quota_volume = get_quota_for_date(
                db=db,
                object_id=obj.id,
                transport_type_id=transport_type.id,
                target_date=booking_date,
                direction=direction_enum,
            )
        if quota and total_quota_volume is not None:
            if cubes is None:
                errors.append(schemas.BookingImportError(row_number=idx, message="cubes ?????????? ??? ??? ? ??????"))
                continue
            key = (obj.id, transport_type.id, direction_enum.value, booking_date)
            extra_used = provisional_usage.get(key, 0.0)
            used_volume = calculate_used_volume(
                db=db,
                object_id=obj.id,
                transport_type_id=transport_type.id,
                target_date=booking_date,
                direction=direction_enum,
            ) + extra_used
            remaining_volume = total_quota_volume - used_volume
            if not quota.allow_overbooking and cubes > remaining_volume:
                errors.append(
                    schemas.BookingImportError(
                        row_number=idx,
                        message=f"–†—ü–°–ā–†¬Ķ–†–Ü–°‚ÄĻ–°‚ā¨–†¬Ķ–†–Ö–†¬į –†—Ē–†–Ü–†—ē–°‚Äö–†¬į –†–Ö–†¬į {booking_date}. –†—õ–°–É–°‚Äö–†¬į–°‚Äö–†—ē–†—Ē {remaining_volume}, –†¬∑–†¬į–°–Ź–†–Ü–†¬Ľ–†¬Ķ–†–Ö–†—ē {cubes}",
                    )
                )
                continue

        new_booking = models.Booking(
            user_id=current_user.id,
            vehicle_type_id=vehicle_type.id,
            vehicle_plate="",
            driver_full_name=driver_name or "",
            driver_phone=driver_phone or "",
            status="confirmed",
            supplier_id=supplier.id if supplier else None,
            zone_id=zone_id,
            transport_type_id=transport_type.id if transport_type else None,
            cubes=cubes,
            transport_sheet=transport_sheet,
            booking_type=direction_enum,
        )
        db.add(new_booking)
        db.flush()

        for s in chosen_chain:
            db.add(models.BookingTimeSlot(booking_id=new_booking.id, time_slot_id=s.id))

        # Flush immediately so subsequent rows in the same import see the newly occupied slots
        # when they calculate availability/capacity.
        db.flush()

        if quota and total_quota_volume is not None and cubes is not None and transport_type:
            key = (obj.id, transport_type.id, direction_enum.value, booking_date)
            provisional_usage[key] = provisional_usage.get(key, 0.0) + cubes

        created += 1

    if created:
        db.commit()
    else:
        db.rollback()

    return schemas.BookingImportResult(created=created, errors=errors)

