from collections import Counter, defaultdict
from datetime import date, datetime, time, timedelta
from io import BytesIO
from typing import Any

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from fastapi.encoders import jsonable_encoder
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.utils.datetime import from_excel
from sqlalchemy import and_
from sqlalchemy.orm import Session, joinedload

from .. import models, schemas
from ..db import get_db
from ..deps import get_current_user
from .bookings import _serialize_booking

router = APIRouter()


def _require_admin(user: models.User) -> None:
    if user.role != models.UserRole.admin:
        raise HTTPException(status_code=403, detail="Недостаточно прав")


def normalize_tl_number(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().upper()


def _format_comparison_date(value: date) -> str:
    return value.isoformat()


def _format_comparison_time(value: time) -> str:
    return value.replace(microsecond=0).isoformat()


def _parse_file_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, time):
        return None
    if isinstance(value, (int, float)):
        try:
            parsed = from_excel(value)
            if isinstance(parsed, datetime):
                return parsed.date()
            if isinstance(parsed, date) and not isinstance(parsed, datetime):
                return parsed
        except Exception:
            return None
    text = str(value).strip()
    if not text:
        return None
    normalized = text.replace("T", " ")
    for fmt in (
        "%Y-%m-%d",
        "%d.%m.%Y",
        "%d/%m/%Y",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%d.%m.%Y %H:%M:%S",
        "%d.%m.%Y %H:%M",
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
    ):
        try:
            return datetime.strptime(normalized, fmt).date()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text).date()
    except ValueError:
        return None


def _parse_file_time(value: Any) -> time | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.time().replace(microsecond=0)
    if isinstance(value, time):
        return value.replace(microsecond=0)
    if isinstance(value, date):
        return None
    if isinstance(value, (int, float)):
        try:
            parsed = from_excel(value)
            if isinstance(parsed, datetime):
                return parsed.time().replace(microsecond=0)
            if isinstance(parsed, time):
                return parsed.replace(microsecond=0)
        except Exception:
            return None
    text = str(value).strip()
    if not text:
        return None
    normalized = text.replace("T", " ")
    for fmt in (
        "%H:%M:%S",
        "%H:%M",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%d.%m.%Y %H:%M:%S",
        "%d.%m.%Y %H:%M",
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
    ):
        try:
            return datetime.strptime(normalized, fmt).time().replace(microsecond=0)
        except ValueError:
            continue
    try:
        return time.fromisoformat(text).replace(microsecond=0)
    except ValueError:
        try:
            return datetime.fromisoformat(text).time().replace(microsecond=0)
        except ValueError:
            return None


def _yms_booking_date(yms_data: dict | None) -> date | None:
    if not isinstance(yms_data, dict) or not yms_data.get("booking_date"):
        return None
    try:
        return date.fromisoformat(str(yms_data["booking_date"]))
    except ValueError:
        return None


def _yms_start_time(yms_data: dict | None) -> time | None:
    if not isinstance(yms_data, dict) or not yms_data.get("start_time"):
        return None
    try:
        return time.fromisoformat(str(yms_data["start_time"])).replace(microsecond=0)
    except ValueError:
        return None


def _date_time_differences(file_date: date | None, file_time: time | None, yms_data: dict | None) -> list[dict]:
    differences = []
    if file_date is not None:
        yms_date = _yms_booking_date(yms_data)
        if yms_date is not None and file_date != yms_date:
            differences.append({
                "field": "booking_date",
                "label": "Дата записи",
                "file_value": _format_comparison_date(file_date),
                "yms_value": _format_comparison_date(yms_date),
                "message": "Дата в файле не совпадает с записью YMS",
            })
    if file_time is not None:
        yms_time = _yms_start_time(yms_data)
        if yms_time is not None and file_time != yms_time:
            differences.append({
                "field": "start_time",
                "label": "Время записи",
                "file_value": _format_comparison_time(file_time),
                "yms_value": _format_comparison_time(yms_time),
                "message": "Время в файле не совпадает с записью YMS",
            })
    return differences


def _file_data_with_comparison_fields(file_row: dict) -> dict:
    file_data = dict(file_row.get("data") or {})
    if file_row.get("file_date") is not None:
        file_data["Дата записи (файл)"] = _format_comparison_date(file_row["file_date"])
    if file_row.get("file_time") is not None:
        file_data["Время записи (файл)"] = _format_comparison_time(file_row["file_time"])
    return file_data


def _profile_to_dict(profile: models.DataComparisonProfile) -> dict:
    return {
        "id": profile.id,
        "name": profile.name,
        "object_id": profile.object_id,
        "object_name": profile.object.name if profile.object else None,
        "direction": profile.direction.value if hasattr(profile.direction, "value") else profile.direction,
        "tl_column_name": profile.tl_column_name,
        "tl_column_letter": profile.tl_column_letter,
        "file_start_row": profile.file_start_row,
        "file_end_row": profile.file_end_row,
        "status_filters": profile.status_filters or [],
        "yms_filters": profile.yms_filters or {},
        "file_settings": profile.file_settings or {},
        "comparison_settings": profile.comparison_settings or {},
        "is_active": profile.is_active,
        "created_at": profile.created_at.isoformat() if profile.created_at else None,
        "updated_at": profile.updated_at.isoformat() if profile.updated_at else None,
    }


def _row_to_dict(row: models.DataComparisonRunRow) -> dict:
    return {
        "id": row.id,
        "run_id": row.run_id,
        "tl_number_original": row.tl_number_original,
        "tl_number_normalized": row.tl_number_normalized,
        "status": row.status,
        "file_row_number": row.file_row_number,
        "booking_id": row.booking_id,
        "file_data": row.file_data,
        "yms_data": row.yms_data,
        "differences": row.differences or [],
    }


def _run_to_dict(run: models.DataComparisonRun, include_rows: bool = False) -> dict:
    payload = {
        "id": run.id,
        "profile_id": run.profile_id,
        "profile_name": run.profile.name if run.profile else None,
        "user_id": run.user_id,
        "date_from": run.date_from.isoformat(),
        "date_to": run.date_to.isoformat(),
        "extended_date_from": run.extended_date_from.isoformat(),
        "extended_date_to": run.extended_date_to.isoformat(),
        "source_file_name": run.source_file_name,
        "status": run.status,
        "summary": run.summary or {},
        "error_message": run.error_message,
        "created_at": run.created_at.isoformat() if run.created_at else None,
        "started_at": run.started_at.isoformat() if run.started_at else None,
        "finished_at": run.finished_at.isoformat() if run.finished_at else None,
    }
    if include_rows:
        payload["rows"] = [_row_to_dict(row) for row in run.rows]
    return payload


def _normalize_excel_column_letter(value: str | None) -> str | None:
    if value is None:
        return None
    normalized = str(value).strip().upper()
    if not normalized:
        return None
    try:
        column_index_from_string(normalized)
    except ValueError:
        raise HTTPException(status_code=400, detail="Столбец с номером ТЛ должен быть буквой Excel, например A, G или AA")
    return normalized


def _parse_snapshot_column_indices(snapshot_columns: Any, max_column: int) -> set[int] | None:
    if snapshot_columns is None:
        return None
    if isinstance(snapshot_columns, str):
        raw_parts = [part.strip().upper() for part in snapshot_columns.split(",")]
    elif isinstance(snapshot_columns, list):
        raw_parts = [str(part).strip().upper() for part in snapshot_columns]
    else:
        raise HTTPException(status_code=400, detail="Колонки для снимка Excel должны быть строкой A:G или списком колонок")

    parts = [part for part in raw_parts if part]
    if not parts:
        return None

    indices: set[int] = set()
    for part in parts:
        if ":" in part:
            start, end = [value.strip() for value in part.split(":", 1)]
            try:
                start_idx = column_index_from_string(start)
                end_idx = column_index_from_string(end)
            except ValueError:
                raise HTTPException(status_code=400, detail="Колонки для снимка Excel должны быть в формате A:G или A,C,D")
            if end_idx < start_idx:
                raise HTTPException(status_code=400, detail="Конец диапазона колонок Excel не может быть раньше начала")
            indices.update(range(start_idx, end_idx + 1))
        else:
            try:
                indices.add(column_index_from_string(part))
            except ValueError:
                raise HTTPException(status_code=400, detail="Колонки для снимка Excel должны быть в формате A:G или A,C,D")

    return {idx for idx in indices if 1 <= idx <= max_column}


def _worksheet_row_values(ws, row_number: int, allowed_column_indices: set[int] | None = None) -> dict:
    column_indices = range(1, ws.max_column + 1) if allowed_column_indices is None else sorted(allowed_column_indices)
    return {
        ws.cell(row=1, column=col_idx).value or f"Колонка {col_idx}": ws.cell(row=row_number, column=col_idx).value
        for col_idx in column_indices
        if col_idx <= ws.max_column
    }


def _parse_xlsx_rows(
    content: bytes,
    tl_column_name: str,
    tl_column_letter: str | None = None,
    file_start_row: int = 2,
    file_end_row: int | None = None,
    snapshot_columns: Any = None,
    date_column_letter: str | None = None,
    time_column_letter: str | None = None,
) -> list[dict]:
    wb = load_workbook(BytesIO(content), data_only=True)
    ws = wb.active
    snapshot_column_indices = _parse_snapshot_column_indices(snapshot_columns, ws.max_column)
    start_row = file_start_row or 2
    if start_row < 1:
        raise HTTPException(status_code=400, detail="Строка начала должна быть не меньше 1")
    if file_end_row is not None and file_end_row < start_row:
        raise HTTPException(status_code=400, detail="Строка окончания не может быть меньше строки начала")

    normalized_letter = _normalize_excel_column_letter(tl_column_letter)
    normalized_date_letter = _normalize_excel_column_letter(date_column_letter)
    normalized_time_letter = _normalize_excel_column_letter(time_column_letter)
    date_column_index = column_index_from_string(normalized_date_letter) if normalized_date_letter else None
    time_column_index = column_index_from_string(normalized_time_letter) if normalized_time_letter else None
    if date_column_index and date_column_index > ws.max_column:
        raise HTTPException(status_code=400, detail=f"В файле нет столбца даты '{normalized_date_letter}'")
    if time_column_index and time_column_index > ws.max_column:
        raise HTTPException(status_code=400, detail=f"В файле нет столбца времени '{normalized_time_letter}'")
    parsed = []
    if normalized_letter:
        column_index = column_index_from_string(normalized_letter)
        end_row = file_end_row or ws.max_row
        if column_index > ws.max_column:
            raise HTTPException(status_code=400, detail=f"В файле нет столбца '{normalized_letter}'")
        for row_number in range(start_row, end_row + 1):
            tl_original = ws.cell(row=row_number, column=column_index).value
            tl_normalized = normalize_tl_number(tl_original)
            if not tl_normalized:
                continue
            row_values = _worksheet_row_values(ws, row_number, snapshot_column_indices)
            file_date = _parse_file_date(ws.cell(row=row_number, column=date_column_index).value) if date_column_index else None
            file_time = _parse_file_time(ws.cell(row=row_number, column=time_column_index).value) if time_column_index else None
            parsed.append({
                "row_number": row_number,
                "tl_original": str(tl_original).strip() if tl_original is not None else "",
                "tl_normalized": tl_normalized,
                "data": row_values,
                "file_date": file_date,
                "file_time": file_time,
            })
        return parsed

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    if tl_column_name not in headers:
        raise HTTPException(status_code=400, detail=f"В файле не найдена колонка '{tl_column_name}'")
    end_index = file_end_row or len(rows)
    for idx, values in enumerate(rows[start_row - 1:end_index], start=start_row):
        data = {
            headers[col_idx]: values[col_idx] if col_idx < len(values) else None
            for col_idx in range(len(headers))
            if headers[col_idx] and (snapshot_column_indices is None or col_idx + 1 in snapshot_column_indices)
        }
        tl_original = values[headers.index(tl_column_name)] if headers.index(tl_column_name) < len(values) else None
        file_date = _parse_file_date(values[date_column_index - 1]) if date_column_index and date_column_index - 1 < len(values) else None
        file_time = _parse_file_time(values[time_column_index - 1]) if time_column_index and time_column_index - 1 < len(values) else None
        tl_normalized = normalize_tl_number(tl_original)
        if not tl_normalized:
            continue
        parsed.append({
            "row_number": idx,
            "tl_original": str(tl_original).strip() if tl_original is not None else "",
            "tl_normalized": tl_normalized,
            "data": data,
            "file_date": file_date,
            "file_time": file_time,
        })
    return parsed


def _booking_query(db: Session, profile: models.DataComparisonProfile, start_date: date, end_date: date):
    statuses = profile.status_filters or ["confirmed"]
    direction_value = profile.direction.value if hasattr(profile.direction, "value") else profile.direction
    bookings = (
        db.query(models.Booking)
        .join(models.BookingTimeSlot, models.Booking.id == models.BookingTimeSlot.booking_id)
        .join(models.TimeSlot, models.BookingTimeSlot.time_slot_id == models.TimeSlot.id)
        .join(models.Dock, models.TimeSlot.dock_id == models.Dock.id)
        .options(
            joinedload(models.Booking.user),
            joinedload(models.Booking.vehicle_type),
            joinedload(models.Booking.supplier),
            joinedload(models.Booking.zone),
            joinedload(models.Booking.transport_type),
        )
        .filter(
            models.Dock.object_id == profile.object_id,
            models.Booking.booking_type == models.BookingDirection(direction_value),
            models.Booking.status.in_(statuses),
            models.TimeSlot.slot_date >= start_date,
            models.TimeSlot.slot_date <= end_date,
        )
        .all()
    )
    deduped = {}
    for booking in bookings:
        deduped[booking.id] = booking
    return list(deduped.values())


def _index_bookings(db: Session, bookings: list[models.Booking]) -> dict[str, list[dict]]:
    by_tl: dict[str, list[dict]] = defaultdict(list)
    for booking in bookings:
        normalized = normalize_tl_number(booking.transport_sheet)
        if not normalized:
            continue
        serialized = _serialize_booking(db, booking, include_user=True)
        if serialized:
            by_tl[normalized].append(serialized)
    return by_tl


@router.get("/profiles")
def list_profiles(db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    _require_admin(current_user)
    profiles = db.query(models.DataComparisonProfile).options(joinedload(models.DataComparisonProfile.object)).order_by(models.DataComparisonProfile.name).all()
    return [_profile_to_dict(profile) for profile in profiles]


@router.post("/profiles")
def create_profile(payload: schemas.DataComparisonProfileCreate, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    _require_admin(current_user)
    obj = db.query(models.Object).filter(models.Object.id == payload.object_id).first()
    if not obj:
        raise HTTPException(status_code=404, detail="Object not found")
    try:
        direction = models.BookingDirection(payload.direction)
    except Exception:
        raise HTTPException(status_code=400, detail="direction must be 'in' or 'out'")
    tl_column_letter = _normalize_excel_column_letter(payload.tl_column_letter)
    if payload.file_start_row < 1:
        raise HTTPException(status_code=400, detail="Строка начала должна быть не меньше 1")
    if payload.file_end_row is not None and payload.file_end_row < payload.file_start_row:
        raise HTTPException(status_code=400, detail="Строка окончания не может быть меньше строки начала")
    profile = models.DataComparisonProfile(
        name=payload.name,
        object_id=payload.object_id,
        direction=direction,
        tl_column_name=payload.tl_column_name,
        tl_column_letter=tl_column_letter,
        file_start_row=payload.file_start_row,
        file_end_row=payload.file_end_row,
        status_filters=payload.status_filters or ["confirmed"],
        yms_filters=payload.yms_filters or {},
        file_settings=payload.file_settings or {},
        comparison_settings=payload.comparison_settings or {},
        is_active=payload.is_active,
    )
    db.add(profile)
    db.commit()
    db.refresh(profile)
    return _profile_to_dict(profile)


@router.put("/profiles/{profile_id}")
def update_profile(profile_id: int, payload: schemas.DataComparisonProfileUpdate, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    _require_admin(current_user)
    profile = db.query(models.DataComparisonProfile).filter(models.DataComparisonProfile.id == profile_id).first()
    if not profile:
        raise HTTPException(status_code=404, detail="Профиль сверки не найден")
    obj = db.query(models.Object).filter(models.Object.id == payload.object_id).first()
    if not obj:
        raise HTTPException(status_code=404, detail="Object not found")
    try:
        direction = models.BookingDirection(payload.direction)
    except Exception:
        raise HTTPException(status_code=400, detail="direction must be 'in' or 'out'")
    tl_column_letter = _normalize_excel_column_letter(payload.tl_column_letter)

    profile.name = payload.name
    profile.object_id = payload.object_id
    profile.direction = direction
    profile.tl_column_name = payload.tl_column_name
    profile.tl_column_letter = tl_column_letter
    # Row range is intentionally not edited here by the UI; keep stored defaults for legacy clients only.
    profile.status_filters = payload.status_filters or ["confirmed"]
    profile.yms_filters = payload.yms_filters or {}
    profile.file_settings = payload.file_settings or {}
    profile.comparison_settings = payload.comparison_settings or {}
    profile.is_active = payload.is_active

    db.commit()
    db.refresh(profile)
    return _profile_to_dict(profile)


@router.post("/runs")
async def create_run(
    profile_id: int = Form(...),
    date_from: date = Form(...),
    date_to: date = Form(...),
    file_start_row: int | None = Form(None),
    file_end_row: int | None = Form(None),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    _require_admin(current_user)
    if date_from > date_to:
        raise HTTPException(status_code=400, detail="Дата с не может быть позже даты по")
    profile = db.query(models.DataComparisonProfile).filter(models.DataComparisonProfile.id == profile_id, models.DataComparisonProfile.is_active == True).first()
    if not profile:
        raise HTTPException(status_code=404, detail="Профиль сверки не найден")
    start_row = file_start_row if file_start_row is not None else profile.file_start_row
    end_row = file_end_row if file_end_row is not None else profile.file_end_row
    if start_row < 1:
        raise HTTPException(status_code=400, detail="Строка начала должна быть не меньше 1")
    if end_row is not None and end_row < start_row:
        raise HTTPException(status_code=400, detail="Строка окончания не может быть меньше строки начала")
    content = await file.read()
    file_settings = profile.file_settings or {}
    file_rows = _parse_xlsx_rows(
        content,
        profile.tl_column_name,
        profile.tl_column_letter,
        start_row,
        end_row,
        file_settings.get("snapshot_columns"),
        file_settings.get("date_column_letter"),
        file_settings.get("time_column_letter"),
    )
    extended_from = date_from - timedelta(days=2)
    extended_to = date_to + timedelta(days=2)

    primary_bookings = _booking_query(db, profile, date_from, date_to)
    extended_bookings = _booking_query(db, profile, extended_from, extended_to)
    primary_by_tl = _index_bookings(db, primary_bookings)
    extended_by_tl = _index_bookings(db, extended_bookings)

    run = models.DataComparisonRun(
        profile_id=profile.id,
        user_id=current_user.id,
        date_from=date_from,
        date_to=date_to,
        extended_date_from=extended_from,
        extended_date_to=extended_to,
        source_file_name=file.filename or "uploaded.xlsx",
        status="completed",
        started_at=datetime.utcnow(),
        finished_at=datetime.utcnow(),
        summary={},
    )
    db.add(run)
    db.flush()

    summary = {
        "file_rows": len(file_rows),
        "unique_file_tl": len({row["tl_normalized"] for row in file_rows}),
        "yms_rows": sum(len(v) for v in primary_by_tl.values()),
        "matched": 0,
        "found_in_yms_extended_period": 0,
        "missing_in_yms": 0,
        "missing_in_file": 0,
        "duplicate_in_file": 0,
        "duplicate_in_yms": 0,
    }
    datetime_comparison_enabled = bool(file_settings.get("date_column_letter") or file_settings.get("time_column_letter"))
    if datetime_comparison_enabled:
        summary.update({
            "field_mismatch": 0,
            "datetime_matched": 0,
            "datetime_mismatch": 0,
        })
    rows_to_save = []
    file_counts = Counter(row["tl_normalized"] for row in file_rows)
    duplicate_file_tl = {tl for tl, count in file_counts.items() if count > 1}
    seen_file_tl = set()

    for file_row in file_rows:
        tl = file_row["tl_normalized"]
        seen_file_tl.add(tl)
        if tl in duplicate_file_tl:
            status = "duplicate_in_file"
            yms_data = None
            booking_id = None
        elif len(primary_by_tl.get(tl, [])) > 1:
            status = "duplicate_in_yms"
            yms_data = primary_by_tl[tl]
            booking_id = None
        elif primary_by_tl.get(tl):
            status = "matched"
            yms_data = primary_by_tl[tl][0]
            booking_id = yms_data.get("id")
        elif extended_by_tl.get(tl):
            status = "found_in_yms_extended_period"
            yms_data = extended_by_tl[tl][0] if len(extended_by_tl[tl]) == 1 else extended_by_tl[tl]
            booking_id = yms_data.get("id") if isinstance(yms_data, dict) else None
        else:
            status = "missing_in_yms"
            yms_data = None
            booking_id = None
        differences = []
        if status == "matched" and (file_row.get("file_date") is not None or file_row.get("file_time") is not None):
            differences = _date_time_differences(file_row.get("file_date"), file_row.get("file_time"), yms_data if isinstance(yms_data, dict) else None)
            if differences:
                status = "field_mismatch"
                summary["datetime_mismatch"] += 1
            else:
                summary["datetime_matched"] += 1
        summary[status] += 1
        rows_to_save.append(models.DataComparisonRunRow(
            run_id=run.id,
            tl_number_original=file_row["tl_original"],
            tl_number_normalized=tl,
            status=status,
            file_row_number=file_row["row_number"],
            booking_id=booking_id,
            file_data=jsonable_encoder(_file_data_with_comparison_fields(file_row)),
            yms_data=jsonable_encoder(yms_data),
            differences=jsonable_encoder(differences),
        ))

    for tl, bookings in primary_by_tl.items():
        if tl in seen_file_tl:
            continue
        status = "duplicate_in_yms" if len(bookings) > 1 else "missing_in_file"
        summary[status] += 1
        rows_to_save.append(models.DataComparisonRunRow(
            run_id=run.id,
            tl_number_original=bookings[0].get("transport_sheet") or tl,
            tl_number_normalized=tl,
            status=status,
            file_row_number=None,
            booking_id=bookings[0].get("id") if len(bookings) == 1 else None,
            file_data=None,
            yms_data=jsonable_encoder(bookings if len(bookings) > 1 else bookings[0]),
            differences=[],
        ))

    run.summary = summary
    db.add_all(rows_to_save)
    db.commit()
    db.refresh(run)
    return _run_to_dict(run, include_rows=True)


@router.get("/runs")
def list_runs(db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    _require_admin(current_user)
    runs = db.query(models.DataComparisonRun).options(joinedload(models.DataComparisonRun.profile)).order_by(models.DataComparisonRun.created_at.desc()).all()
    return [_run_to_dict(run) for run in runs]


@router.get("/runs/{run_id}")
def get_run(run_id: int, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    _require_admin(current_user)
    run = db.query(models.DataComparisonRun).options(joinedload(models.DataComparisonRun.profile), joinedload(models.DataComparisonRun.rows)).filter(models.DataComparisonRun.id == run_id).first()
    if not run:
        raise HTTPException(status_code=404, detail="Сверка не найдена")
    return _run_to_dict(run, include_rows=True)
