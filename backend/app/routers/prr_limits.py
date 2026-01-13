import json
from io import BytesIO
from typing import Optional, List, Dict, Tuple

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from openpyxl import Workbook, load_workbook

from app import models, schemas
from app.deps import get_db

router = APIRouter()


def _validate_duration(duration: int) -> None:
    if duration < 0 or duration % 30 != 0:
        raise HTTPException(
            status_code=400,
            detail="duration_minutes must be > 0 and multiple of 30",
        )


def _normalize(value: Optional[str]) -> str:
    return (value or "").strip().lower()


@router.get("/template")
def download_prr_limits_template(
    db: Session = Depends(get_db),
):
    wb = Workbook()
    ws = wb.active
    ws.title = "prr_limits"
    headers = ["object_name", "supplier_name", "transport_type", "vehicle_type", "duration_minutes"]
    ws.append(headers)
    ws.append(["Example Object", "", "Example Transport", "Example Vehicle", 60])

    ws_obj = wb.create_sheet("objects")
    ws_obj.append(["object_name"])
    for obj in db.query(models.Object).all():
        ws_obj.append([obj.name])

    ws_sup = wb.create_sheet("suppliers")
    ws_sup.append(["supplier_name"])
    for sup in db.query(models.Supplier).all():
        ws_sup.append([sup.name])

    ws_tt = wb.create_sheet("transport_types")
    ws_tt.append(["transport_type"])
    for tt in db.query(models.TransportTypeRef).all():
        ws_tt.append([tt.name])

    ws_vt = wb.create_sheet("vehicle_types")
    ws_vt.append(["vehicle_type"])
    for vt in db.query(models.VehicleType).all():
        ws_vt.append([vt.name])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="prr_limits_template.xlsx"'},
    )


@router.get("/export")
def export_prr_limits(
    db: Session = Depends(get_db),
):
    wb = Workbook()
    ws = wb.active
    ws.title = "prr_limits"
    headers = ["object_name", "supplier_name", "transport_type", "vehicle_type", "duration_minutes"]
    ws.append(headers)

    limits = db.query(models.PrrLimit).options(
        joinedload(models.PrrLimit.object),
        joinedload(models.PrrLimit.supplier),
        joinedload(models.PrrLimit.transport_type),
        joinedload(models.PrrLimit.vehicle_type),
    ).all()

    for lim in limits:
        ws.append([
            lim.object.name if lim.object else "",
            lim.supplier.name if lim.supplier else "",
            lim.transport_type.name if lim.transport_type else "",
            lim.vehicle_type.name if lim.vehicle_type else "",
            lim.duration_minutes,
        ])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="prr_limits_export.xlsx"'},
    )


@router.post("/import", response_model=schemas.PrrLimitImportResult)
def import_prr_limits(
    file: UploadFile = File(...),
    resolutions: Optional[str] = Form(None),
    db: Session = Depends(get_db),
):
    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Only Excel .xlsx/.xlsm files are supported")

    try:
        wb = load_workbook(BytesIO(file.file.read()))
    except Exception:
        raise HTTPException(status_code=400, detail="Failed to read Excel file")

    if "prr_limits" in wb.sheetnames:
        ws = wb["prr_limits"]
    else:
        ws = wb.active

    expected_headers = ["object_name", "supplier_name", "transport_type", "vehicle_type", "duration_minutes"]
    header_row = [str(cell.value).strip() if cell.value is not None else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    if [h.lower() for h in header_row] != expected_headers:
        raise HTTPException(status_code=400, detail=f"Invalid headers. Expected: {', '.join(expected_headers)}")

    # Maps for lookup
    object_map = {_normalize(o.name): o.id for o in db.query(models.Object).all()}
    supplier_map = {_normalize(s.name): s.id for s in db.query(models.Supplier).all()}
    transport_type_map = {_normalize(t.name): t.id for t in db.query(models.TransportTypeRef).all()}
    vehicle_type_map = {_normalize(v.name): v.id for v in db.query(models.VehicleType).all()}

    existing_limits = db.query(models.PrrLimit).all()
    key_to_limit = {
        (lim.object_id, lim.supplier_id, lim.transport_type_id, lim.vehicle_type_id): lim
        for lim in existing_limits
    }

    resolution_map: Dict[Tuple[str, str, str, str], str] = {}
    if resolutions:
        try:
            parsed = json.loads(resolutions)
            if isinstance(parsed, list):
                for item in parsed:
                    action = item.get("action")
                    if action not in ("keep_existing", "replace_with_new"):
                        continue
                    key = (
                        _normalize(item.get("object_name")),
                        _normalize(item.get("supplier_name")),
                        _normalize(item.get("transport_type")),
                        _normalize(item.get("vehicle_type")),
                    )
                    resolution_map[key] = action
            else:
                raise ValueError
        except Exception:
            raise HTTPException(status_code=400, detail="Invalid resolutions payload")

    errors: List[schemas.PrrLimitImportError] = []
    conflicts: List[schemas.PrrLimitImportConflict] = []
    unresolved_conflict = False

    pending_rows: Dict[Tuple[int, Optional[int], Optional[int], Optional[int]], Dict] = {}

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        raw_obj, raw_supplier, raw_tt, raw_vt, raw_duration = row
        obj_name = (raw_obj or "").strip()
        supplier_name = (raw_supplier or "").strip()
        tt_name = (raw_tt or "").strip()
        vt_name = (raw_vt or "").strip()

        row_errors = []
        obj_id = object_map.get(_normalize(obj_name))
        if not obj_id:
            row_errors.append(f"object '{obj_name}' not found")

        supplier_id = None
        if supplier_name:
            supplier_id = supplier_map.get(_normalize(supplier_name))
            if supplier_id is None:
                row_errors.append(f"supplier '{supplier_name}' not found")

        tt_id = None
        if tt_name:
            tt_id = transport_type_map.get(_normalize(tt_name))
            if tt_id is None:
                row_errors.append(f"transport_type '{tt_name}' not found")

        vt_id = None
        if vt_name:
            vt_id = vehicle_type_map.get(_normalize(vt_name))
            if vt_id is None:
                row_errors.append(f"vehicle_type '{vt_name}' not found")

        try:
            duration = int(raw_duration)
        except Exception:
            row_errors.append("duration_minutes must be integer")
            duration = None

        if duration is not None:
            if duration <= 0 or duration % 30 != 0:
                row_errors.append("duration_minutes must be >0 and multiple of 30")

        if row_errors:
            errors.append(schemas.PrrLimitImportError(row_number=idx, message="; ".join(row_errors)))
            continue

        key_ids = (obj_id, supplier_id, tt_id, vt_id)
        key_names = (_normalize(obj_name), _normalize(supplier_name), _normalize(tt_name), _normalize(vt_name))
        action = resolution_map.get(key_names)

        if key_ids in pending_rows:
            if action == "replace_with_new":
                pending_rows[key_ids] = {
                    "object_id": obj_id,
                    "supplier_id": supplier_id,
                    "transport_type_id": tt_id,
                    "vehicle_type_id": vt_id,
                    "duration": duration,
                }
            elif action == "keep_existing":
                pass
            else:
                conflicts.append(schemas.PrrLimitImportConflict(
                    row_number=idx,
                    object_name=obj_name,
                    supplier_name=supplier_name or None,
                    transport_type=tt_name or None,
                    vehicle_type=vt_name or None,
                    existing_duration=pending_rows[key_ids]["duration"],
                    new_duration=duration,
                    source="file",
                ))
                unresolved_conflict = True
            continue

        existing = key_to_limit.get(key_ids)
        if existing:
            if action == "replace_with_new":
                pending_rows[key_ids] = {
                    "object_id": obj_id,
                    "supplier_id": supplier_id,
                    "transport_type_id": tt_id,
                    "vehicle_type_id": vt_id,
                    "duration": duration,
                }
            elif action == "keep_existing":
                pass
            else:
                conflicts.append(schemas.PrrLimitImportConflict(
                    row_number=idx,
                    object_name=obj_name,
                    supplier_name=supplier_name or None,
                    transport_type=tt_name or None,
                    vehicle_type=vt_name or None,
                    existing_duration=existing.duration_minutes,
                    new_duration=duration,
                    source="database",
                ))
                unresolved_conflict = True
            continue

        pending_rows[key_ids] = {
            "object_id": obj_id,
            "supplier_id": supplier_id,
            "transport_type_id": tt_id,
            "vehicle_type_id": vt_id,
            "duration": duration,
        }

    if unresolved_conflict:
        return schemas.PrrLimitImportResult(created=0, updated=0, errors=errors, conflicts=conflicts)

    created = 0
    updated = 0
    for key, payload in pending_rows.items():
        existing = key_to_limit.get(key)
        if existing:
            existing.duration_minutes = payload["duration"]
            updated += 1
        else:
            db.add(models.PrrLimit(
                object_id=payload["object_id"],
                supplier_id=payload["supplier_id"],
                transport_type_id=payload["transport_type_id"],
                vehicle_type_id=payload["vehicle_type_id"],
                duration_minutes=payload["duration"],
            ))
            created += 1

    if created or updated:
        db.commit()
    else:
        db.rollback()

    return schemas.PrrLimitImportResult(
        created=created,
        updated=updated,
        errors=errors,
        conflicts=conflicts,
    )


@router.get("/duration/", response_model=schemas.PrrLimit)
def get_duration(
    object_id: int,
    supplier_id: Optional[int] = None,
    transport_type_id: Optional[int] = None,
    vehicle_type_id: Optional[int] = None,
    db: Session = Depends(get_db)
):
    # 1. Exact match
    query = db.query(models.PrrLimit).filter(
        models.PrrLimit.object_id == object_id,
        models.PrrLimit.supplier_id == supplier_id,
        models.PrrLimit.transport_type_id == transport_type_id,
        models.PrrLimit.vehicle_type_id == vehicle_type_id
    )
    result = query.first()
    if result:
        return result

    # 2. Less specific rules
    # object + supplier + transport
    query = db.query(models.PrrLimit).filter(
        models.PrrLimit.object_id == object_id,
        models.PrrLimit.supplier_id == supplier_id,
        models.PrrLimit.transport_type_id == transport_type_id,
        models.PrrLimit.vehicle_type_id == None
    )
    result = query.first()
    if result:
        return result

    # object + supplier + vehicle
    query = db.query(models.PrrLimit).filter(
        models.PrrLimit.object_id == object_id,
        models.PrrLimit.supplier_id == supplier_id,
        models.PrrLimit.transport_type_id == None,
        models.PrrLimit.vehicle_type_id == vehicle_type_id
    )
    result = query.first()
    if result:
        return result

    # object + transport + vehicle
    query = db.query(models.PrrLimit).filter(
        models.PrrLimit.object_id == object_id,
        models.PrrLimit.supplier_id == None,
        models.PrrLimit.transport_type_id == transport_type_id,
        models.PrrLimit.vehicle_type_id == vehicle_type_id
    )
    result = query.first()
    if result:
        return result

    # object + supplier
    query = db.query(models.PrrLimit).filter(
        models.PrrLimit.object_id == object_id,
        models.PrrLimit.supplier_id == supplier_id,
        models.PrrLimit.transport_type_id == None,
        models.PrrLimit.vehicle_type_id == None
    )
    result = query.first()
    if result:
        return result

    # object + transport
    query = db.query(models.PrrLimit).filter(
        models.PrrLimit.object_id == object_id,
        models.PrrLimit.supplier_id == None,
        models.PrrLimit.transport_type_id == transport_type_id,
        models.PrrLimit.vehicle_type_id == None
    )
    result = query.first()
    if result:
        return result

    # object + vehicle
    query = db.query(models.PrrLimit).filter(
        models.PrrLimit.object_id == object_id,
        models.PrrLimit.supplier_id == None,
        models.PrrLimit.transport_type_id == None,
        models.PrrLimit.vehicle_type_id == vehicle_type_id
    )
    result = query.first()
    if result:
        return result

    # object only
    query = db.query(models.PrrLimit).filter(
        models.PrrLimit.object_id == object_id,
        models.PrrLimit.supplier_id == None,
        models.PrrLimit.transport_type_id == None,
        models.PrrLimit.vehicle_type_id == None
    )
    result = query.first()
    if result:
        return result

    raise HTTPException(status_code=404, detail="No matching PRR limit found")


@router.post("/", response_model=schemas.PrrLimit)
def create_prr_limit(prr_limit: schemas.PrrLimitCreate, db: Session = Depends(get_db)):
    _validate_duration(prr_limit.duration_minutes)
    db_prr_limit = models.PrrLimit(**prr_limit.dict())
    db.add(db_prr_limit)
    db.commit()
    db.refresh(db_prr_limit)
    return db_prr_limit


@router.get("/", response_model=List[schemas.PrrLimit])
def read_prr_limits(skip: int = 0, limit: int = 100, db: Session = Depends(get_db)):
    prr_limits = db.query(models.PrrLimit).offset(skip).limit(limit).all()
    return prr_limits


@router.get("/{prr_limit_id}", response_model=schemas.PrrLimit)
def read_prr_limit(prr_limit_id: int, db: Session = Depends(get_db)):
    db_prr_limit = db.query(models.PrrLimit).filter(models.PrrLimit.id == prr_limit_id).first()
    if db_prr_limit is None:
        raise HTTPException(status_code=404, detail="PrrLimit not found")
    return db_prr_limit


@router.put("/{prr_limit_id}", response_model=schemas.PrrLimit)
def update_prr_limit(prr_limit_id: int, prr_limit: schemas.PrrLimitUpdate, db: Session = Depends(get_db)):
    db_prr_limit = db.query(models.PrrLimit).filter(models.PrrLimit.id == prr_limit_id).first()
    if db_prr_limit is None:
        raise HTTPException(status_code=404, detail="PrrLimit not found")

    update_payload = prr_limit.dict(exclude_unset=True)
    if "duration_minutes" in update_payload:
        _validate_duration(update_payload["duration_minutes"])

    for key, value in update_payload.items():
        setattr(db_prr_limit, key, value)

    db.commit()
    db.refresh(db_prr_limit)
    return db_prr_limit


@router.delete("/{prr_limit_id}", response_model=schemas.PrrLimit)
def delete_prr_limit(prr_limit_id: int, db: Session = Depends(get_db)):
    db_prr_limit = db.query(models.PrrLimit).filter(models.PrrLimit.id == prr_limit_id).first()
    if db_prr_limit is None:
        raise HTTPException(status_code=404, detail="PrrLimit not found")
    db.delete(db_prr_limit)
    db.commit()
    return db_prr_limit
