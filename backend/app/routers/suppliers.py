from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Response
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from typing import List
from io import BytesIO
from ..db import get_db
from ..models import Supplier, UserSupplier, UserRole, VehicleType, TransportTypeRef, Zone
from ..schemas import (
    Supplier as SupplierSchema,
    SupplierCreate,
    SupplierUpdate,
    SupplierWithZone,
    UserSupplier as UserSupplierSchema,
    UserSupplierCreate,
    SupplierImportResult,
    SupplierImportError,
)
from ..deps import get_current_user
from openpyxl import Workbook, load_workbook

router = APIRouter(prefix="/api/suppliers", tags=["suppliers"])


@router.get("/", response_model=List[SupplierWithZone])
def get_suppliers(db: Session = Depends(get_db)):
    return (
        db.query(Supplier)
        .options(
            joinedload(Supplier.zone),
            joinedload(Supplier.vehicle_types),
            joinedload(Supplier.transport_types),
        )
        .all()
    )


@router.get("/my", response_model=List[SupplierWithZone])
def get_my_suppliers(
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    if current_user.role == UserRole.admin:
        return (
            db.query(Supplier)
            .options(
                joinedload(Supplier.zone),
                joinedload(Supplier.vehicle_types),
                joinedload(Supplier.transport_types),
            )
            .all()
        )

    user_suppliers = db.query(UserSupplier).filter(UserSupplier.user_id == current_user.id).all()
    supplier_ids = [us.supplier_id for us in user_suppliers]
    return (
        db.query(Supplier)
        .options(
            joinedload(Supplier.zone),
            joinedload(Supplier.vehicle_types),
            joinedload(Supplier.transport_types),
        )
        .filter(Supplier.id.in_(supplier_ids))
        .all()
    )


@router.post("/", response_model=SupplierSchema)
def create_supplier(
    supplier: SupplierCreate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    if current_user.role != UserRole.admin:
        raise HTTPException(status_code=403, detail="Недостаточно прав")

    # Связи
    vehicle_types = db.query(VehicleType).filter(VehicleType.id.in_(supplier.vehicle_type_ids)).all()
    if len(vehicle_types) != len(supplier.vehicle_type_ids):
        raise HTTPException(status_code=404, detail="One or more vehicle types not found")
        
    transport_types = db.query(TransportTypeRef).filter(TransportTypeRef.id.in_(supplier.transport_type_ids)).all()
    if len(transport_types) != len(supplier.transport_type_ids):
        raise HTTPException(status_code=404, detail="One or more transport types not found")

    data = supplier.dict()
    data.pop("vehicle_type_ids", None)
    data.pop("transport_type_ids", None)

    db_supplier = Supplier(**data)
    db_supplier.vehicle_types = vehicle_types
    db_supplier.transport_types = transport_types
    db.add(db_supplier)
    db.commit()
    db.refresh(db_supplier)
    return db_supplier


@router.get("/{supplier_id}", response_model=SupplierWithZone)
def get_supplier(supplier_id: int, db: Session = Depends(get_db)):
    supplier = (
        db.query(Supplier)
        .options(
            joinedload(Supplier.zone),
            joinedload(Supplier.vehicle_types),
            joinedload(Supplier.transport_types),
        )
        .filter(Supplier.id == supplier_id)
        .first()
    )
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")
    return supplier


@router.put("/{supplier_id}", response_model=SupplierSchema)
def update_supplier(
    supplier_id: int,
    supplier: SupplierUpdate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    if current_user.role != UserRole.admin:
        raise HTTPException(status_code=403, detail="Недостаточно прав")

    db_supplier = db.query(Supplier).filter(Supplier.id == supplier_id).first()
    if not db_supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")

    # Обновляем связи
    if supplier.vehicle_type_ids is not None:
        vehicle_types = db.query(VehicleType).filter(VehicleType.id.in_(supplier.vehicle_type_ids)).all()
        if len(vehicle_types) != len(supplier.vehicle_type_ids):
            raise HTTPException(status_code=404, detail="One or more vehicle types not found")
        db_supplier.vehicle_types = vehicle_types

    if supplier.transport_type_ids is not None:
        transport_types = db.query(TransportTypeRef).filter(TransportTypeRef.id.in_(supplier.transport_type_ids)).all()
        if len(transport_types) != len(supplier.transport_type_ids):
            raise HTTPException(status_code=404, detail="One or more transport types not found")
        db_supplier.transport_types = transport_types

    # Обновляем остальные поля
    update_data = supplier.dict(exclude_unset=True)
    update_data.pop("vehicle_type_ids", None)
    update_data.pop("transport_type_ids", None)

    for key, value in update_data.items():
        setattr(db_supplier, key, value)

    db.commit()
    db.refresh(db_supplier)
    return db_supplier


@router.delete("/{supplier_id}")
def delete_supplier(
    supplier_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    if current_user.role != UserRole.admin:
        raise HTTPException(status_code=403, detail="Недостаточно прав")

    db_supplier = db.query(Supplier).filter(Supplier.id == supplier_id).first()
    if not db_supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")

    db.delete(db_supplier)
    db.commit()
    return {"message": "Supplier deleted"}


@router.get("/import/template")
def download_import_template(
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """
    Сформировать XLSX-шаблон для импорта поставщиков с актуальными справочниками зон и типов ТС.
    """
    if current_user.role != UserRole.admin:
        raise HTTPException(status_code=403, detail="Недостаточно прав")

    wb = Workbook()
    ws = wb.active
    ws.title = "suppliers"
    ws.append(["name", "zone_name", "vehicle_types", "comment"])
    ws.append(["ООО Пример", "Эрго/решетки/корпус", "Фура 20',Газель", "Комментарий (опционально)"])

    ws_zones = wb.create_sheet("zones")
    ws_zones.append(["zone_name"])
    zones = db.query(Zone).all()
    for z in zones:
        ws_zones.append([z.name])

    ws_vehicle_types = wb.create_sheet("vehicle_types")
    ws_vehicle_types.append(["vehicle_type_name"])
    for vt in db.query(VehicleType).all():
        ws_vehicle_types.append([vt.name])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="supplier_import_template.xlsx"'},
    )


@router.post("/import", response_model=SupplierImportResult)
def import_suppliers(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """
    Импорт поставщиков из XLSX. Валидные строки сохраняются, ошибки возвращаются в ответе.
    """
    if current_user.role != UserRole.admin:
        raise HTTPException(status_code=403, detail="Недостаточно прав")

    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Ожидается Excel файл (.xlsx)")

    try:
        content = BytesIO(file.file.read())
        wb = load_workbook(content)
        ws = wb.active
    except Exception:
        raise HTTPException(status_code=400, detail="Не удалось прочитать Excel файл")

    expected_headers = ["name", "zone_name", "vehicle_types", "comment"]
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    if [h.lower() for h in headers] != expected_headers:
        raise HTTPException(status_code=400, detail=f"Ожидается заголовок: {', '.join(expected_headers)}")

    zone_map = {z.name.strip().lower(): z for z in db.query(Zone).all()}
    vehicle_type_map = {vt.name.strip().lower(): vt for vt in db.query(VehicleType).all()}
    existing_supplier_names = {s.name.strip().lower() for s in db.query(Supplier).all()}

    errors: list[SupplierImportError] = []
    created = 0
    names_in_file: set[str] = set()

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        raw_name, raw_zone_name, raw_vehicle_types, raw_comment = row
        name = (raw_name or "").strip()
        zone_name = (raw_zone_name or "").strip()
        vehicle_types_raw = (raw_vehicle_types or "").strip()
        comment = (raw_comment or "") if raw_comment else None

        if not name and not zone_name and not vehicle_types_raw and not comment:
            continue  # пустая строка

        row_errors = []
        name_key = name.lower()
        if not name:
            row_errors.append("name обязателен")
        elif name_key in existing_supplier_names or name_key in names_in_file:
            row_errors.append("поставщик с таким name уже существует")

        if not zone_name:
            row_errors.append("zone_name обязателен")
        else:
            zone = zone_map.get(zone_name.lower())
            if not zone:
                row_errors.append(f"zone_name '{zone_name}' не найден")

        vehicle_types: list[VehicleType] = []
        if vehicle_types_raw:
            for vt_name in [v.strip() for v in vehicle_types_raw.split(",") if v.strip()]:
                vt = vehicle_type_map.get(vt_name.lower())
                if not vt:
                    row_errors.append(f"vehicle_type '{vt_name}' не найден")
                else:
                    vehicle_types.append(vt)

        if row_errors:
            errors.append(SupplierImportError(row_number=idx, message="; ".join(row_errors)))
            continue

        # Все проверки прошли
        supplier = Supplier(name=name, comment=comment, zone_id=zone_map[zone_name.lower()].id)
        supplier.vehicle_types = vehicle_types
        db.add(supplier)
        names_in_file.add(name_key)
        created += 1

    if created:
        db.commit()
    else:
        db.rollback()

    return SupplierImportResult(created=created, errors=errors)


@router.get("/user/{user_id}", response_model=List[SupplierWithZone])
def get_user_suppliers(user_id: int, db: Session = Depends(get_db)):
    user_suppliers = db.query(UserSupplier).filter(UserSupplier.user_id == user_id).all()
    supplier_ids = [us.supplier_id for us in user_suppliers]
    return (
        db.query(Supplier)
        .options(
            joinedload(Supplier.zone),
            joinedload(Supplier.vehicle_types),
            joinedload(Supplier.transport_types),
        )
        .filter(Supplier.id.in_(supplier_ids))
        .all()
    )


@router.post("/user/{user_id}/supplier/{supplier_id}", response_model=UserSupplierSchema)
def add_user_supplier(
    user_id: int,
    supplier_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    if current_user.id != user_id and current_user.role != UserRole.admin:
        raise HTTPException(status_code=403, detail="Недостаточно прав")

    supplier = db.query(Supplier).filter(Supplier.id == supplier_id).first()
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")

    existing = (
        db.query(UserSupplier)
        .filter(
            UserSupplier.user_id == user_id,
            UserSupplier.supplier_id == supplier_id,
        )
        .first()
    )
    if existing:
        raise HTTPException(status_code=400, detail="Relation already exists")

    user_supplier = UserSupplier(user_id=user_id, supplier_id=supplier_id)
    db.add(user_supplier)
    db.commit()
    db.refresh(user_supplier)
    return user_supplier


@router.delete("/user/{user_id}/supplier/{supplier_id}")
def remove_user_supplier(
    user_id: int,
    supplier_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    if current_user.id != user_id and current_user.role != UserRole.admin:
        raise HTTPException(status_code=403, detail="Недостаточно прав")

    user_supplier = (
        db.query(UserSupplier)
        .filter(
            UserSupplier.user_id == user_id,
            UserSupplier.supplier_id == supplier_id,
        )
        .first()
    )

    if not user_supplier:
        raise HTTPException(status_code=404, detail="Relation not found")

    db.delete(user_supplier)
    db.commit()
    return {"message": "Relation deleted"}
