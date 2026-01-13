from collections import defaultdict
from datetime import date, timedelta
from io import BytesIO
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session, joinedload

from .. import models, schemas
from ..db import get_db
from ..deps import require_admin
from ..quota_utils import get_quota_for_date, resolve_direction, used_volume_by_date

router = APIRouter()


def _validate_overrides(payload: schemas.VolumeQuotaBase, overrides: List[schemas.VolumeQuotaOverrideBase]):
    seen_dates: set[date] = set()
    for ov in overrides:
        if ov.override_date in seen_dates:
            raise HTTPException(status_code=400, detail="Override dates must be unique within a quota")
        seen_dates.add(ov.override_date)
        if ov.override_date.year != payload.year or ov.override_date.month != payload.month:
            raise HTTPException(status_code=400, detail="Override date must be within the same year and month as the quota")
        if ov.override_date.weekday() != payload.day_of_week:
            raise HTTPException(status_code=400, detail="Override date must match the quota weekday")
        if ov.volume <= 0:
            raise HTTPException(status_code=400, detail="Override volume must be greater than 0")


def _ensure_no_overlap(
    db: Session,
    payload: schemas.VolumeQuotaBase,
    direction_enum: models.BookingDirection,
    exclude_id: int | None = None,
):
    requested_types = set(payload.transport_type_ids or [])
    if not requested_types:
        raise HTTPException(status_code=400, detail="transport_type_ids is required")

    query = (
        db.query(models.VolumeQuota)
        .join(models.VolumeQuota.transport_types)
        .options(joinedload(models.VolumeQuota.transport_types))
        .filter(
            models.VolumeQuota.object_id == payload.object_id,
            models.VolumeQuota.direction == direction_enum,
            models.VolumeQuota.year == payload.year,
            models.VolumeQuota.month == payload.month,
            models.VolumeQuota.day_of_week == payload.day_of_week,
        )
    )
    if exclude_id:
        query = query.filter(models.VolumeQuota.id != exclude_id)

    existing = query.all()
    for q in existing:
        existing_types = {t.id for t in q.transport_types}
        if existing_types.intersection(requested_types):
            raise HTTPException(
                status_code=400,
                detail=f"Quota already exists for the selected transport types (conflict with quota #{q.id})",
            )


def _load_transport_types(db: Session, ids: List[int]):
    transport_types = db.query(models.TransportTypeRef).filter(models.TransportTypeRef.id.in_(ids)).all() if ids else []
    if len(transport_types) != len(ids):
        raise HTTPException(status_code=404, detail="One or more transport types not found")
    return transport_types


def _normalize(value: Optional[str]) -> str:
    return (value or "").strip()


def _normalize_lower(value: Optional[str]) -> str:
    return _normalize(value).lower()


def _parse_bool(value) -> bool:
    if value is None:
        return True
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    return _normalize_lower(str(value)) in {"1", "true", "yes", "y", "да", "ok"}


def _parse_direction(value: str) -> models.BookingDirection:
    try:
        return resolve_direction(_normalize_lower(value))
    except Exception as exc:
        raise ValueError("direction must be 'in' or 'out'") from exc


def _parse_day_of_week(value) -> int:
    try:
        num = int(value)
    except Exception as exc:
        raise ValueError("day_of_week must be an integer") from exc
    if 1 <= num <= 7:
        num = num - 1  # convert to 0-based (Mon=0)
    if num < 0 or num > 6:
        raise ValueError("day_of_week must be between 1-7 (Mon-Sun) or 0-6")
    return num


def _parse_transport_types(raw: str, tt_map: dict[str, int]) -> List[int]:
    parts = [_normalize_lower(p) for p in (raw or "").replace(";", ",").split(",") if _normalize_lower(p)]
    seen: set[int] = set()
    result: list[int] = []
    for part in parts:
        tt_id = tt_map.get(part)
        if not tt_id:
            raise ValueError(f"transport type '{part}' not found")
        if tt_id not in seen:
            seen.add(tt_id)
            result.append(tt_id)
    if not result:
        raise ValueError("transport_types is required")
    return result


@router.get("/", response_model=List[schemas.VolumeQuota])
def list_volume_quotas(db: Session = Depends(get_db), _: models.User = Depends(require_admin)):
    quotas = (
        db.query(models.VolumeQuota)
        .options(joinedload(models.VolumeQuota.transport_types), joinedload(models.VolumeQuota.overrides))
        .all()
    )
    return quotas


@router.get("/template")
def download_volume_quota_template(
    db: Session = Depends(get_db),
    _: models.User = Depends(require_admin),
):
    wb = Workbook()
    ws = wb.active
    ws.title = "Quotas"
    headers = ["object", "direction", "year", "month", "day_of_week", "transport_types", "volume", "allow_overbooking"]
    ws.append(headers)
    ws.append(["Example Object", "in", 2026, 1, 1, "Example Transport", 500, False])

    ws_ov = wb.create_sheet("Overrides")
    ws_ov.append(["object", "direction", "date", "transport_types", "volume"])
    ws_ov.append(["Example Object", "in", "2026-01-15", "Example Transport", 300])

    ws_obj = wb.create_sheet("objects")
    ws_obj.append(["object"])
    for obj in db.query(models.Object).all():
        ws_obj.append([obj.name])

    ws_tt = wb.create_sheet("transport_types")
    ws_tt.append(["transport_type"])
    for tt in db.query(models.TransportTypeRef).all():
        ws_tt.append([tt.name])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="volume_quotas_template.xlsx"'},
    )


@router.post("/import", response_model=schemas.VolumeQuotaImportResult)
def import_volume_quotas(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: models.User = Depends(require_admin),
):
    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Only Excel .xlsx/.xlsm files are supported")

    try:
        wb = load_workbook(BytesIO(file.file.read()))
    except Exception:
        raise HTTPException(status_code=400, detail="Failed to read Excel file")

    quotas_sheet = None
    overrides_sheet = None
    for name in wb.sheetnames:
        normalized = name.strip().lower()
        if normalized == "quotas" and not quotas_sheet:
            quotas_sheet = wb[name]
        if normalized == "overrides" and not overrides_sheet:
            overrides_sheet = wb[name]
    if quotas_sheet is None:
        quotas_sheet = wb.active
    if overrides_sheet is None and "Overrides" in wb.sheetnames:
        overrides_sheet = wb["Overrides"]

    expected_quota_headers = ["object", "direction", "year", "month", "day_of_week", "transport_types", "volume", "allow_overbooking"]
    header_row = [str(cell.value).strip() if cell.value is not None else "" for cell in next(quotas_sheet.iter_rows(min_row=1, max_row=1))]
    if [h.lower() for h in header_row] != expected_quota_headers:
        raise HTTPException(status_code=400, detail=f"Invalid 'Quotas' headers. Expected: {', '.join(expected_quota_headers)}")

    expected_override_headers = ["object", "direction", "date", "transport_types", "volume"]
    if overrides_sheet:
        header_row_ov = [str(cell.value).strip() if cell.value is not None else "" for cell in next(overrides_sheet.iter_rows(min_row=1, max_row=1))]
        if [h.lower() for h in header_row_ov] != expected_override_headers:
            raise HTTPException(status_code=400, detail=f"Invalid 'Overrides' headers. Expected: {', '.join(expected_override_headers)}")

    # Reference maps
    object_map = {_normalize_lower(o.name): o.id for o in db.query(models.Object).all()}
    tt_map = {_normalize_lower(t.name): t.id for t in db.query(models.TransportTypeRef).all()}

    # Existing quotas grouped by (object, direction, year, month, day_of_week)
    existing_quotas = db.query(models.VolumeQuota).options(
        joinedload(models.VolumeQuota.transport_types),
        joinedload(models.VolumeQuota.overrides),
    ).all()
    existing_by_key: dict[tuple, list] = defaultdict(list)
    existing_by_id: dict[int, models.VolumeQuota] = {}
    for q in existing_quotas:
        key = (q.object_id, q.direction, q.year, q.month, q.day_of_week)
        tt_ids = {t.id for t in q.transport_types}
        existing_by_key[key].append((q, tt_ids))
        existing_by_id[q.id] = q

    errors: list[schemas.VolumeQuotaImportError] = []
    pending: dict[tuple, dict] = {}
    base_key_to_pending: dict[tuple, list] = defaultdict(list)
    existing_override_updates: dict[int, dict[date, float]] = defaultdict(dict)

    def add_error(sheet: str, row_number: int, message: str):
        errors.append(schemas.VolumeQuotaImportError(sheet=sheet, row_number=row_number, message=message))

    # Parse base quotas
    for idx, row in enumerate(quotas_sheet.iter_rows(min_row=2, values_only=True), start=2):
        raw_object, raw_dir, raw_year, raw_month, raw_dow, raw_tt, raw_volume, raw_over = row
        obj_name = _normalize(raw_object)
        obj_id = object_map.get(_normalize_lower(obj_name))
        row_errors: list[str] = []
        if not obj_id:
            row_errors.append(f"object '{obj_name}' not found")

        try:
            direction_enum = _parse_direction(raw_dir)
        except Exception as exc:
            row_errors.append(str(exc))
            direction_enum = None

        try:
            year = int(raw_year)
        except Exception:
            row_errors.append("year must be an integer")
            year = None

        try:
            month = int(raw_month)
            if month < 1 or month > 12:
                row_errors.append("month must be between 1 and 12")
        except Exception:
            row_errors.append("month must be an integer")
            month = None

        try:
            day_of_week = _parse_day_of_week(raw_dow)
        except Exception as exc:
            row_errors.append(str(exc))
            day_of_week = None

        try:
            transport_type_ids = _parse_transport_types(raw_tt, tt_map)
        except Exception as exc:
            row_errors.append(str(exc))
            transport_type_ids = []

        try:
            volume = float(raw_volume)
            if volume <= 0:
                row_errors.append("volume must be greater than 0")
        except Exception:
            row_errors.append("volume must be a number")
            volume = None

        allow_overbooking = _parse_bool(raw_over)

        if row_errors:
            add_error("Quotas", idx, "; ".join(row_errors))
            continue

        key = (obj_id, direction_enum, year, month, day_of_week, frozenset(transport_type_ids))
        base_key = (obj_id, direction_enum, year, month, day_of_week)

        if key in pending:
            add_error("Quotas", idx, "Duplicate quota for same object/direction/date/transport types within file")
            continue

        # Check conflicts with pending in the same base key
        for other in base_key_to_pending.get(base_key, []):
            if other["tt_set"].intersection(set(transport_type_ids)):
                add_error("Quotas", idx, "Conflicts with another quota in file for overlapping transport types")
                break
        else:
            # Check conflicts with existing DB quotas
            matching_existing = existing_by_key.get(base_key, [])
            conflict = False
            existing_match_id: int | None = None
            for q, tt_ids in matching_existing:
                if tt_ids == set(transport_type_ids):
                    existing_match_id = q.id
                    continue
                if tt_ids.intersection(transport_type_ids):
                    add_error("Quotas", idx, f"Conflicts with existing quota #{q.id} (overlapping transport types)")
                    conflict = True
                    break
            if conflict:
                continue

            pending[key] = {
                "object_id": obj_id,
                "direction": direction_enum,
                "year": year,
                "month": month,
                "day_of_week": day_of_week,
                "volume": volume,
                "allow_overbooking": allow_overbooking,
                "transport_type_ids": transport_type_ids,
                "tt_set": set(transport_type_ids),
                "kind": "update" if existing_match_id else "create",
                "existing_id": existing_match_id,
                "overrides": {},
            }
            base_key_to_pending[base_key].append(pending[key])

    # Parse overrides (optional sheet)
    if overrides_sheet:
        for idx, row in enumerate(overrides_sheet.iter_rows(min_row=2, values_only=True), start=2):
            raw_object, raw_dir, raw_date, raw_tt, raw_volume = row
            obj_name = _normalize(raw_object)
            obj_id = object_map.get(_normalize_lower(obj_name))
            row_errors: list[str] = []
            if not obj_id:
                row_errors.append(f"object '{obj_name}' not found")

            try:
                direction_enum = _parse_direction(raw_dir)
            except Exception as exc:
                row_errors.append(str(exc))
                direction_enum = None

            try:
                override_date = raw_date if isinstance(raw_date, date) else date.fromisoformat(str(raw_date))
            except Exception:
                row_errors.append("date must be ISO (YYYY-MM-DD)")
                override_date = None

            try:
                transport_type_ids = _parse_transport_types(raw_tt, tt_map)
            except Exception as exc:
                row_errors.append(str(exc))
                transport_type_ids = []

            try:
                volume = float(raw_volume)
                if volume <= 0:
                    row_errors.append("volume must be greater than 0")
            except Exception:
                row_errors.append("volume must be a number")
                volume = None

            if row_errors:
                add_error("Overrides", idx, "; ".join(row_errors))
                continue

            base_key = (obj_id, direction_enum, override_date.year, override_date.month, override_date.weekday())
            target = None

            for candidate in base_key_to_pending.get(base_key, []):
                if candidate["tt_set"] == set(transport_type_ids):
                    target = candidate
                    break

            existing_match: Optional[models.VolumeQuota] = None
            if target is None:
                for q, tt_ids in existing_by_key.get(base_key, []):
                    if tt_ids == set(transport_type_ids):
                        existing_match = q
                        break

            if target is None and existing_match is None:
                add_error("Overrides", idx, "Base quota not found for this override (add it to 'Quotas' sheet first)")
                continue

            # Date consistency check
            if target:
                if override_date.year != target["year"] or override_date.month != target["month"] or override_date.weekday() != target["day_of_week"]:
                    add_error("Overrides", idx, "Override date does not match quota year/month/day_of_week")
                    continue
                if override_date in target["overrides"]:
                    add_error("Overrides", idx, "Duplicate override date for the same quota in file")
                    continue
                target["overrides"][override_date] = volume
            else:
                # Existing quota only touched by overrides
                existing_updates = existing_override_updates[existing_match.id]
                if override_date in existing_updates:
                    add_error("Overrides", idx, "Duplicate override date for the same quota in file")
                    continue
                existing_updates[override_date] = volume

    created = 0
    updated = 0

    # Apply changes for quotas present in file
    for payload in pending.values():
        if payload["kind"] == "create":
            quota = models.VolumeQuota(
                object_id=payload["object_id"],
                direction=payload["direction"],
                year=payload["year"],
                month=payload["month"],
                day_of_week=payload["day_of_week"],
                volume=payload["volume"],
                allow_overbooking=payload["allow_overbooking"],
            )
            quota.transport_types = _load_transport_types(db, payload["transport_type_ids"])
            quota.overrides = [models.VolumeQuotaOverride(override_date=d, volume=v) for d, v in payload["overrides"].items()]
            db.add(quota)
            created += 1
        else:
            quota = existing_by_id.get(payload["existing_id"])
            if not quota:
                add_error("Quotas", 0, f"Quota #{payload['existing_id']} not found during update")
                continue
            quota.object_id = payload["object_id"]
            quota.direction = payload["direction"]
            quota.year = payload["year"]
            quota.month = payload["month"]
            quota.day_of_week = payload["day_of_week"]
            quota.volume = payload["volume"]
            quota.allow_overbooking = payload["allow_overbooking"]
            quota.transport_types = _load_transport_types(db, payload["transport_type_ids"])
            quota.overrides.clear()
            for d, v in payload["overrides"].items():
                quota.overrides.append(models.VolumeQuotaOverride(override_date=d, volume=v))
            updated += 1

    # Apply overrides-only updates for existing quotas not touched in base sheet
    for quota_id, overrides in existing_override_updates.items():
        quota = existing_by_id.get(quota_id)
        if not quota:
            continue
        # Replace overrides for the given dates; keep others
        existing_map = {ov.override_date: ov for ov in list(quota.overrides)}
        for d, v in overrides.items():
            if d in existing_map:
                existing_map[d].volume = v
            else:
                quota.overrides.append(models.VolumeQuotaOverride(override_date=d, volume=v))
        updated += 1

    db.commit()

    return schemas.VolumeQuotaImportResult(
        created=created,
        updated=updated,
        errors=errors,
    )


@router.post("/", response_model=schemas.VolumeQuota)
def create_volume_quota(
    payload: schemas.VolumeQuotaCreate,
    db: Session = Depends(get_db),
    _: models.User = Depends(require_admin),
):
    try:
        direction_enum = resolve_direction(payload.direction)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    if payload.month < 1 or payload.month > 12:
        raise HTTPException(status_code=400, detail="month must be between 1 and 12")
    if payload.day_of_week < 0 or payload.day_of_week > 6:
        raise HTTPException(status_code=400, detail="day_of_week must be between 0 (Mon) and 6 (Sun)")
    if payload.volume <= 0:
        raise HTTPException(status_code=400, detail="volume must be greater than 0")

    _validate_overrides(payload, payload.overrides)
    _ensure_no_overlap(db, payload, direction_enum)

    quota = models.VolumeQuota(
        object_id=payload.object_id,
        direction=direction_enum,
        year=payload.year,
        month=payload.month,
        day_of_week=payload.day_of_week,
        volume=payload.volume,
        allow_overbooking=payload.allow_overbooking,
    )
    quota.transport_types = _load_transport_types(db, payload.transport_type_ids)
    quota.overrides = [models.VolumeQuotaOverride(override_date=ov.override_date, volume=ov.volume) for ov in payload.overrides]

    db.add(quota)
    db.commit()
    db.refresh(quota)
    return quota


@router.put("/{quota_id}", response_model=schemas.VolumeQuota)
def update_volume_quota(
    quota_id: int,
    payload: schemas.VolumeQuotaUpdate,
    db: Session = Depends(get_db),
    _: models.User = Depends(require_admin),
):
    quota = (
        db.query(models.VolumeQuota)
        .options(joinedload(models.VolumeQuota.transport_types), joinedload(models.VolumeQuota.overrides))
        .filter(models.VolumeQuota.id == quota_id)
        .first()
    )
    if not quota:
        raise HTTPException(status_code=404, detail="Quota not found")

    try:
        direction_enum = resolve_direction(payload.direction)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    if payload.month < 1 or payload.month > 12:
        raise HTTPException(status_code=400, detail="month must be between 1 and 12")
    if payload.day_of_week < 0 or payload.day_of_week > 6:
        raise HTTPException(status_code=400, detail="day_of_week must be between 0 (Mon) and 6 (Sun)")
    if payload.volume <= 0:
        raise HTTPException(status_code=400, detail="volume must be greater than 0")

    _validate_overrides(payload, payload.overrides)
    _ensure_no_overlap(db, payload, direction_enum, exclude_id=quota_id)

    quota.object_id = payload.object_id
    quota.direction = direction_enum
    quota.year = payload.year
    quota.month = payload.month
    quota.day_of_week = payload.day_of_week
    quota.volume = payload.volume
    quota.allow_overbooking = payload.allow_overbooking
    quota.transport_types = _load_transport_types(db, payload.transport_type_ids)

    quota.overrides.clear()
    for ov in payload.overrides:
        quota.overrides.append(models.VolumeQuotaOverride(override_date=ov.override_date, volume=ov.volume))

    db.commit()
    db.refresh(quota)
    return quota


@router.delete("/{quota_id}")
def delete_volume_quota(
    quota_id: int,
    db: Session = Depends(get_db),
    _: models.User = Depends(require_admin),
):
    quota = db.query(models.VolumeQuota).filter(models.VolumeQuota.id == quota_id).first()
    if not quota:
        raise HTTPException(status_code=404, detail="Quota not found")
    db.delete(quota)
    db.commit()
    return {"message": "Deleted"}


@router.get("/availability", response_model=List[schemas.VolumeQuotaAvailability])
def quota_availability(
    object_id: int,
    transport_type_id: int,
    direction: str = Query(..., description="in or out"),
    from_date: date = Query(...),
    to_date: date = Query(...),
    db: Session = Depends(get_db),
):
    if from_date > to_date:
        raise HTTPException(status_code=400, detail="from_date must be before to_date")

    try:
        direction_enum = resolve_direction(direction)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    quotas = (
        db.query(models.VolumeQuota)
        .join(models.VolumeQuota.transport_types)
        .options(joinedload(models.VolumeQuota.overrides))
        .filter(
            models.VolumeQuota.object_id == object_id,
            models.VolumeQuota.direction == direction_enum,
            models.TransportTypeRef.id == transport_type_id,
        )
        .all()
    )
    quota_map = {(q.year, q.month, q.day_of_week): q for q in quotas}
    usage_map = used_volume_by_date(db, object_id, transport_type_id, from_date, to_date, direction_enum)

    results: list[schemas.VolumeQuotaAvailability] = []
    current = from_date
    while current <= to_date:
        key = (current.year, current.month, current.weekday())
        quota = quota_map.get(key)
        used_volume = usage_map.get(current, 0.0)

        if quota:
            override = next((ov for ov in quota.overrides if ov.override_date == current), None)
            total_volume = override.volume if override else quota.volume
            remaining_volume = total_volume - used_volume
            results.append(
                schemas.VolumeQuotaAvailability(
                    date=current,
                    total_volume=float(total_volume),
                    used_volume=float(used_volume),
                    remaining_volume=float(remaining_volume),
                    allow_overbooking=quota.allow_overbooking,
                    quota_id=quota.id,
                )
            )
        else:
            results.append(
                schemas.VolumeQuotaAvailability(
                    date=current,
                    total_volume=None,
                    used_volume=float(used_volume),
                    remaining_volume=float(0.0),
                    allow_overbooking=None,
                    quota_id=None,
                )
            )

        current += timedelta(days=1)

    return results
