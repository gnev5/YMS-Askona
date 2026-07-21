import pytest
from fastapi.testclient import TestClient
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from datetime import date, time, datetime
from io import BytesIO

from openpyxl import load_workbook

from app.db import Base, get_db
from app import models, schemas # Keep models and schemas imported at module level
from app.deps import get_current_user




# Настройка тестовой базы данных
SQLALCHEMY_DATABASE_URL = "sqlite:///:memory:"

@pytest.fixture(scope="session")
def db_engine():
    """Returns an engine for the test database."""
    return create_engine(SQLALCHEMY_DATABASE_URL, connect_args={"check_same_thread": False})


@pytest.fixture(scope="session")
def setup_db(db_engine):
    """Creates the database schema once per session."""
    Base.metadata.create_all(bind=db_engine)
    yield
    Base.metadata.drop_all(bind=db_engine)


@pytest.fixture(scope="function")
def db_session(db_engine, setup_db):
    """Provides a transactional scope for tests."""
    connection = db_engine.connect()
    transaction = connection.begin()
    session = sessionmaker(autocommit=False, autoflush=False, bind=connection)()
    
    yield session
    
    session.close()
    transaction.rollback()
    connection.close()


@pytest.fixture(scope="function")
def test_user_fixture(db_session):
    test_user = models.User(email="test@user.com", password_hash="hash", full_name="Test User")
    db_session.add(test_user)
    db_session.commit()
    db_session.refresh(test_user)
    return test_user


@pytest.fixture(scope="function")
def test_client(db_session, test_user_fixture):
    """Фикстура для создания тестового клиента с очищенной БД и переопределенными зависимостями"""
    # app.main creates tables on import using app.db.engine; point it at the test DB.
    import app.db as app_db
    app_db.engine = db_session.get_bind()

    from app.main import app

    app.dependency_overrides[get_db] = lambda: db_session
    app.dependency_overrides[get_current_user] = lambda: test_user_fixture
    with TestClient(app) as client:
        yield client
    app.dependency_overrides.pop(get_db)
    app.dependency_overrides.pop(get_current_user)






def test_create_booking_capacity_limit(test_client, db_session, test_user_fixture):

    client = test_client

    # 1. Создаем тестовые данные

    test_object = models.Object(name="Test Object", object_type="warehouse", capacity_in=1, capacity_out=1)

    db_session.add(test_object)

    db_session.commit()



    test_dock = models.Dock(name="Test Dock", dock_type="entrance", object_id=test_object.id)

    db_session.add(test_dock)

    db_session.commit()



    test_vehicle_type = models.VehicleType(name="Test Vehicle", duration_minutes=30)

    db_session.add(test_vehicle_type)

    db_session.commit()



    test_slot = models.TimeSlot(

        dock_id=test_dock.id,

        slot_date=date.today(),

        start_time=time(10, 0),

        end_time=time(10, 30),

        capacity=1,

    )

    db_session.add(test_slot)

    db_session.commit()



    # 2. Создаем первое бронирование (должно пройти успешно)

    booking_data_1 = {

        "vehicle_type_id": test_vehicle_type.id,

        "booking_date": str(date.today()),

        "start_time": "10:00",

        "object_id": test_object.id,

        "vehicle_plate": "123",

        "driver_full_name": "driver",

        "driver_phone": "12345",

    }

    response1 = client.post("/api/bookings/", json=booking_data_1)

    assert response1.status_code == 200



    # 3. Создаем второе бронирование на тот же слот (должно провалиться из-за лимита)

    booking_data_2 = {

        "vehicle_type_id": test_vehicle_type.id,

        "booking_date": str(date.today()),

        "start_time": "10:00",

        "object_id": test_object.id,

        "vehicle_plate": "456",

        "driver_full_name": "driver2",

        "driver_phone": "54321",

    }

    response2 = client.post("/api/bookings/", json=booking_data_2)

    assert response2.status_code == 409

    assert "не найдено доступных временных слотов" in response2.json()["detail"].lower()


def test_create_booking_on_exit_dock_fails(test_client, db_session):
    """
    Тест проверяет, что нельзя создать бронь на ВХОД для дока с типом "Выход".
    Ожидаемый результат: Ошибка 409, так как доступных слотов найдено не будет.
    """
    # 1. Arrange: Создаем тестовые данные
    test_object = models.Object(name="Test Object", object_type="warehouse")
    db_session.add(test_object)
    db_session.commit()

    exit_dock = models.Dock(name="Exit Dock", dock_type="exit", object_id=test_object.id)
    db_session.add(exit_dock)
    db_session.commit()

    test_vehicle_type = models.VehicleType(name="Test Vehicle", duration_minutes=30)
    db_session.add(test_vehicle_type)
    db_session.commit()

    test_slot = models.TimeSlot(
        dock_id=exit_dock.id,
        slot_date=date.today(),
        start_time=time(14, 0),
        end_time=time(14, 30),
        capacity=1,
    )
    db_session.add(test_slot)
    db_session.commit()

    # 2. Act: Пытаемся создать бронирование на этот слот
    booking_data = {
        "vehicle_type_id": test_vehicle_type.id,
        "booking_date": str(date.today()),
        "start_time": "14:00",
        "object_id": test_object.id, # Важно указать объект
        "vehicle_plate": "EXIT-TEST",
        "driver_full_name": "Test Driver",
        "driver_phone": "1234567890",
    }

    response = test_client.post("/api/bookings/", json=booking_data)

    # 3. Assert: Проверяем, что получили ошибку
    assert response.status_code == 409
    assert "не найдено доступных временных слотов" in response.json()["detail"].lower()





def test_create_booking_does_not_fallback_to_dock_from_other_supplier_zone(test_client, db_session, test_user_fixture):
    zone_a = models.Zone(name="Zone A")
    zone_b = models.Zone(name="Zone B")
    db_session.add_all([zone_a, zone_b])
    db_session.commit()

    supplier = models.Supplier(name="Supplier A", zone_id=zone_a.id)
    db_session.add(supplier)
    db_session.commit()

    test_object = models.Object(name="Zone Test Object", object_type="warehouse")
    db_session.add(test_object)
    db_session.commit()

    dock_zone_a = models.Dock(name="Dock Zone A", dock_type="entrance", object_id=test_object.id)
    dock_zone_b = models.Dock(name="Dock Zone B", dock_type="entrance", object_id=test_object.id)
    dock_zone_a.available_zones.append(zone_a)
    dock_zone_b.available_zones.append(zone_b)
    db_session.add_all([dock_zone_a, dock_zone_b])
    db_session.commit()

    test_vehicle_type = models.VehicleType(name="Zone Vehicle", duration_minutes=30)
    db_session.add(test_vehicle_type)
    db_session.commit()

    slot_zone_a = models.TimeSlot(
        dock_id=dock_zone_a.id,
        slot_date=date.today(),
        start_time=time(11, 0),
        end_time=time(11, 30),
        capacity=1,
    )
    slot_zone_b = models.TimeSlot(
        dock_id=dock_zone_b.id,
        slot_date=date.today(),
        start_time=time(11, 0),
        end_time=time(11, 30),
        capacity=1,
    )
    db_session.add_all([slot_zone_a, slot_zone_b])
    db_session.commit()

    existing_booking = models.Booking(
        user_id=test_user_fixture.id,
        vehicle_type_id=test_vehicle_type.id,
        status="confirmed",
        booking_type=models.BookingDirection.inbound,
    )
    db_session.add(existing_booking)
    db_session.flush()
    db_session.add(models.BookingTimeSlot(booking_id=existing_booking.id, time_slot_id=slot_zone_a.id))
    db_session.commit()

    booking_data = {
        "vehicle_type_id": test_vehicle_type.id,
        "booking_date": str(date.today()),
        "start_time": "11:00",
        "object_id": test_object.id,
        "supplier_id": supplier.id,
        "zone_id": zone_a.id,
        "vehicle_plate": "ZONE-001",
        "driver_full_name": "Zone Driver",
        "driver_phone": "70000000000",
        "time_slot_id": slot_zone_a.id,
        "booking_type": "in",
    }

    response = test_client.post("/api/bookings/", json=booking_data)

    assert response.status_code == 409
    assert "не найдено доступных временных слотов" in response.json()["detail"].lower()


def _create_basic_booking(client, db_session, *, slot_time=time(10, 0), vehicle_plate="CANCEL-1"):
    test_object = models.Object(name=f"Object {vehicle_plate}", object_type="warehouse", capacity_in=1, capacity_out=1)
    db_session.add(test_object)
    db_session.commit()

    test_dock = models.Dock(name=f"Dock {vehicle_plate}", dock_type="entrance", object_id=test_object.id)
    db_session.add(test_dock)
    db_session.commit()

    test_vehicle_type = models.VehicleType(name=f"Vehicle {vehicle_plate}", duration_minutes=30)
    db_session.add(test_vehicle_type)
    db_session.commit()

    test_slot = models.TimeSlot(
        dock_id=test_dock.id,
        slot_date=date.today(),
        start_time=slot_time,
        end_time=time(slot_time.hour, slot_time.minute + 30),
        capacity=1,
    )
    db_session.add(test_slot)
    db_session.commit()

    response = client.post("/api/bookings/", json={
        "vehicle_type_id": test_vehicle_type.id,
        "booking_date": str(date.today()),
        "start_time": slot_time.strftime("%H:%M"),
        "object_id": test_object.id,
        "vehicle_plate": vehicle_plate,
        "driver_full_name": "Cancel Driver",
        "driver_phone": "12345",
    })
    assert response.status_code == 200, response.text
    return response.json(), test_object, test_slot, test_vehicle_type


def test_cancelled_booking_remains_in_my_bookings_and_slot_can_be_rebooked(test_client, db_session):
    booking, test_object, _test_slot, test_vehicle_type = _create_basic_booking(test_client, db_session)

    cancel_response = test_client.put(f"/api/bookings/{booking['id']}/cancel")
    assert cancel_response.status_code == 200

    my_response = test_client.get("/api/bookings/my")
    assert my_response.status_code == 200
    returned = {item["id"]: item for item in my_response.json()}
    assert returned[booking["id"]]["status"] == "cancelled"
    assert returned[booking["id"]]["vehicle_plate"] == "CANCEL-1"

    # Отмененная запись остается в истории, но ее слот не должен блокировать новую бронь.
    rebook_response = test_client.post("/api/bookings/", json={
        "vehicle_type_id": test_vehicle_type.id,
        "booking_date": str(date.today()),
        "start_time": "10:00",
        "object_id": test_object.id,
        "vehicle_plate": "CANCEL-2",
        "driver_full_name": "Rebook Driver",
        "driver_phone": "67890",
    })
    assert rebook_response.status_code == 200, rebook_response.text


def _create_export_report_booking(
    db_session,
    user,
    *,
    supplier_name: str,
    transport_type_name: str,
    transport_enum: models.TransportType,
    slot_date: date,
    start: time,
    end: time,
    cubes: float | None,
):
    suffix = f"{supplier_name}-{transport_type_name}-{start.strftime('%H%M')}"[:60]
    obj = models.Object(name=f"Object {suffix}", object_type="warehouse", capacity_in=10, capacity_out=10)
    zone = models.Zone(name=f"Zone {suffix}")
    vehicle_type = models.VehicleType(name=f"Vehicle {suffix}", duration_minutes=30)
    transport_type = db_session.query(models.TransportTypeRef).filter(
        models.TransportTypeRef.name == transport_type_name
    ).first()
    if transport_type is None:
        transport_type = models.TransportTypeRef(name=transport_type_name, enum_value=transport_enum)
        db_session.add(transport_type)
    db_session.add_all([obj, zone, vehicle_type])
    db_session.commit()

    supplier = models.Supplier(name=supplier_name, zone_id=zone.id)
    db_session.add(supplier)
    db_session.commit()

    dock = models.Dock(name=f"Dock {suffix}", dock_type="entrance", object_id=obj.id)
    db_session.add(dock)
    db_session.commit()

    slot = models.TimeSlot(dock_id=dock.id, slot_date=slot_date, start_time=start, end_time=end, capacity=2)
    db_session.add(slot)
    db_session.commit()

    booking = models.Booking(
        user_id=user.id,
        vehicle_type_id=vehicle_type.id,
        vehicle_plate=f"PLATE-{start.strftime('%H%M')}",
        driver_full_name="Report Driver",
        driver_phone="12345",
        supplier_id=supplier.id,
        zone_id=zone.id,
        transport_type_id=transport_type.id,
        cubes=cubes,
        transport_sheet=f"TL-{start.strftime('%H%M')}",
        booking_type=models.BookingDirection.inbound,
        status="confirmed",
        created_at=datetime(2026, 1, 1, 10, 0),
    )
    db_session.add(booking)
    db_session.commit()
    db_session.add(models.BookingTimeSlot(booking_id=booking.id, time_slot_id=slot.id))
    db_session.commit()
    db_session.refresh(booking)
    return booking


def test_bookings_export_adds_production_and_purchased_summary_sheets(test_client, db_session, test_user_fixture):
    base_date = date(2026, 1, 10)
    bookings = [
        _create_export_report_booking(
            db_session,
            test_user_fixture,
            supplier_name="Литвуд Лопатина 5",
            transport_type_name="собственное производство",
            transport_enum=models.TransportType.own_production,
            slot_date=base_date,
            start=time(21, 0),
            end=time(21, 30),
            cubes=1,
        ),
        _create_export_report_booking(
            db_session,
            test_user_fixture,
            supplier_name="Литвуд Лопатина 9",
            transport_type_name="собственное производство",
            transport_enum=models.TransportType.own_production,
            slot_date=base_date,
            start=time(22, 30),
            end=time(23, 0),
            cubes=1,
        ),
        _create_export_report_booking(
            db_session,
            test_user_fixture,
            supplier_name="Почайка паллеты",
            transport_type_name="собственное производство",
            transport_enum=models.TransportType.own_production,
            slot_date=base_date,
            start=time(20, 0),
            end=time(20, 30),
            cubes=1,
        ),
        _create_export_report_booking(
            db_session,
            test_user_fixture,
            supplier_name="Софт Слип (Ковров)",
            transport_type_name="собственное производство",
            transport_enum=models.TransportType.own_production,
            slot_date=base_date,
            start=time(23, 0),
            end=time(23, 30),
            cubes=75,
        ),
        _create_export_report_booking(
            db_session,
            test_user_fixture,
            supplier_name="Картон служебный",
            transport_type_name="собственное производство",
            transport_enum=models.TransportType.own_production,
            slot_date=base_date,
            start=time(19, 0),
            end=time(19, 30),
            cubes=999,
        ),
        _create_export_report_booking(
            db_session,
            test_user_fixture,
            supplier_name="Литвуд Лопатина 5",
            transport_type_name="магистраль",
            transport_enum=models.TransportType.own_production,
            slot_date=base_date,
            start=time(17, 0),
            end=time(17, 30),
            cubes=999,
        ),
        _create_export_report_booking(
            db_session,
            test_user_fixture,
            supplier_name="Поставщик А",
            transport_type_name="закупная",
            transport_enum=models.TransportType.purchased,
            slot_date=base_date,
            start=time(18, 0),
            end=time(18, 30),
            cubes=10,
        ),
        _create_export_report_booking(
            db_session,
            test_user_fixture,
            supplier_name="Поставщик Б",
            transport_type_name="закупная",
            transport_enum=models.TransportType.purchased,
            slot_date=base_date,
            start=time(22, 0),
            end=time(22, 30),
            cubes=12,
        ),
    ]

    response = test_client.post("/api/bookings/export/xlsx", json=[booking.id for booking in bookings])
    assert response.status_code == 200, response.text

    wb = load_workbook(BytesIO(response.content))
    assert "Собственное производство" in wb.sheetnames
    assert "Закупная" in wb.sheetnames

    own_ws = wb["Собственное производство"]
    assert [cell.value for cell in own_ws[1]] == ["Направление", "2026-01-10", "2026-01-11", "Общий итог"]
    own_rows = {row[0]: row[1:] for row in own_ws.iter_rows(min_row=2, values_only=True)}
    assert own_rows["Лопатина"] == (38, 38, 76)
    assert own_rows["Почаевский"] == (64, 0, 64)
    assert own_rows["Софт Слип"] == (0, 75, 75)
    assert own_rows["Социалистическая"] == (0, 0, 0)
    assert own_rows["ЦРСГП"] == (0, 0, 0)
    assert own_rows["Общий итог"] == (102, 113, 215)
    assert "Картон служебный" not in own_rows

    purchased_ws = wb["Закупная"]
    assert [cell.value for cell in purchased_ws[1]] == ["Тип перевозки", "2026-01-10", "2026-01-11", "Общий итог"]
    purchased_rows = {row[0]: row[1:] for row in purchased_ws.iter_rows(min_row=2, values_only=True)}
    assert purchased_rows["Закупная"] == (10, 12, 22)



def test_cancelled_booking_export_includes_grey_row(test_client, db_session):
    booking, _test_object, _test_slot, _test_vehicle_type = _create_basic_booking(
        test_client,
        db_session,
        slot_time=time(12, 0),
        vehicle_plate="EXPORT-CANCEL",
    )
    assert test_client.put(f"/api/bookings/{booking['id']}/cancel").status_code == 200

    response = test_client.post("/api/bookings/export/xlsx", json=[booking["id"]])
    assert response.status_code == 200, response.text

    wb = load_workbook(BytesIO(response.content))
    ws = wb["bookings"]
    row = [cell.value for cell in ws[2]]
    assert row[10] == "cancelled"
    assert row[12] == booking["id"]
    assert ws.cell(row=2, column=1).fill.fgColor.rgb == "FFE5E7EB"
