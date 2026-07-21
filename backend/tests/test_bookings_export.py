import pytest
from datetime import date, time, timedelta
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook, Workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.db import Base, get_db
from app import models
from app.deps import get_current_user


SQLALCHEMY_DATABASE_URL = "sqlite:///:memory:"


@pytest.fixture(scope="session")
def db_engine():
    return create_engine(SQLALCHEMY_DATABASE_URL, connect_args={"check_same_thread": False})


@pytest.fixture(scope="session")
def setup_db(db_engine):
    Base.metadata.create_all(bind=db_engine)
    yield
    Base.metadata.drop_all(bind=db_engine)


@pytest.fixture(scope="function")
def db_session(db_engine, setup_db):
    connection = db_engine.connect()
    transaction = connection.begin()
    session = sessionmaker(autocommit=False, autoflush=False, bind=connection)()

    yield session

    session.close()
    transaction.rollback()
    connection.close()


@pytest.fixture(scope="function")
def test_user_fixture(db_session):
    test_user = models.User(email="export@user.com", password_hash="hash", full_name="Export User")
    db_session.add(test_user)
    db_session.commit()
    db_session.refresh(test_user)
    return test_user


@pytest.fixture(scope="function")
def test_client(db_session, test_user_fixture):
    from app.main import app

    app.dependency_overrides[get_db] = lambda: db_session
    app.dependency_overrides[get_current_user] = lambda: test_user_fixture
    with TestClient(app) as client:
        yield client
    app.dependency_overrides.pop(get_db)
    app.dependency_overrides.pop(get_current_user)


def test_export_bookings_start_end_variant(test_client, db_session, test_user_fixture):
    start_date = date(2026, 6, 5)
    end_date = start_date + timedelta(days=1)

    zone = models.Zone(name="Export Zone")
    supplier = models.Supplier(name="Export Supplier", zone=zone)
    transport_type = models.TransportTypeRef(
        name="Export Transport",
        enum_value=models.TransportType.purchased,
    )
    vehicle_type = models.VehicleType(name="Export Vehicle", duration_minutes=60)
    test_object = models.Object(name="Export Object", object_type="warehouse")
    dock = models.Dock(name="Export Dock", dock_type="entrance", object=test_object)

    slot_start = models.TimeSlot(
        dock=dock,
        slot_date=start_date,
        start_time=time(23, 0),
        end_time=time(23, 30),
        capacity=1,
    )
    slot_end = models.TimeSlot(
        dock=dock,
        slot_date=end_date,
        start_time=time(0, 0),
        end_time=time(0, 30),
        capacity=1,
    )

    booking = models.Booking(
        user_id=test_user_fixture.id,
        vehicle_type=vehicle_type,
        supplier=supplier,
        zone=zone,
        transport_type=transport_type,
        vehicle_plate="A123BC77",
        driver_full_name="Export Driver",
        driver_phone="79990000000",
        cubes=12.5,
        transport_sheet="TL-001",
        status="confirmed",
        booking_type=models.BookingDirection.inbound,
    )

    db_session.add_all([
        zone,
        supplier,
        transport_type,
        vehicle_type,
        test_object,
        dock,
        slot_start,
        slot_end,
        booking,
    ])
    db_session.commit()

    db_session.add_all([
        models.BookingTimeSlot(booking_id=booking.id, time_slot_id=slot_start.id),
        models.BookingTimeSlot(booking_id=booking.id, time_slot_id=slot_end.id),
    ])
    db_session.commit()

    response = test_client.post("/api/bookings/export/xlsx?variant=start-end", json=[booking.id])

    assert response.status_code == 200
    assert "my_bookings_start_end_export_" in response.headers["content-disposition"]

    workbook = load_workbook(BytesIO(response.content))
    sheet = workbook["bookings"]

    headers = [sheet.cell(row=1, column=idx).value for idx in range(1, 12)]
    assert headers == [
        "\u0414\u0430\u0442\u0430 \u043d\u0430\u0447\u0430\u043b\u0430",
        "\u0412\u0440\u0435\u043c\u044f \u043d\u0430\u0447\u0430\u043b\u0430",
        "\u0414\u0430\u0442\u0430 \u043e\u043a\u043e\u043d\u0447\u0430\u043d\u0438\u044f",
        "\u0412\u0440\u0435\u043c\u044f \u043e\u043a\u043e\u043d\u0447\u0430\u043d\u0438\u044f",
        "\u0422\u0440\u0430\u043d\u0441\u043f\u043e\u0440\u0442\u043d\u044b\u0439 \u043b\u0438\u0441\u0442",
        "\u041f\u043e\u0441\u0442\u0430\u0432\u0449\u0438\u043a",
        "\u041a\u0443\u0431\u044b",
        "\u0422\u0438\u043f \u0422\u0421",
        "\u041e\u0431\u044a\u0435\u043a\u0442",
        "\u0417\u043e\u043d\u0430",
        "\u0422\u0438\u043f \u043f\u0435\u0440\u0435\u0432\u043e\u0437\u043a\u0438",
    ]

    values = [sheet.cell(row=2, column=idx).value for idx in range(1, 12)]
    assert values == [
        "05.06.2026",
        "23:00",
        "06.06.2026",
        "00:30",
        "TL-001",
        "Export Supplier",
        12.5,
        "Export Vehicle",
        "Export Object",
        "Export Zone",
        "Export Transport",
    ]


def test_bookings_export_adds_production_report_matrix_sheet():
    from app.routers.bookings import _append_bookings_report_sheets

    wb = Workbook()
    wb.active.title = "bookings"
    _append_bookings_report_sheets(wb, [
        {
            "booking_date": "2026-01-10",
            "start_time": "21:00:00",
            "supplier_name": "Литвуд Лопатина 5",
            "transport_type_name": "собственное производство",
            "cubes": 1,
        },
        {
            "booking_date": "2026-01-10",
            "start_time": "22:30:00",
            "supplier_name": "Литвуд Лопатина 9",
            "transport_type_name": "собственное производство",
            "cubes": 1,
        },
        {
            "booking_date": "2026-01-10",
            "start_time": "20:00:00",
            "supplier_name": "Почайка паллеты",
            "transport_type_name": "собственное производство",
            "cubes": 1,
        },
        {
            "booking_date": "2026-01-10",
            "start_time": "23:00:00",
            "supplier_name": "Софт Слип (Ковров)",
            "transport_type_name": "собственное производство",
            "cubes": 75,
        },
        {
            "booking_date": "2026-01-10",
            "start_time": "19:00:00",
            "supplier_name": "Картон служебный",
            "transport_type_name": "собственное производство",
            "cubes": 999,
        },
        {
            "booking_date": "2026-01-10",
            "start_time": "17:00:00",
            "supplier_name": "Литвуд Лопатина 5",
            "transport_type_name": "магистраль",
            "cubes": 999,
        },
        {
            "booking_date": "2026-01-10",
            "start_time": "18:00:00",
            "supplier_name": "Поставщик А",
            "transport_type_name": "закупная",
            "cubes": 10,
        },
        {
            "booking_date": "2026-01-10",
            "start_time": "22:00:00",
            "supplier_name": "Поставщик Б",
            "transport_type_name": "закупная",
            "cubes": 12,
        },
    ])

    assert "Собственное производство" in wb.sheetnames
    assert "Закупная" not in wb.sheetnames
    own_ws = wb["Собственное производство"]
    assert [cell.value for cell in own_ws[1]] == [
        "Дата",
        "Лопатина",
        "Почаевский",
        "Солвис",
        "Софт Слип",
        "Социалистическая",
        "ЦРСГП",
        "Закупная",
        "Общий итог",
    ]
    own_rows = {row[0]: row[1:] for row in own_ws.iter_rows(min_row=2, values_only=True)}
    assert own_rows["2026-01-10"] == (38, 64, 0, 0, 0, 0, 10, 112)
    assert own_rows["2026-01-11"] == (38, 0, 0, 75, 0, 0, 12, 125)
    assert own_rows["Общий итог"] == (76, 64, 0, 75, 0, 0, 22, 237)
    assert "Картон служебный" not in own_rows
