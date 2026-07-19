from datetime import date, time as dt_time
from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app import models
from app.db import Base, get_db
from app.deps import get_current_user


SQLALCHEMY_DATABASE_URL = "sqlite:///:memory:"


@pytest.fixture(scope="session")
def db_engine():
    return create_engine(SQLALCHEMY_DATABASE_URL, connect_args={"check_same_thread": False})


@pytest.fixture(scope="session")
def setup_db(db_engine):
    Base.metadata.create_all(bind=db_engine)
    yield
    Base.metadata.drop_all(bind=db_engine)


@pytest.fixture(scope="function")
def db_session(db_engine, setup_db):
    connection = db_engine.connect()
    transaction = connection.begin()
    session = sessionmaker(autocommit=False, autoflush=False, bind=connection)()
    yield session
    session.close()
    transaction.rollback()
    connection.close()


@pytest.fixture(scope="function")
def admin_user(db_session):
    user = models.User(
        email="comparison-admin@example.com",
        password_hash="hash",
        full_name="Comparison Admin",
        role=models.UserRole.admin,
    )
    db_session.add(user)
    db_session.commit()
    db_session.refresh(user)
    return user


@pytest.fixture(scope="function")
def test_client(db_session, admin_user):
    import app.db as app_db
    app_db.engine = db_session.get_bind()

    from app.main import app

    app.dependency_overrides[get_db] = lambda: db_session
    app.dependency_overrides[get_current_user] = lambda: admin_user
    with TestClient(app) as client:
        yield client
    app.dependency_overrides.pop(get_db)
    app.dependency_overrides.pop(get_current_user)


def _xlsx_with_tl_numbers(*tl_numbers: str) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.append(["Номер ТЛ", "Комментарий"])
    for number in tl_numbers:
        ws.append([number, "test"])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _xlsx_with_tl_in_column_g() -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.append(["header"])
    ws.append(["skip-before-range", None, None, None, None, None, "TL-SKIP-BEFORE"])
    ws.append(["included-1", None, None, None, None, None, " tl-010 "])
    ws.append(["included-2", None, None, None, None, None, "TL-011"])
    ws.append(["skip-after-range", None, None, None, None, None, "TL-SKIP-AFTER"])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _xlsx_with_time_cell_in_file_data() -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.append(["Плановое время", "Номер ТЛ"])
    ws.append([dt_time(8, 30), "TL-TIME"])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _xlsx_for_snapshot_columns() -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.append(["Дата", "Ворота", "Номер ТЛ", "Лишнее поле", "Смена"])
    ws.append([date(2026, 7, 10), "Ворота 1", "TL-SNAPSHOT", "не сохранять", "Смена А"])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _add_booking(db_session, *, user_id: int, object_id: int, direction: str, tl_number: str, slot_date: date):
    dock_type = models.DockType.entrance if direction == "in" else models.DockType.exit
    dock = models.Dock(name=f"Dock {tl_number}", dock_type=dock_type, object_id=object_id)
    vehicle_type = models.VehicleType(name=f"Truck {tl_number}", duration_minutes=30)
    db_session.add_all([dock, vehicle_type])
    db_session.flush()
    slot = models.TimeSlot(
        dock_id=dock.id,
        slot_date=slot_date,
        start_time=dt_time(10, 0),
        end_time=dt_time(10, 30),
        capacity=1,
    )
    booking = models.Booking(
        user_id=user_id,
        vehicle_type_id=vehicle_type.id,
        transport_sheet=tl_number,
        booking_type=models.BookingDirection(direction),
        status="confirmed",
    )
    db_session.add_all([slot, booking])
    db_session.flush()
    db_session.add(models.BookingTimeSlot(booking_id=booking.id, time_slot_id=slot.id))
    db_session.commit()
    return booking


def test_comparison_run_matches_tl_case_insensitive_and_checks_extended_period(test_client, db_session, admin_user):
    obj = models.Object(name="РЦ Тест", object_type=models.ObjectType.warehouse)
    db_session.add(obj)
    db_session.commit()

    _add_booking(db_session, user_id=admin_user.id, object_id=obj.id, direction="in", tl_number="TL-001", slot_date=date(2026, 7, 10))
    _add_booking(db_session, user_id=admin_user.id, object_id=obj.id, direction="in", tl_number="TL-002", slot_date=date(2026, 7, 12))
    _add_booking(db_session, user_id=admin_user.id, object_id=obj.id, direction="in", tl_number="TL-003", slot_date=date(2026, 7, 8))

    profile_response = test_client.post(
        "/api/data-comparisons/profiles",
        json={
            "name": "РЦ Тест — Вход",
            "object_id": obj.id,
            "direction": "in",
            "tl_column_name": "Номер ТЛ",
            "status_filters": ["confirmed"],
        },
    )
    assert profile_response.status_code == 200
    profile_id = profile_response.json()["id"]

    upload = _xlsx_with_tl_numbers(" tl-001 ", "TL-003", "TL-999", "TL-DUP", "tl-dup")
    response = test_client.post(
        "/api/data-comparisons/runs",
        data={"profile_id": str(profile_id), "date_from": "2026-07-10", "date_to": "2026-07-12"},
        files={"file": ("comparison.xlsx", upload, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["summary"] == {
        "file_rows": 5,
        "unique_file_tl": 4,
        "yms_rows": 2,
        "matched": 1,
        "found_in_yms_extended_period": 1,
        "missing_in_yms": 1,
        "missing_in_file": 1,
        "duplicate_in_file": 2,
        "duplicate_in_yms": 0,
    }

    rows_by_tl = {row["tl_number_normalized"]: row for row in payload["rows"]}
    assert rows_by_tl["TL-001"]["status"] == "matched"
    assert rows_by_tl["TL-003"]["status"] == "found_in_yms_extended_period"
    assert rows_by_tl["TL-999"]["status"] == "missing_in_yms"
    assert rows_by_tl["TL-002"]["status"] == "missing_in_file"
    duplicate_rows = [row for row in payload["rows"] if row["status"] == "duplicate_in_file"]
    assert len(duplicate_rows) == 2

    history = test_client.get("/api/data-comparisons/runs")
    assert history.status_code == 200
    assert history.json()[0]["summary"]["found_in_yms_extended_period"] == 1

    detail = test_client.get(f"/api/data-comparisons/runs/{payload['id']}")
    assert detail.status_code == 200
    assert detail.json()["rows"] == payload["rows"]


def test_profile_can_read_tl_by_excel_column_letter_and_row_range(test_client, db_session, admin_user):
    obj = models.Object(name="РЦ Диапазон", object_type=models.ObjectType.warehouse)
    db_session.add(obj)
    db_session.commit()

    _add_booking(db_session, user_id=admin_user.id, object_id=obj.id, direction="in", tl_number="TL-010", slot_date=date(2026, 7, 10))
    _add_booking(db_session, user_id=admin_user.id, object_id=obj.id, direction="in", tl_number="TL-011", slot_date=date(2026, 7, 10))
    _add_booking(db_session, user_id=admin_user.id, object_id=obj.id, direction="in", tl_number="TL-SKIP-AFTER", slot_date=date(2026, 7, 10))

    profile_response = test_client.post(
        "/api/data-comparisons/profiles",
        json={
            "name": "РЦ Диапазон — Вход",
            "object_id": obj.id,
            "direction": "in",
            "tl_column_letter": "g",
            "file_start_row": 3,
            "file_end_row": 4,
            "status_filters": ["confirmed"],
        },
    )
    assert profile_response.status_code == 200
    profile_payload = profile_response.json()
    assert profile_payload["tl_column_letter"] == "G"
    assert profile_payload["file_start_row"] == 3
    assert profile_payload["file_end_row"] == 4

    response = test_client.post(
        "/api/data-comparisons/runs",
        data={"profile_id": str(profile_payload["id"]), "date_from": "2026-07-10", "date_to": "2026-07-10"},
        files={"file": ("comparison.xlsx", _xlsx_with_tl_in_column_g(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["summary"]["file_rows"] == 2
    assert payload["summary"]["matched"] == 2
    assert payload["summary"]["missing_in_file"] == 1
    assert [row["file_row_number"] for row in payload["rows"] if row["status"] == "matched"] == [3, 4]
    assert {row["tl_number_normalized"] for row in payload["rows"] if row["status"] == "matched"} == {"TL-010", "TL-011"}
    assert all(row["tl_number_normalized"] != "TL-SKIP-BEFORE" for row in payload["rows"])


def test_profile_can_be_updated_for_separate_profiles_section(test_client, db_session):
    first_obj = models.Object(name="РЦ Старый", object_type=models.ObjectType.warehouse)
    second_obj = models.Object(name="РЦ Новый", object_type=models.ObjectType.warehouse)
    db_session.add_all([first_obj, second_obj])
    db_session.commit()

    create_response = test_client.post(
        "/api/data-comparisons/profiles",
        json={
            "name": "Профиль до редактирования",
            "object_id": first_obj.id,
            "direction": "in",
            "tl_column_letter": "g",
            "status_filters": ["confirmed"],
        },
    )
    assert create_response.status_code == 200
    profile_id = create_response.json()["id"]

    update_response = test_client.put(
        f"/api/data-comparisons/profiles/{profile_id}",
        json={
            "name": "Профиль после редактирования",
            "object_id": second_obj.id,
            "direction": "out",
            "tl_column_letter": "h",
            "status_filters": ["confirmed", "completed"],
            "is_active": False,
        },
    )

    assert update_response.status_code == 200
    payload = update_response.json()
    assert payload["name"] == "Профиль после редактирования"
    assert payload["object_id"] == second_obj.id
    assert payload["object_name"] == "РЦ Новый"
    assert payload["direction"] == "out"
    assert payload["tl_column_letter"] == "H"
    assert payload["status_filters"] == ["confirmed", "completed"]
    assert payload["is_active"] is False

    profiles = test_client.get("/api/data-comparisons/profiles")
    assert profiles.status_code == 200
    stored = next(profile for profile in profiles.json() if profile["id"] == profile_id)
    assert stored["name"] == "Профиль после редактирования"


def test_run_row_range_is_launch_parameter_not_profile_setting(test_client, db_session, admin_user):
    obj = models.Object(name="РЦ Запуск", object_type=models.ObjectType.warehouse)
    db_session.add(obj)
    db_session.commit()

    _add_booking(db_session, user_id=admin_user.id, object_id=obj.id, direction="in", tl_number="TL-010", slot_date=date(2026, 7, 10))
    _add_booking(db_session, user_id=admin_user.id, object_id=obj.id, direction="in", tl_number="TL-011", slot_date=date(2026, 7, 10))
    _add_booking(db_session, user_id=admin_user.id, object_id=obj.id, direction="in", tl_number="TL-SKIP-AFTER", slot_date=date(2026, 7, 10))

    profile_response = test_client.post(
        "/api/data-comparisons/profiles",
        json={
            "name": "РЦ Запуск — Вход",
            "object_id": obj.id,
            "direction": "in",
            "tl_column_letter": "g",
            "status_filters": ["confirmed"],
        },
    )
    assert profile_response.status_code == 200
    profile_payload = profile_response.json()
    assert profile_payload["file_start_row"] == 2
    assert profile_payload["file_end_row"] is None

    response = test_client.post(
        "/api/data-comparisons/runs",
        data={
            "profile_id": str(profile_payload["id"]),
            "date_from": "2026-07-10",
            "date_to": "2026-07-10",
            "file_start_row": "3",
            "file_end_row": "4",
        },
        files={"file": ("comparison.xlsx", _xlsx_with_tl_in_column_g(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["summary"]["file_rows"] == 2
    assert payload["summary"]["matched"] == 2
    assert payload["summary"]["missing_in_file"] == 1
    assert [row["file_row_number"] for row in payload["rows"] if row["status"] == "matched"] == [3, 4]
    assert {row["tl_number_normalized"] for row in payload["rows"] if row["status"] == "matched"} == {"TL-010", "TL-011"}


def test_run_persists_file_and_yms_json_with_time_values(test_client, db_session, admin_user):
    obj = models.Object(name="РЦ Время", object_type=models.ObjectType.warehouse)
    db_session.add(obj)
    db_session.commit()

    _add_booking(db_session, user_id=admin_user.id, object_id=obj.id, direction="in", tl_number="TL-TIME", slot_date=date(2026, 7, 10))

    profile_response = test_client.post(
        "/api/data-comparisons/profiles",
        json={
            "name": "РЦ Время — Вход",
            "object_id": obj.id,
            "direction": "in",
            "tl_column_letter": "b",
            "status_filters": ["confirmed"],
        },
    )
    assert profile_response.status_code == 200

    response = test_client.post(
        "/api/data-comparisons/runs",
        data={
            "profile_id": str(profile_response.json()["id"]),
            "date_from": "2026-07-10",
            "date_to": "2026-07-10",
            "file_start_row": "2",
        },
        files={"file": ("comparison.xlsx", _xlsx_with_time_cell_in_file_data(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 200
    payload = response.json()
    matched = next(row for row in payload["rows"] if row["status"] == "matched")
    assert matched["file_data"]["Плановое время"] == "08:30:00"
    assert matched["yms_data"]["start_time"] == "10:00:00"

def test_profile_snapshot_columns_limit_saved_excel_file_data(test_client, db_session, admin_user):
    obj = models.Object(name="РЦ Снимок", object_type=models.ObjectType.warehouse)
    db_session.add(obj)
    db_session.commit()

    _add_booking(db_session, user_id=admin_user.id, object_id=obj.id, direction="in", tl_number="TL-SNAPSHOT", slot_date=date(2026, 7, 10))

    profile_response = test_client.post(
        "/api/data-comparisons/profiles",
        json={
            "name": "РЦ Снимок — Вход",
            "object_id": obj.id,
            "direction": "in",
            "tl_column_letter": "c",
            "status_filters": ["confirmed"],
            "file_settings": {"snapshot_columns": "A:C,E"},
        },
    )
    assert profile_response.status_code == 200
    assert profile_response.json()["file_settings"]["snapshot_columns"] == "A:C,E"

    response = test_client.post(
        "/api/data-comparisons/runs",
        data={
            "profile_id": str(profile_response.json()["id"]),
            "date_from": "2026-07-10",
            "date_to": "2026-07-10",
            "file_start_row": "2",
        },
        files={"file": ("comparison.xlsx", _xlsx_for_snapshot_columns(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 200
    matched = next(row for row in response.json()["rows"] if row["status"] == "matched")
    assert matched["file_data"] == {
        "Дата": "2026-07-10T00:00:00",
        "Ворота": "Ворота 1",
        "Номер ТЛ": "TL-SNAPSHOT",
        "Смена": "Смена А",
    }

